# Import dependencies
import pandas as pd
import openpyxl
import time
import pyeasylib.excellib as excellib



class DataHyperlink:
    
    def __init__(self, 
                 source_sheetname,
                 source_cell,
                 field_to_source_locations,
                 reference_sheetname,
                 header_rows,
                 field_to_data,
                 wb = None):
        
        '''
        Class to create hyperlinks in a sheet, and links it to 
        data from another sheet.
        
        Definition:
            - source    : the place where the hyperlinks are
            - reference : the place where the hyperlinks point to
        
        Parameters:
            - source_sheetname
              the sheetname of the source data, i.e. where to insert 
              the hyperlinks
            
            - field_to_source_locations:
              a dictionary that links the name of the hyperlink, to 
              the location at the source sheet
            
            - reference_sheetname
              the sheetname of the data to be referenced, i.e. where
              to point the hyperlinks to
              
            - field_to_data
              a dictionary of fieldname and the dataframe
              
            - reference_header_rows
              a list of header rows for the reference data
              
            - wb
              openpyxl workbook. If not specified, will be auto-created
              
        '''
        
        # the place of the hyperlink
        self.source_sheetname           = source_sheetname
        self.field_to_source_locations  = field_to_source_locations
        
        # Place of the reference data, i.e. the cell after clicking on the hyperlink
        self.source_cell                = source_cell
        self.reference_sheetname        = reference_sheetname
        self.header_rows                = header_rows
        self.field_to_data              = field_to_data
        
        # workbook -> will be created if not provided
        self.wb                         = wb
        
        # Initialise a new wb if not present.
        self.wb = self._get_wb()
            

    def _get_wb(self):
        
        if getattr(self, "wb") is None:
            print ("creating a new wb...")
            self.wb = openpyxl.Workbook()
        
        return self.wb
    
    def write_reference_data(self):
            
        # Get attributes
        reference_sheetname = self.reference_sheetname
        header_rows         = self.header_rows
        field_to_data       = self.field_to_data
        
        # Get wb and ws
        wb = self._get_wb()
        if reference_sheetname in wb.sheetnames:
            ws = wb[reference_sheetname]
        else:
            ws = wb.create_sheet(reference_sheetname)
        
        # Write header
        row = 1
        for header in header_rows:
            ws.cell(row, 1, header)
            row += 1
        
        # Get a container to save the fieldname loc
        field_to_target_cell = {}
            
        # Write data
        row += 1 # advance  1 empty row
        for fieldname, df in field_to_data.items():
            
            # # Write the field name
            # ws.cell(row, 1, fieldname)
        
            # # Save loc
            # field_to_target_cell[fieldname] = (row, 1)
            
            # write the data, starting from 1 row after field
            data_row = row + 1
            excellib.df_to_worksheet(df, ws, 
                                     index = True, header = True, 
                                     startrow = data_row, startcol = 1)
            
            # Update row
            row += df.shape[0] + 3 # with two empty rows
        
        self.field_to_target_cell = field_to_target_cell        

    def write_source_data(self):
        
        # Get attrs
        field_to_source_locations = self.field_to_source_locations
        source_sheetname     = self.source_sheetname
        field_to_target_cell = self.field_to_target_cell
        reference_sheetname   = self.reference_sheetname
        
        # Get wb and ws
        wb = self._get_wb()
        # ws = wb.create_sheet(source_sheetname)
        ws = wb[source_sheetname]
        
        # Write
        for fieldname, loc in field_to_source_locations.items():
            
            # Write the fieldname
            ws[loc] = fieldname
            
            # Get the target cell
            # target_row, target_col = field_to_target_cell[fieldname]
            # target_col_letter = openpyxl.utils.get_column_letter(target_col)
            # target_cell = f"{target_col_letter}{target_row}"
            target_cell = self.source_cell
            
            # Set the hyperlink
            hyperlink = f"#'{reference_sheetname}'!{target_cell}"            
            ws[loc].hyperlink = hyperlink
            ws[loc].style = 'Hyperlink'

        
        

if __name__ == "__main__":
        
    # Main sheet for this: main_data
    source_sheetname = "source_data"
    field_to_source_locations = {
        "Apple": "G9",
        "Banana": "A7",
        "Coconut": "A11"
        }
    
    
    # Source data to be in: source_data
    reference_sheetname = "reference_data"
    header_rows = ["This is comment row number 1.",

                   ]
    
    field_to_data = {
        "38492.34": pd.DataFrame([["A", 1, 2, 3],
                               ["B", 4, 4, 5],
                               ["C", 5, 3, 2]]),
          
        "Banana": pd.DataFrame([["X", 0.3, 0.3, 0.3],
                                ["Y", 0.4, 0.4, 0.5]]),
        
        "Coconut": pd.DataFrame([["Q", 3],
                                 ["W", 3],
                                 ["E", 3],
                                 ["R", 6],
                                 ["T", 7]])    
        }
    
    
    # Expected output: see expected_output.xlsx

    source_sheetname = "<5100-xx>Investment sub-lead"
    field_to_source_locations = {"<<<link>>>": "E39"}
    source_cell = "A9"
    reference_sheetname = "TB"
    import luna.common as common
    tb_class = common.TBLoader_From_LunaHub(50060, 2023)
    field_to_data = {"<<<link>>>" : tb_class.get_data_by_fy(2023).drop(["L/S (interval)"], axis = 1)}
    wb = openpyxl.load_workbook(r"D:\workspace\luna\personal_workspace\db\funds_test.xlsx")

    
    if False:
        
        self = DataHyperlink(source_sheetname, field_to_source_locations, reference_sheetname, header_rows, field_to_data, wb)
        self.write_reference_data()
        
        self.write_source_data()
        self.wb.save('test.xlsx')
        

    # Try again where the wb is already present
    if True:
        self = DataHyperlink(source_sheetname, source_cell, field_to_source_locations, reference_sheetname, header_rows, field_to_data, wb)
        self.write_reference_data()
        
        self.write_source_data()
        self.wb.save(r"D:\workspace\luna\personal_workspace\db\funds_test.xlsx")

    if True:

        import webbrowser
        webbrowser.open(r"D:\workspace\luna\personal_workspace\db\funds_test.xlsx")

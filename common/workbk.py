import pandas as pd
import numpy as np
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

from copy import copy

class CreditDropdownList:
    ROW = "J2:J1048576" #TODO: Update this to not hardcode?
    
    def __init__(self, filepath):
          self.filepath = filepath
          self.data = None

    def create_dropdown_list(self):

          self.create_wb()
      #     self.data = self.read_wb()
      #     return self.data
    
    def create_wb(self):
            
            if False:
                  df = pd.read_excel(self.filepath)

                  wb = openpyxl.load_workbook(self.filepath)
                  ws = wb.active
      
                  col_letter = openpyxl.utils.get_column_letter(df.shape[1]+1)
      
                  # create new col and format header
                  ws[f'{col_letter}1'] = f'Declare for current FY?'
                  old_cell = ws['A1']
                  new_cell = ws[f'{col_letter}1']
                  new_cell.border = copy(old_cell.border)
                  new_cell.font = copy(old_cell.font)
            
                  dv = DataValidation(type     = "list",
                                    formula1 = '"Yes, No"',
                                    allow_blank =True
                                    )
                  
                  # Only add a last column if df has data
                  if df.shape[0] > 0:    
                        dv.add(f'{col_letter}2:{col_letter}{df.shape[0] + 1}')
                        ws.add_data_validation(dv)
                  else:
                        pass

                  wb.save(self.filepath)
                  wb.close()

            if True:
                  wb = openpyxl.load_workbook(self.filepath)
                  ws = wb.active

                  # Create data validation 
                  
                  _list = ["Yes", "No"]
                  dv = DataValidation(type="list", formula1='"{}"'.format(','.join(_list)))
                  
                  # Error message
                  dv.error ='Your entry is not in the list'
                  dv.errorTitle = 'Invalid Entry'

                  
                  ws.add_data_validation(dv)

                  dv.add(CreditDropdownList.ROW)

                  wb.save(self.filepath) 
    
    def read_wb(self):
            
            file = input("File path with credit quality tags: ")
            data = pd.read_excel(file)
            self.data = data
            
            return self.data
            
if __name__ == "__main__":


      if False:
            
            workbk_fp = r"D:\gohjiawey\Desktop\Form 3\Credit Quality.xlsx"
    
            workbk = CreditDropdownList(workbk_fp)
            processed_workbk = workbk.create_dropdown_list()
            processed_workbk 

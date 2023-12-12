'''
Sample script for Dacia and Jia Wey
'''

# Import standard libs
import os
import datetime
import pandas as pd
import numpy as np
import re
from fuzzywuzzy import fuzz, process
import sys
sys.path.append("D:\gohjiawey\Desktop\Form 3\CODES")


# Import luna package and fsvi package

import luna
import luna.common as common
import luna.fsvi as fsvi

from luna.common.gl import GLProcessor
import calendar

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

from luna.common.workbk import CreditDropdownList

class MASForm1_Generator:
    
    def __init__(self, tb_class, mapper_class,
                 gl_class, fy = 2022):
        
        
        self.tb_class       = tb_class
        #self.aged_ar_class  = aged_ar_class
        self.mapper_class   = mapper_class
        self.fy             = fy
        self.gl_class       = gl_class

        self.main()
        
    def main(self):
        
        self._prepare_output_container()   
        
        self.map_f1_balances()

        self.gl_ageing()
        self.map_div_declared()
        self.map_future_tax_benefits()
        self.map_base_capital()

        self.collect_manual_inputs()
        self.map_unsecured_rlt_corp()
        self.map_other_unsecured_loans()

        self.create_deductions_tempdf()
        self.update_deductions_tempdf()
        
        self.map_financial_resources()

        self.awp_processing()
        self.map_operational_rr()
        self.map_total_operational_rr()
        self.map_total_risk_requirement()
        self.map_fr_trr_ratio()


        self.create_aa_tempdf()
        self.map_aa_df()        
        self.update_fees_receivables_tempdf(aged_receivables_fp)
        self.update_adjusted_asset_tempdf()
        self.map_adjusted_asset()
        self.map_avg_adjusted_asset()
        self.map_aa_threshold()
        self.column_mapper()
        self.output_excel()
        self.output_excel_trr()
        

    def _prepare_output_container(self):

        # make a copy of the input template
        outputdf = self.mapper_class.df_processed.copy()

        # Add an output column
        outputdf["Balance"] = None

        # Save as attr
        self.outputdf = outputdf
    
    def add_bal_to_template_by_varname(self, varname, value):
        
        row = self.mapper_class.varname_to_index.at[varname]

        # Map value to template's Balance column
        self.outputdf.loc[row, "Balance"] = value
                
        # Excel row 
        print(f"The varname ({varname}) is at ExcelRow: {row}")

    def get_f1_balances(self):
        # need to use Form1's output then add varname column (14)
        # extract the varname, amt
        # map to form 2 by varname matching 

        f1_varname_list = ["puc_ord_shares", "puc_pref_share_noncumulative", "puc_reserve_fund", 
                           "puc_unappr_profit_or_loss", "puc_net_head_office_funds", "puc_pref_share_cumulative", 
                           "current_liab_redeemable_pref_share", "noncurrent_liab_redeemable_pref_share", 
                           "puc_rev_reserve", "puc_other_reserves", "noncurrent_asset_goodwill_ia", 
                           "current_asset_other_prepayment", "noncurrent_asset_investment_in_subsi", "total_asset"]
        
        # Read Form 1 Mapping Output
        f1_fp = r"D:\gohjiawey\Desktop\Form 3\f1_map_mg.xlsx"
        
        # Extract var_name & Balance columns
        df = pd.read_excel(f1_fp)
        filtered_df = df[df["var_name"].isin(f1_varname_list)]
        filtered_df = filtered_df[["var_name", "Balance"]]

        return filtered_df

    def map_f1_balances(self):
        filtered_df = self.get_f1_balances().copy()
        # to delete
        print(filtered_df)

        total_balance = 0
        varname = "current_liab_redeemable_pref_share, noncurrent_liab_redeemable_pref_share"

        for index, arow in filtered_df.iterrows():
            name = arow["var_name"]
            if name in ["current_liab_redeemable_pref_share", "noncurrent_liab_redeemable_pref_share"]:
                amt = arow["Balance"] 
                total_balance += amt
                # Map amt to Template 
                self.add_bal_to_template_by_varname(varname, total_balance)  
                # to delete
                print(f" Mapped {name} : {amt}")  
            
            else:
                amt = arow["Balance"]
                # Map amt to Template  
                self.add_bal_to_template_by_varname(name, amt) 
                # to delete
                print(f" Mapped {name} : {amt}")   
        
        print(f" Mapped {varname} : {total_balance}") 


    def _filter_gl(self, month_end):
        gl = self.gl_class.gl.copy()

        gl["Posting Date"] = pd.to_datetime(gl["Posting Date"], 
                                                dayfirst=True)
        gl["Month"] = gl["Posting Date"].dt.month
        filtered_gl = gl[gl["Month"]<=month_end]

        return filtered_gl
    
    def _get_ending_balance(self, filtered_gl):
        self.gl_mvmt = filtered_gl.groupby([
            "GL Account No", 
            "GL Account Name"]).agg({"Amount": "sum", 
                                     "Opening Balance": "first"})
        self.gl_mvmt.rename(columns={"Amount": "GL Movement"}, inplace=True)
        self.gl_mvmt["Ending Balance"] = self.gl_mvmt["GL Movement"] + self.gl_mvmt["Opening Balance"]
        self.gl_mvmt.reset_index(inplace=True)
        return self.gl_mvmt
    
    def _get_gl_tb(self, gl, tb):
        # Get GL with TB L/S codes 
        gl_tb = tb.merge(gl, how="left", left_on="Account No", right_on="GL Account No")

        gl_tb["Ending Balance"] = gl_tb["Ending Balance"].fillna(gl_tb["Value"])

        gl_tb["L/S"] = pd.to_numeric(gl_tb["L/S"], errors='coerce')
        
        return gl_tb

    def gl_ageing(self):
        # get the tb for the current fy
        tb_df = tb_class.get_data_by_fy(fy).copy()

        month_end = fy_end_date.month
        
        first_month = self._filter_gl(month_end-2)
        second_month = self._filter_gl(month_end-1)
        third_month = self._filter_gl(month_end)
        
        first_month = self._get_ending_balance(first_month)
        second_month = self._get_ending_balance(second_month)
        third_month = self._get_ending_balance(third_month)

        first_month = self._get_gl_tb(first_month, tb_df)
        second_month = self._get_gl_tb(second_month, tb_df)
        third_month = self._get_gl_tb(third_month, tb_df)

        return month_end, first_month, second_month, third_month
    
    def map_div_declared(self):
        varname = "upl_div_declared"
        # Dividend GL movement 
        month_end, first_month, second_month, third_month = self.gl_ageing()
        print()
        
        div = third_month[third_month["L/S"]== 6900.4]
        div = div[div["Name"].str.contains("Dividend", case=False)]
        div_declared = div["GL Movement"].sum()

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, div_declared) 
            
        # to delete
        print(f" Mapped {varname} : {div_declared}")    
    
    def map_future_tax_benefits(self):
        varname = "dfr_future_incometax_benefits"
        month_end, first_month, second_month, third_month = self.gl_ageing()
  
        ftb = third_month[third_month["L/S"]== 5850]
        future_tax_benefits = ftb["Value"].sum()

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, future_tax_benefits) 
            
        # to delete
        print(f" Mapped {varname} : {future_tax_benefits}")   
    
    def map_base_capital(self):
    
        # Get relevant amt from Templates
        # Add:
        puc_ord_shares_varname = "puc_ord_shares"
        puc_ord_shares_row = self.mapper_class.varname_to_index.at[puc_ord_shares_varname]

        puc_unappr_profit_or_loss_varname = "puc_unappr_profit_or_loss"
        puc_unappr_profit_or_loss_row = self.mapper_class.varname_to_index.at[puc_unappr_profit_or_loss_varname]
        
        puc_balance = self.outputdf.loc[puc_ord_shares_row:puc_unappr_profit_or_loss_row, "Balance"].sum()

        print(puc_balance)
        print(type(puc_balance))

        # Less:
        upl_div_declared_varname = "upl_div_declared"
        upl_div_declared_row = self.mapper_class.varname_to_index.at[upl_div_declared_varname]

        upl_interim_loss_varname = "upl_interim_loss"
        upl_interim_loss_row = self.mapper_class.varname_to_index.at[upl_interim_loss_varname]

        upl_balance = self.outputdf.loc[upl_div_declared_row:upl_interim_loss_row, "Balance"].sum()

        # Amt (base capital = puc + reserve fund - upl)
        base_capital = puc_balance - upl_balance
        print(base_capital)
        print(type(base_capital))


        # Net head office funds 
        puc_net_head_office_funds_varname = "puc_net_head_office_funds"
        puc_net_head_office_funds_row = self.mapper_class.varname_to_index.at[puc_net_head_office_funds_varname]
        puc_net_head_office_funds = self.outputdf.loc[puc_net_head_office_funds_row, "Balance"]

        # Total Base Capital or Net Head Office Funds 
        total_base_capital_nhof = base_capital + puc_net_head_office_funds
        print(total_base_capital_nhof)
        print(type(total_base_capital_nhof))

        based_capital_varname = "base_capital"
        bc_total_base_capital_varname = "bc_total_base_capital_nhof"
        fr_base_capital_nhof_varname = "fr_base_capital_nhof"

        # Map amt to Template  
        self.add_bal_to_template_by_varname(based_capital_varname, base_capital)                    # Base Capital
        self.add_bal_to_template_by_varname(bc_total_base_capital_varname, total_base_capital_nhof) # Total Base capital or Net Head Office Funds
        self.add_bal_to_template_by_varname(fr_base_capital_nhof_varname, total_base_capital_nhof)  # (1) Base capital or Net Head Office Funds  

        # to delete
        print(f" Mapped {based_capital_varname} : {base_capital}")  
        print(f" Mapped {bc_total_base_capital_varname} : {total_base_capital_nhof}")  
        print(f" Mapped {fr_base_capital_nhof_varname} : {total_base_capital_nhof}")  

    def collect_manual_inputs(self):
        '''
        To run only when the user_inputs parameter is not specified when
        the class is initialised.
        '''
        question_list = [
            "Unsecured amount due from related corporations: $",    
            "Account no for Other unsecured loans and advances: ",  
            "Amount of Off-balance sheet items for last 3 months (separate each amount with a comma): ", 
            "Amount of Payment on behalf for last 3 months (separate each amount with a comma): ", 
            "Name of non_clms:",                                                                    
            "Select which fields should be included for income or expenses not derived from ordinary activities and not expected to recur frequently or regularly (Choose: Interest income, Dividend, Other revenue):"
        ]

        # 0
        # NA
        # 0,0,0
        # 644.46,123.89, 6097.46
        # Asia Corporate Jet Singapore Pte Ltd
        # Other revenue

        varname_list = ["dfr_unsecured_rlt_corporations", 
                        "dfr_other_unsecured_loans", 
                        "aa_off_bs_items", 
                        "payment_on_behalf", 
                        "non_clms", 
                        "non_freq_income_exp"
                        ]

        self.user_inputs = pd.DataFrame({"Question": question_list, 
                                       "Answer": ""}, index = varname_list)
        
        self.user_inputs["Answer"] = self.user_inputs["Question"].apply(input)
        
    
#### (II) Financial Resources 

    def map_unsecured_rlt_corp(self):

        varname = "dfr_unsecured_rlt_corporations"
        
        answer = self.user_inputs.at["dfr_unsecured_rlt_corporations", "Answer"]
        unsecured_rlt_corp = float(answer)

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, unsecured_rlt_corp) 
            
        # to delete
        print(f" Mapped {varname} : {unsecured_rlt_corp}")   

    def map_other_unsecured_loans(self):

        varname = "dfr_other_unsecured_loans"

        tb_df = tb_class.get_data_by_fy(fy).copy()

        answer = self.user_inputs.at["dfr_other_unsecured_loans", "Answer"]

        if answer == "NA":
            other_unsecured_loans = 0.0
            print(f"No other unsecured loans and advances accounts. The amount will be set to 0.00 {other_unsecured_loans}")

        else:
            acc_no = answer.split(",")
            other_unsecured_loans = tb_df[tb_df["Account No"].isin(acc_no)]["Value"].sum()

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, other_unsecured_loans) 
            
        # to delete
        print(f" Mapped {varname} : {other_unsecured_loans}")   


########## 3 months temp df of Deductions from Financial Resources for AA (c) Deductions from Financial Resources 

    def create_deductions_tempdf(self):
        # Deductions from Financial Resources sections
        # Create 3 columns for last 3 months
        dfr_intangible_assets_varname = "noncurrent_asset_goodwill_ia"
        dfr_intangible_assets__row = self.mapper_class.varname_to_index.at[dfr_intangible_assets_varname]

        fr_total_deductions_varname = "fr_total_deductions"
        fr_total_deductions_row = self.mapper_class.varname_to_index.at[fr_total_deductions_varname]
        
        deductions_df = self.outputdf.copy()
        deductions_df = deductions_df.loc[dfr_intangible_assets__row:fr_total_deductions_row]
        # can consider using col name instead of number
        deductions_df = deductions_df.iloc[:, [0,1,2,3,-1]] #Header1-4, Balance column

        month_end = fy_end_date.month

        self.col_name_3 = calendar.month_abbr[month_end] + '-'+ str(fy)[-2:] # e.g. Dec
        deductions_df.rename(columns={"Balance": self.col_name_3}, inplace=True) 
        self.col_name_2 = calendar.month_abbr[month_end-2] + '-'+ str(fy)[-2:]  # e.g.Nov
        self.col_name_1 = calendar.month_abbr[month_end-1] + '-'+ str(fy)[-2:] # e.g.Oct

        col_names = deductions_df.columns.tolist()
        col_names.insert(-1,self.col_name_2)
        col_names.insert(-1,self.col_name_1)
        deductions_df = deductions_df.reindex(columns = col_names)

        self.deductions_df = deductions_df
    
    def _map_tempdf_by_ls(self, acc_list, field_varname, type=None):
        month_end, first_month, second_month, third_month = self.gl_ageing()
        deductions_df = self.deductions_df.copy()
        # print(deductions_df)

        if type == None:
            oct = first_month[first_month["L/S"].isin(acc_list)]["Ending Balance"].sum()
            nov = second_month[second_month["L/S"].isin(acc_list)]["Ending Balance"].sum()
   
            indexrow = self.mapper_class.varname_to_index.at[field_varname]

            deductions_df.at[indexrow, self.col_name_1] = oct
            deductions_df.at[indexrow, self.col_name_2] = nov

        elif type == "Prepayment": 
            prepay = first_month[first_month["L/S"].isin(acc_list)]
            oct = prepay[prepay["Name"].str.contains("(?i)pre")]["Ending Balance"].sum()
            prepay = second_month[second_month["L/S"].isin(acc_list)]
            nov = prepay[prepay["Name"].str.contains("(?i)pre")]["Ending Balance"].sum()

            indexrow = self.mapper_class.varname_to_index.at[field_varname]

            deductions_df.at[indexrow, self.col_name_1] = oct
            deductions_df.at[indexrow, self.col_name_2] = nov

            print(oct, nov)


        elif type == "Cash":

            depo = self.verify_credit_quality()
            
            accounts = depo[depo["Credit Quality Grade 1?"]=="Yes"]["Account No"].to_list()

            # Get ending petty cash / cash in hand balance for 3 months
            cash = first_month[first_month["L/S"].isin(acc_list)]
            oct_cash = cash[cash["Name"].str.contains("(?i)cash")]["Ending Balance"].sum()
            oct_depo = cash[cash["Account No"].isin(accounts)]["Ending Balance"].sum()

            cash = second_month[second_month["L/S"].isin(acc_list)]
            nov_cash = cash[cash["Name"].str.contains("(?i)cash")]["Ending Balance"].sum()
            nov_depo = cash[cash["Account No"].isin(accounts)]["Ending Balance"].sum()

            cash = third_month[third_month["L/S"].isin(acc_list)]
            dec_cash = cash[cash["Name"].str.contains("(?i)cash")]["Ending Balance"].sum()
            dec_depo = cash[cash["Account No"].isin(accounts)]["Ending Balance"].sum()

            cash = [oct_cash,nov_cash,dec_cash]
            depo = [oct_depo,nov_depo,dec_depo]

            return cash,depo

        self.deductions_df = deductions_df
    
    def update_deductions_tempdf(self):
        # Map amounts from the relevant L/S to the correct field for each month
        self._map_tempdf_by_ls([5700.1, 5700.1, 5800.1, 5800.2], "noncurrent_asset_goodwill_ia")
        self._map_tempdf_by_ls([5850], "dfr_future_incometax_benefits")
        self._map_tempdf_by_ls([5400.1,5200.2], "current_asset_other_prepayment", type="Prepayment")
        self._map_tempdf_by_ls([5100.3, 5100.4, 5100.5], "dfr_capinvst_subsidiary_associate")

        # Get column total
        fr_total_deductions_varname = "fr_total_deductions"
        fr_total_deductions_row = self.mapper_class.varname_to_index.at[fr_total_deductions_varname]
        self.deductions_df.iloc[:,-3:] = self.deductions_df.iloc[:,-3:].astype("float")
        total_deductions = self.deductions_df.iloc[:,-3:].sum()

        self.deductions_df.loc[fr_total_deductions_row, self.deductions_df.columns[-3:]] = total_deductions
        
        print(self.deductions_df)
        return self.deductions_df

    def map_financial_resources(self):
        # Base Capital or Net Head Funds
        fr_base_capital_nhof_varname = "fr_base_capital_nhof"
        fr_base_capital_nhof_row  = self.mapper_class.varname_to_index.at[fr_base_capital_nhof_varname]
        fr_base_capital = self.outputdf.loc[fr_base_capital_nhof_row, "Balance"]
        print(fr_base_capital)
        print(type(fr_base_capital))

        # Get Add items (Base Capital or Net Head Funds)
        puc_pref_share_cumulative_varname = "puc_pref_share_cumulative"
        puc_pref_share_cumulative_row = self.mapper_class.varname_to_index.at[puc_pref_share_cumulative_varname]

        bc_cltv_impairment_allowance_varname = "bc_cltv_impairment_allowance"
        bc_cltv_impairment_allowance_row = self.mapper_class.varname_to_index.at[bc_cltv_impairment_allowance_varname]
        bc_sum_balance = self.outputdf.loc[puc_pref_share_cumulative_row:bc_cltv_impairment_allowance_row, "Balance"].sum()
        print(bc_sum_balance)
        print(type(bc_sum_balance))
        
        # Get Less items (deductions from Financial Resources)
        dfr_intangible_assets_varname = "noncurrent_asset_goodwill_ia"
        dfr_intangible_assets_row = self.mapper_class.varname_to_index.at[dfr_intangible_assets_varname]

        dfr_other_assets_nonconvertible_cash_varname = "dfr_other_assets_nonconvertible_cash"
        dfr_other_assets_nonconvertible_cash_row = self.mapper_class.varname_to_index.at[dfr_other_assets_nonconvertible_cash_varname]

        fr_sum_balance = self.outputdf.loc[dfr_intangible_assets_row:dfr_other_assets_nonconvertible_cash_row, "Balance"].sum()
        print(fr_sum_balance)
        print(type(fr_sum_balance))

        # Amt (fr = bc + add items - less items) 
        fr = fr_base_capital + bc_sum_balance - fr_sum_balance 


        fr_total_deductions_varname = "fr_total_deductions"
        varname = "fr_financial_resources_anhof"

        # Map amt to Template  
        self.add_bal_to_template_by_varname(fr_total_deductions_varname, fr_sum_balance) # Total Deductions from Financial Resources
        self.add_bal_to_template_by_varname(varname, fr)                                 # Financial Resources or Adjusted Net Head Office Funds ("FR")
            
        # to delete
        print(f" Mapped {fr_total_deductions_varname} : {fr_sum_balance}")   
        print(f" Mapped {varname} : {fr}")   


#### (III) Total Risk Requirement

    def awp_processing(self):
        # Processing to get the annual gross income table
        awp_fp = r"P:\YEAR 2023\TECHNOLOGY\Technology users\FS Vertical\f2\MG Based capital calculation Dec 2021-1.xlsx"
        awp = pd.read_excel(awp_fp,sheet_name="2. FR+TRR")
        start = awp[awp.iloc[:,0].str.
                         contains("Definition", 
                                  na=False)].index[0]
        end = awp[awp.iloc[:,1].str.
                       contains("Adjusted annual gross income", 
                                na = False)].index[0]

        awp = awp.iloc[start:end+1, 1:5]
        awp.columns = awp.iloc[0]
        awp = awp.iloc[1:]
        awp.reset_index(drop=True, inplace=True)

        self.awp = awp

        return self.awp
    
    def f3_output_processing(self):
        '''
        process from Datahub 
        '''
        # Get F3 output from Datahub 
        f3_fp = r"D:\gohjiawey\Desktop\Form 3\draft_MG - Copy.xlsx"
        f3 = pd.read_excel(f3_fp, sheet_name=1)

        # Combine Headers columns together
        f3["Header 1"] = f3["Header 1"].fillna(f3["Header 2"])
        f3["Header 1"] = f3["Header 1"].fillna(f3["Header 3"])
        f3["Header 1"] = f3["Header 1"].fillna(f3["Header 4"])

        # Pivot years 
        f3_melt = f3.melt(id_vars=['Header 1', 'var_name'], value_vars=[self.fy-1 ,self.fy], var_name="FY", value_name="Amount")
        
        # Filter for 3 years
        years = [self.fy-3, self.fy-2, self.fy-1]
        f3_melt_filtered = f3_melt[f3_melt["FY"].isin(years)]

        # Get template 
        awp_template = pd.read_excel("D:\gohjiawey\Desktop\Form 3\CODES\luna\parameters\mas_forms_tb_mapping.xlsx", sheet_name= "Form 2 - AGI")
        
        # Split var_name to list 
        awp_template = awp_template.rename(columns={"Unnamed: 1": str(self.fy-3), "Unnamed: 2": str(self.fy-2), "Unnamed: 3": str(self.fy-1)})
        awp_template["var_name"] = awp_template["var_name"].apply(lambda x: [i.strip() for i in str(x).split(",")])
        
        # Overwrite var_name for income or expense not derived from ordinary activities based on manual input received
        answer = self.user_inputs.at["non_freq_income_exp", "Answer"]
        answer = answer.split(",")
        varname_list = []
        for i in answer:
            if re.search("(?i)div", i):
                varname_list.append("rev_dividend")
            elif re.search("(?i)other rev", i):
                varname_list.append("rev_other_revenue")
            elif re.search("(?i)interest"):
                varname_list.append("rev_int_others")
        
        required_varname_list = []
        for index, row in awp_template.iterrows():
            name = row['var_name']
            for x in name: 
                if x in varname_list:
                    awp_template.at[index,"var_name"] = varname_list
                if x != 'nan':
                    required_varname_list.append(x)
    
        
        # Filter for relevant var_name
        data = f3_melt_filtered[f3_melt_filtered['var_name'].isin(required_varname_list)]
        # Replace NaN values in the 'Amount' column with 0
        data['Amount'].fillna(0.00, inplace=True)
        # Filter away NaN in 'var_name' column
        awp_template = awp_template[awp_template['var_name'].apply(lambda x: x != ['nan'])]


        '''
        Create a dataframe for each of the 3 years (e.g. 2021,2020,2019).
        filtered_data frame is a list of dataframes. index 0 is empty. index 1 is fy-1, index 2 is fy-2 and so on
        '''
        column_names = data.columns.tolist()
        filtered_data = ["",pd.DataFrame(columns=column_names),pd.DataFrame(columns=column_names),pd.DataFrame(columns=column_names)]

        for index, row in data.iterrows():
            if row['FY'] == self.fy-1:
                filtered_data[1].loc[0] = row
                filtered_data[1].index = filtered_data[1].index - 1
            elif row['FY'] == self.fy-2:
                filtered_data[2].loc[0] = row
                filtered_data[2].index = filtered_data[2].index - 1
            elif row['FY'] == self.fy-3:
                filtered_data[3].loc[-1] = row
                filtered_data[3].index = filtered_data[3].index - 1
        
        for i in range(len(filtered_data)):
            if(i==0):
                continue
            filtered_data[i] = filtered_data[i].sort_index()



        for year_index in range(len(filtered_data)):
            if year_index ==0 :
                continue
            # Map to awp_template (maybe helper function for each self.fy)
            awp_template.loc[:,str(self.fy-year_index)] = 0.00
            for i, row_awp in awp_template.iterrows():
                varname = row_awp["var_name"]
                for index, row in filtered_data[year_index].iterrows():
                    name = row["var_name"]
                    amt = row["Amount"]
                    if name in varname:
                        print(name, varname)
                        awp_template.at[i,str(self.fy-year_index)] += amt
        
        
        # Create Adjusted annual gross income
        awp_template = awp_template.reset_index(drop=True)
        index_aagi = len(awp_template["var_name"])
        awp_template.at[index_aagi,"Annual gross income ="] = "Adjusted Annual gross income"

        for year_index in range(len(filtered_data)):
            if(year_index==0):
                continue    
            rev_total_revenue_amt = 0
            less_amt = 0
            for i, row_awp in awp_template.iterrows():
                if(i==index_aagi):
                    continue
                row_name = row_awp["Annual gross income ="] 
                if(row_name == "- total revenue as per reported in respective year's Form 3 ** (previously Form 6)"):
                    print("HELLO")
                    print(awp_template.loc[i, str(self.fy-year_index)])
                    rev_total_revenue_amt = awp_template.loc[i, str(self.fy-year_index)]
                else:
                    less_amt += awp_template.loc[i, str(self.fy-year_index)]
            adjusted_annual_gross_income = rev_total_revenue_amt - less_amt
            awp_template.loc[index_aagi, str(self.fy-year_index)] = adjusted_annual_gross_income
        
        self.awp = awp_template

    def map_operational_rr(self):

        # Get Average annual gross income from preceding 3 years
        avg_agi = self.awp.iloc[-1, -3:].mean()

        # Separate Average annual gross income to amounts above and below 10 mil
        avg_agi_above_10m = avg_agi-10000000
        if avg_agi_above_10m>0:
            avg_agi_below_10m = 10000000
        else:
            avg_agi_below_10m = avg_agi
            avg_agi_above_10m = 0

        # 5% of amounts above 10 mil + 2% of amounts below 10 mil
        a = (avg_agi_below_10m*0.05) + (avg_agi_above_10m*0.02)

        # ORR highest of a and 100,000
        orr = max(a, 100000)

        # Map amt to Template  
        operational_varname = "orr_operational"
        self.add_bal_to_template_by_varname(operational_varname, orr) 

        # to delete
        print(f" Mapped {operational_varname} : {orr}")   

    def map_total_operational_rr(self):
        # Get OOR amt & OORR from Template
        operational_varname = "orr_operational"
        operational_row = self.mapper_class.varname_to_index.at[operational_varname]
        operational = self.outputdf.at[operational_row, "Balance"]

        other_operational_varname = "other_orr_operational"
        other_operational_row = self.mapper_class.varname_to_index.at[other_operational_varname]
        other_operational = self.outputdf.at[other_operational_row, "Balance"]
        # Check if the cell value is considered 'empty'
        if pd.isna(other_operational) or other_operational == "":
            other_operational = 0

        total_orr = operational + other_operational
        print(total_orr)

        # Map amt to Template  
        total_operational_varname = "trr_total_operational_risk_req"
        self.add_bal_to_template_by_varname(total_operational_varname, total_orr) 

        # to delete
        print(f" Mapped {total_operational_varname} : {total_orr}") 

    def map_total_risk_requirement(self):
        
        # Assume CRR, PRR, URR, LERR, FSRR all 0
        total_operational_varname = "trr_total_operational_risk_req"
        total_orr_row = self.mapper_class.varname_to_index.at[total_operational_varname]
        total_orr = self.outputdf.at[total_orr_row, "Balance"]

        # Map amt to Template  
        total_risk_req_varname = "ttr_total_risk_req"
        self.add_bal_to_template_by_varname(total_risk_req_varname, total_orr) 

        # to delete
        print(f" Mapped {total_risk_req_varname} : {total_orr}") 
        
    def map_fr_trr_ratio(self):
        # Get FR amt from Template
        fr_financial_resources_anhof_varname = "fr_financial_resources_anhof"
        fr_financial_resources_anhof_row = self.mapper_class.varname_to_index.at[fr_financial_resources_anhof_varname]
        fr = self.outputdf.at[fr_financial_resources_anhof_row, "Balance"]


        # Get trr amt from Template
        total_risk_req_varname = "ttr_total_risk_req"
        total_risk_req__row = self.mapper_class.varname_to_index.at[total_risk_req_varname]
        trr = self.outputdf.at[total_risk_req__row, "Balance"]
        
        # fr/trr *100 
        ratio = round(fr/trr*100, 2)

        # Map amt to Template  
        fr_trr_ratio_varname = "ttr_fr_trr_ratio"
        self.add_bal_to_template_by_varname(fr_trr_ratio_varname, ratio) 

        # to delete
        print(f" Mapped {fr_trr_ratio_varname} : {ratio}") 


#### (IV) Adjusted Assets 

    def create_aa_tempdf(self):
        # Adjusted Asset sections
        # Create 3 columns for last 3 month
        on_balance_assets_varname = "total_asset"
        on_balance_assets_row = self.mapper_class.varname_to_index.at[on_balance_assets_varname]

        aa_adjusted_assets_varname = "aa_adjusted_assets"
        aa_adjusted_assets_row = self.mapper_class.varname_to_index.at[aa_adjusted_assets_varname]
        
        aa_df = self.mapper_class.df_processed.copy()
        aa_df = aa_df.loc[on_balance_assets_row:aa_adjusted_assets_row]
        aa_df = aa_df.iloc[:, 0:5]


        aa_df.rename(columns={"Amount": self.col_name_3}, inplace=True) 
        col_names = aa_df.columns.tolist()
        col_names.insert(-1,self.col_name_2)
        col_names.insert(-1,self.col_name_1)
        aa_df = aa_df.reindex(columns = col_names)

        condition = aa_df.iloc[:,-1].notna() # change 999999 to Nan values in the last column 
        aa_df.iloc[condition, -1] = np.nan

        self.aa_df = aa_df

    def verify_credit_quality(self):
        
        tb_df = tb_class.get_data_by_fy(fy).copy()
        
        interval_list = [pd.Interval(5000,5000,closed = "both")]
        cash = tb_df[tb_df["L/S (interval)"].apply(lambda x:x in interval_list)]

        depo = cash[~cash["Name"].str.contains("(?i)cash")]
        
        new_cols = depo.columns.to_list() + ["Credit Quality Grade 1?"]
        
        depo = depo.reindex(columns=new_cols)
   
        
        print("Please tag if the following deposits are credit quality grade 1")
        
        depo.to_excel("Credit Quality.xlsx")
        dropdownlist = CreditDropdownList("Credit Quality.xlsx")
        depo = dropdownlist.create_dropdown_list()

        return depo
   
    def map_aa_df(self):
        month_end, first_month, second_month, third_month = self.gl_ageing()
        aa_df = self.aa_df.copy()
        print(aa_df)

        # On Balance sheet assets
        # from GL
         
        dec = third_month[third_month["L/S"]<6000]["Ending Balance"].sum()
        nov = second_month[second_month["L/S"]<6000]["Ending Balance"].sum()
        oct = first_month[first_month["L/S"]<6000]["Ending Balance"].sum()

        # pd.Interval(5000,6000, closed='left')

        on_balance_assets_varname = "total_asset"
        on_balance_assets_row = self.mapper_class.varname_to_index.at[on_balance_assets_varname]

        # Map to aa_df
        aa_df.at[on_balance_assets_row,self.col_name_1] = oct
        aa_df.at[on_balance_assets_row,self.col_name_2] = nov
        aa_df.at[on_balance_assets_row,self.col_name_3] = dec


        # Off balance sheet items 
        # Manual input
        answer  =  self.user_inputs.at["aa_off_bs_items", "Answer"]
        off_bal = answer.split(",")

        off_balance_assets_varname = "aa_off_bs_items"
        off_balance_assets_row = self.mapper_class.varname_to_index.at[off_balance_assets_varname]

        # Map to aa_df
        aa_df.loc[off_balance_assets_row, aa_df.columns[-3:]] = off_bal


        # Deductions from FR
        # from deductions_df
        fr_total_deductions_varname = "fr_total_deductions"
        fr_total_deductions_row = self.mapper_class.varname_to_index.at[fr_total_deductions_varname]

        total_deductions = self.deductions_df.loc[fr_total_deductions_row, self.deductions_df.columns[-3:]] 
        
        aa_fr_total_deductions_varname = "aa_fr_total_deductions"
        aa_fr_total_deductions_row = self.mapper_class.varname_to_index.at[aa_fr_total_deductions_varname]

        # Map to aa_df
        aa_df.loc[aa_fr_total_deductions_row, aa_df.columns[-3:]] = total_deductions


        # Cash and Deposit credit quality grade 1

        cash, depo = self._map_tempdf_by_ls([5000], "cash", type="Cash")

        aa_corp_own_cash_cashequiv_varname = "aa_corp_own_cash_cashequiv"
        aa_corp_own_cash_cashequiv_row = self.mapper_class.varname_to_index.at[ aa_corp_own_cash_cashequiv_varname]
        
        aa_df.loc[aa_corp_own_cash_cashequiv_row, aa_df.columns[-3:]] = cash

        aa_corp_own_deposit_bank_credit_quality_1_varname = "aa_corp_own_deposit_bank_credit_quality_1"
        aa_corp_own_deposit_bank_credit_quality_1_row = self.mapper_class.varname_to_index.at[aa_corp_own_deposit_bank_credit_quality_1_varname]

        aa_df.loc[aa_corp_own_deposit_bank_credit_quality_1_row, aa_df.columns[-3:]] = depo

        # Set values in aa_df to numeric 
        aa_df = aa_df.apply(pd.to_numeric, errors='ignore') 
        print(aa_df)

        self.aa_df = aa_df


#### AR processing for fees_receivables 
    def arprocessing(self, aged_receivables_fp, sheet_name):
        # Load the AR class
        aged_ar_class = common.AgedReceivablesReader_Format1(aged_receivables_fp, 
                                                        sheet_name = sheet_name, # Set the sheet name
                                                        variance_threshold = 0.1) # 1E-9) # To relax criteria if required.
        
        
        aged_group_dict = {"0-90": ["0 - 30", "31 - 60", "61 - 90"],
                           ">90": ["91 - 120", "121 - 150", "150+"]}
        
        # Then we get the AR by company (index) and by new bins (columns)
        aged_df_by_company = aged_ar_class.get_AR_by_new_groups(aged_group_dict)
        self.aged_df_by_company = aged_df_by_company

        return self.aged_df_by_company
        
    def update_fees_receivables_tempdf(self, aged_receivables_fp):

        # To get the accounts receivables data for each month (already in SGD)
        thirdmonth = self.arprocessing(aged_receivables_fp,sheet_name = 0)
        secondmonth = self.arprocessing(aged_receivables_fp,sheet_name = 1)
        firstmonth = self.arprocessing(aged_receivables_fp, sheet_name = 2)

        # This company is non-CLMS, so need to remove from the computation
        answer = self.user_inputs.at["non_clms", "Answer"]
        non_clms = answer.split(",")

        # Filter away/remove non_clms for each month 
        thirdmonth_amt = thirdmonth[~thirdmonth.index.isin(non_clms)]["0-90"].sum()     #1660896.34
        secondmonth_amt = secondmonth[~secondmonth.index.isin(non_clms)]["0-90"].sum()  #65991.49
        firstmonth_amt = firstmonth[~firstmonth.index.isin(non_clms)]["0-90"].sum()     #1440337.46

        amt_list = [firstmonth_amt, secondmonth_amt, thirdmonth_amt]
        print(amt_list)
        print(type(amt_list[0]))

        # Less payment on behalf 
        answer = self.user_inputs.at["payment_on_behalf", "Answer"]
        paymt_behalf = answer.split(",") 

        print(paymt_behalf)
        print(type(paymt_behalf))
        paymt_behalf = [-float(x) for x in paymt_behalf] # convert to float and make it negative
        print(paymt_behalf)


        fees_receivables_varname="aa_fee_receivables_cis_cef_ca_within3mths"
        fees_receivables_row = self.mapper_class.varname_to_index.at[fees_receivables_varname]

        # Map payment on behalf & clms amt to aa_df
        self.aa_df.loc[fees_receivables_row, self.aa_df.columns[-3:]] = paymt_behalf

        for i, amt in enumerate(amt_list):
            self.aa_df.loc[fees_receivables_row, self.aa_df.columns[-3 + i]] += amt
    
        print(self.aa_df)
    
    def update_adjusted_asset_tempdf(self):
        
        # Get index rows for mapping later 
        aa_adjusted_assets_varname = "aa_adjusted_assets"
        aa_adjusted_assets_row = self.mapper_class.varname_to_index.at[aa_adjusted_assets_varname]
  
        off_balance_assets_varname = "aa_off_bs_items"
        off_balance_assets_row = self.mapper_class.varname_to_index.at[off_balance_assets_varname]

        aa_fr_total_deductions_varname = "aa_fr_total_deductions"
        aa_fr_total_deductions_row = self.mapper_class.varname_to_index.at[off_balance_assets_varname]

        fees_receivables_varname="aa_fee_receivables_cis_cef_ca_within3mths"
        fees_receivables_row = self.mapper_class.varname_to_index.at[fees_receivables_varname]

        print(aa_adjusted_assets_row)
        print(off_balance_assets_row)
        print(aa_fr_total_deductions_row)

      
        # Sum part 1(a) to (b)
        bal_items = self.aa_df.loc[:off_balance_assets_row, :].sum(axis = 0, numeric_only = True)  
        print(bal_items)

        # Sum part 1(c) to 1(f)
        remaining_items = self.aa_df.loc[aa_fr_total_deductions_row:fees_receivables_row, :].sum(axis = 0, numeric_only = True)
        print(remaining_items)

        # Adjusted assets = sum of 1(a,b) - sum of 1(c-f)
        total = bal_items - remaining_items
        print(total)
        
        # Map Adjusted asset amt to aa_df
        self.aa_df.loc[aa_adjusted_assets_row, self.aa_df.columns[-3:]] = total 

        # Compute Average adjusted asset 
        avg_adj_asset = np.mean(self.aa_df.loc[aa_adjusted_assets_row, self.aa_df.columns[-3:]])

        # Save as attr
        self.avg_adj_asset = avg_adj_asset

        print(avg_adj_asset)
        print(self.aa_df) 

    def map_adjusted_asset(self):

        aa_df = self.aa_df.copy()
                
        # Filter for the last month (Dec) amts 
        append_df = aa_df.iloc[:,-1]

        # Get index rows for mapping later 
        total_asset_varname = "total_asset"
        total_asset_row = self.mapper_class.varname_to_index.at[total_asset_varname]

        aa_adjusted_assets_varname = "aa_adjusted_assets"
        aa_adjusted_assets_row = self.mapper_class.varname_to_index.at[aa_adjusted_assets_varname]

        # Map to Template
        self.outputdf.loc[total_asset_row:aa_adjusted_assets_row, "Balance"] = append_df

    def map_avg_adjusted_asset(self):

        aa_avg_adjusted_assets_varname = "aa_avg_adjusted_assets"

        # Map amt to Template  
        self.add_bal_to_template_by_varname(aa_avg_adjusted_assets_varname, self.avg_adj_asset)

        # to delete
        print(f" Mapped {aa_avg_adjusted_assets_varname} : {self.avg_adj_asset}") 

    def map_aa_threshold(self):
        # Determine Adjusted Assets Threshold (5* TFR or $10M)
        # Get tfr from Template
        fr_varname = "fr_financial_resources_anhof"
        fr_row = self.mapper_class.varname_to_index.at[fr_varname]
        fr = self.outputdf.at[fr_row, "Balance"]

        print(fr)
        print(fr*5)
        threshold = min(5*(fr), 10000000)
        print(threshold)

        aa_adjusted_assets_threshold_varname = "aa_adjusted_assets_threshold"        
       
        # Map amt to Template  
        self.add_bal_to_template_by_varname(aa_adjusted_assets_threshold_varname, threshold)

        # to delete
        print(f" Mapped {aa_adjusted_assets_threshold_varname} : {threshold}") 

    def column_mapper(self):

        # Map the Balance amounts to the correct field in F1; whether in the Amount or Subtotal column
        for i in self.outputdf.index:
            if pd.notna(self.outputdf.at[i,"Amount"]) and pd.notna(self.outputdf.at[i+1,"Amount"]): #if the 2 rows in Amount column is not empty and row 1's Balance is not empty then compute subtotal
                if pd.isna(self.outputdf.at[i,"Balance"]):
                    subtotal = subtotal
                else:
                    subtotal += self.outputdf.at[i,"Balance"]
            elif pd.notna(self.outputdf.at[i,"Amount"]) and pd.isna(self.outputdf.at[i+1, #if the next row in Amount column is empty
                    "Amount"]):
                if pd.isna(self.outputdf.at[i,"Balance"]) and subtotal != 0: #if the Balance for a row is empty but the subtotal is not 0
                    self.outputdf.at[i,"Subtotal"] = subtotal
                    subtotal = 0
                elif pd.notna(self.outputdf.at[i,"Balance"]):      #if the Balance for a row is not empty 
                    subtotal += self.outputdf.at[i,"Balance"]
                    self.outputdf.at[i,"Subtotal"] = subtotal
                    subtotal = 0
            else: 
                subtotal = 0

        for i in self.outputdf.index:
            if pd.notna(self.outputdf.at[i, "Amount"]):                # if Amount column is not empty
                self.outputdf.at[i, "Amount"] = self.outputdf.at[i, "Balance"]
            elif pd.isna(self.outputdf.at[i, "Amount"]) and pd.notna(self.outputdf.at[i, # if Amount is empty and Subtotal not empty
                    "Subtotal"]):
                self.outputdf.at[i, "Subtotal"] = self.outputdf.at[i, "Balance"]
            elif pd.isna(self.outputdf.at[i, "Amount"]) and pd.isna(self.outputdf.at[i, # if Amount is empty and Subtotal not empty
                    "Subtotal"]):
                self.outputdf.at[i, "Subtotal"] = self.outputdf.at[i, "Balance"]  # if both Amount and Subtotal is empty 



#### Output for auditors (AAA)
    def output_excel(self):
        
        template_fp = r"P:\YEAR 2023\TECHNOLOGY\Technology users\FS Vertical\f2\Average Adjusted Assets Template.xlsx"
        
        aa_df = self.aa_df.copy()

        # To open the workbook 
        # workbook object is created
        wb = openpyxl.load_workbook(template_fp)
        
        # Get workbook active sheet object
        sheet = wb.active


        # first row (on-balance sheet assets)
        first_row = list(aa_df.iloc[0, -3:])

        # Inserting values into cells D9:F9
        columns = ['D', 'E', 'F']
        for col, value in zip(columns, first_row):
            sheet[col + '9'] = value
        
        # second row (off-balance sheet assets)
        second_row = list(aa_df.iloc[1, -3:])       
        # Inserting values into cells D11:F11
        for col, value in zip(columns, second_row):
            sheet[col + '11'] = value
        
        # To Less: 
        # third row (Cash & cash equivalents)
        third_row = list(aa_df.iloc[5, -3:])
        third_row = list(map(lambda x: -float(x), third_row)) # convert to float and make it negative
        #Inserting values into celss D13:f13
        for col, value in zip(columns, third_row):
            sheet[col + '13'] = value
        
        # fourth row (Deposits)
        fourth_row = list(aa_df.iloc[6, -3:])
        fourth_row = list(map(lambda x: -float(x), fourth_row)) # convert to float and make it negative
        #Inserting values into celss D14:f14
        for col, value in zip(columns, fourth_row):
            sheet[col + '14'] = value

        # fifth row (deductions from FR)
        fifth_row = list(aa_df.iloc[3, -3:])
        fifth_row = list(map(lambda x: -float(x), fifth_row)) # convert to float and make it negative
        #Inserting values into celss D15:f15
        for col, value in zip(columns, fifth_row):
            sheet[col + '15'] = value
        
        # sixth row (receivables owed by corporations)
        sixth_row = list(aa_df.iloc[7, -3:])
        sixth_row = list(map(lambda x: -float(x), sixth_row)) # convert to float and make it negative
        #Inserting values into celss D17:f17
        for col, value in zip(columns, sixth_row):
            sheet[col + '17'] = value

        # seventh row (receivables owed by corporations)
        seventh_row = list(aa_df.iloc[8, -3:])
        seventh_row = list(map(lambda x: -float(x), seventh_row)) # convert to float and make it negative
        #Inserting values into cells D18:f18
        for col, value in zip(columns, seventh_row):
            sheet[col + '18'] = value


        # Sum (a) to (iv) to get Adjusted asset 
        for col in range(4,7):
            formulas = "=SUM({}{}:{}{})".format(get_column_letter(col),
                                                9,
                                                get_column_letter(col),
                                                20
                                                )
            sheet.cell(row= 21 , column= col).value = formulas

        # Average adjusted asset 
        sheet['I21'] = "=AVERAGE(D21:F21)"

        # 5*FR
        fr_financial_resources_anhof_varname = "fr_financial_resources_anhof"
        fr_financial_resources_anhof_row = self.mapper_class.varname_to_index.at[fr_financial_resources_anhof_varname]
        print(fr_financial_resources_anhof_row)

        fr = self.outputdf.at[fr_financial_resources_anhof_row, "Balance"]

        print(fr)
        sheet['I26'] = float(fr)*5

        # Adjusted assets threshold 
        sheet['I29'] = "=MIN(I26,I27)"

        # Has AAA exceeded the above threshold?
        sheet['I31']= '=IF((I21<I26)*(I21<I27),"No","Yes")'
        
        wb.save("updated_workbook.xlsx")

#### Output for auditors (TRR)
    def output_excel_trr(self):
        template_fp = r"P:\YEAR 2023\TECHNOLOGY\Technology users\FS Vertical\f2\TRR.xlsx"

         # To open the workbook 
        # workbook object is created
        wb = openpyxl.load_workbook(template_fp)
        
        # Get workbook active sheet object
        sheet = wb.active

        # Retrieve from datahub (need current year)
        datahub = pd.read_excel("D:\gohjiawey\Desktop\Form 3\draftf2 - Copy.xlsx")
        datahub_currentfy = datahub[datahub['FY'] == self.fy]

    ###For current year

        # Base capital (row 8,9,10)        
        puc_list  = ["puc_ord_shares", "puc_pref_share_noncumulative", "puc_reserve_fund"]

        for i, row in enumerate(datahub_currentfy["var_name"]): 
            if row in (puc_list):
                value = datahub_currentfy.at[i, "Balance"]
                sheet["I"+ str(i+6)] = value #row 8,9,10

        # unappropriated profit or loss
        puc_unappr_profit_or_loss = datahub_currentfy.at[5, "Balance"]
        sheet["I12"] =  puc_unappr_profit_or_loss
        print(puc_unappr_profit_or_loss)

        # dividend declared, interim loss 
        puc_less_list = ["upl_div_declared", "upl_interim_loss"]
        for i, row in enumerate(datahub_currentfy["var_name"]): 
            if row in (puc_less_list):
                value = datahub_currentfy.at[i, "Balance"]
                sheet["I"+ str(i+8)] = value  #row 14,15

        # Net head office funds 
        puc_net_head_office_funds = datahub_currentfy.at[9, "Balance"]
        sheet["I17"] = puc_net_head_office_funds
        print(puc_net_head_office_funds)

        # Base Capital/Net Head Office Funds 
        sheet["I19"] =  "=SUM(I8:I17)"

        # Financial Resources 
        fr_add_list = ["puc_pref_share_cumulative", "current_liab_redeemable_pref_share, noncurrent_liab_redeemable_pref_share", 
         "bc_qual_subord_loans_temp", "puc_rev_reserve", "puc_other_reserves","bc_interim_unappr_profit",
         "bc_cltv_impairment_allowance"]
        
        for i, row in enumerate(datahub_currentfy["var_name"]):
            if row in (fr_add_list):
                value = datahub_currentfy.at[i, "Balance"]
                sheet["I"+ str(i+7)] = value #row 22 to 28

        
        fr_less_list = [ "noncurrent_asset_goodwill_ia", "dfr_future_incometax_benefits", "current_asset_other_prepayment", 
                "dfr_charged_assets", "dfr_unsecured_directors_cnntpersons", "dfr_unsecured_rlt_corporations",
                "dfr_other_unsecured_loans", "dfr_capinvst_subsidiary_associate", "noncurrent_asset_investment_in_subsi", 
                "dfr_other_assets_nonconvertible_cash"]

                     
        for i, row in enumerate(datahub_currentfy["var_name"]):
            if row in (fr_less_list):
                value = datahub_currentfy.at[i, "Balance"]
                sheet["I"+ str(i+5)] = -value #set to negative for less   #row 29 to 38 

        # Total Financial Resources ("FR")
        sheet["I40"] = "=SUM(I19:I38)"
        
        # ORR Highest of:
        sheet["I48"] = "=MAX(F50,F53)"

        # 5% of average annual gross income..
        sheet["F50"] = "=SUM(E92:E94)"

        # Total Risk Requirement ("TRR")
        sheet["I66"] = "=I48+I64"

        # Ratio: Financial Resources / Total Risk Requirement ("FR/TRR")
        sheet["I71"] = "=I40/I66"
                

    ###For last 3 years
        # Retrieve from datahub 
        datahub_form3 = pd.read_excel("D:\gohjiawey\Desktop\Form 3\draft - Copy.xlsx")
        
        datahub_form3_current = datahub_form3[datahub_form3["FY"] == self.fy]

        required_varname_list = ["rev_total_revenue",
                                "exp_fee_expense", 
                                 "exp_comm_expense_otherbroker", 
                                 "exp_comm_expense_agents", 
                                 "exp_int_expense",
                                 "rev_int_others", 
                                 "rev_dividend", 
                                 "rev_other_revenue" 
                                 ]

        varname_dict = {}
        for i, row in enumerate(datahub_form3_current["var_name"]):
            if row in (required_varname_list):
                value = datahub_form3_current.at[i, "Balance"]
                varname_dict[row] = value

        # Total Revenue        
        sheet["E78"] = varname_dict.get("rev_total_revenue", 0)
        
        # Fees expense 
        sheet["E79"] = varname_dict.get("exp_fee_expense", 0)

        # Commission expense 
        agents = varname_dict.get('exp_comm_expense_agents', 0)  # Default to 0 if key not present
        otherbroker = varname_dict.get('exp_comm_expense_otherbroker', 0)  # Default to 0 if key not present
        sheet["E80"] = agents + otherbroker
        
        # Interest expense 
        sheet["E81"] = varname_dict.get("exp_int_expense", 0)


        # Adjusted annual gross income 
        for col in range(3,6):
            formulas = "=SUM({}{}:{}{})".format(get_column_letter(col),
                                            78,
                                            get_column_letter(col),
                                            85
                                            )
            sheet.cell(row= 86 , column= col).value = formulas
            

        # Average annual gross income 
        sheet["E88"] = "=AVERAGE(C86,D86,E86)"

        # seperated into amts < S$10,000,000
        sheet["E89"] = "=MIN(E88,10000000)"

        # seperated into amts > S$10,000,000
        sheet["E90"] = "=MAX(0,E88-10000000)"

        # 5% average annual gross below S$10m
        sheet["E92"] = "=E89*0.05"

        # 2% of average annual gross income above S$10m
        sheet["E94"] = "=E90*0.02"

        # Save the workbook
        wb.save("updated_trr.xlsx")
        
        
if __name__ == "__main__":

    # Get the luna folderpath 
    luna_init_file = luna.__file__
    luna_folderpath = os.path.dirname(luna_init_file)
    print (f"Your luna library is at {luna_folderpath}.")
    
    # Get the template folderpath
    template_folderpath = os.path.join(luna_folderpath, "templates")
    
    # AGED RECEIVABLES
    if True:
        #aged_receivables_fp = os.path.join(template_folderpath, "aged_receivables.xlsx")
        #aged_receivables_fp = r"D:\Desktop\owgs\CODES\luna\personal_workspace\dacia\aged_receivables_template.xlsx"
        aged_receivables_fp = r"P:\YEAR 2023\TECHNOLOGY\Technology users\FS Vertical\Form 1\f1 input data\clean_AR_listing.xlsx"
        print (f"Your aged_receivables_fp is at {aged_receivables_fp}.")
        
        # # Load the AR class
        # aged_ar_class = common.AgedReceivablesReader_Format1(aged_receivables_fp, 
        #                                                 sheet_name = 0,            # Set the sheet name
        #                                                 variance_threshold = 0.1) # 1E-9) # To relax criteria if required.
        
        # aged_group_dict = {"0-90": ["0 - 30", "31 - 60", "61 - 90"],
        #                    ">90": ["91 - 120", "121 - 150", "150+"]}
        
        # # Then we get the AR by company (index) and by new bins (columns)
        # aged_df_by_company = aged_ar_class.get_AR_by_new_groups(aged_group_dict)

        # print(aged_df_by_company)
        
    # TB
    if True:
        #tb_fp = os.path.join(template_folderpath, "tb.xlsx")
        tb_fp = r"P:\YEAR 2023\TECHNOLOGY\Technology users\FS Vertical\f2\f2_tb_used.xlsx"
        
        
        print (f"Your tb_filepath is at {tb_fp}.")
        
        # Load the tb
        fy_end_date = datetime.date(2022, 12, 31)
        tb_class = common.TBReader_ExcelFormat1(tb_fp, 
                                                sheet_name = 0,
                                                fy_end_date = fy_end_date)
        
        
        # Get data by fy
        fy = 2022
        tb2022 = tb_class.get_data_by_fy(fy)
        
    # Form 2 mapping
    if True:
        
        mas_tb_mapping_fp = os.path.join(luna_folderpath, "parameters", "mas_forms_tb_mapping.xlsx")
        print (f"Your mas_tb_mapping_fp is at {mas_tb_mapping_fp}.")
        
        # Load the class
        mapper_class = fsvi.mas.MASTemplateReader_Form1(mas_tb_mapping_fp, sheet_name = "Form 2 - TB mapping")
    
        # process df is here:
        df_processed = mapper_class.df_processed  # need to build methods
    
    # GL
    if True: 
        
        gl_fp = r"P:\YEAR 2023\TECHNOLOGY\Technology users\FS Vertical\f2\GL FY2023.xlsx"
        print (f"Your gl_filepath is at {gl_fp}.")  

        # Load the gl
        gl_class = GLProcessor(gl_fp)

        gl_processed = gl_class.gl
        gl_processed
    
    # CLASS
    fy=2022
    self = MASForm1_Generator(tb_class,
                              mapper_class, gl_class, fy=fy)
    
    # Get df by varname
    # filtered_tb = self.filter_tb_by_varname('current_asset_trade_debt_other')
    
    # Output to excel 
    self.outputdf.to_excel("draftf2.xlsx") 


    

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd
import re
from copy import copy
from datetime import datetime

import luna
from luna.fsvi.mas.template_reader import MASTemplateReader_Form3
from luna.lunahub import tables
import os

class OutputFormatter:

    COLUMN_MAPPER = {"target_ocr_prevfy_excelcol"   : "E",
                     "target_ocr_currfy_excelcol"   : "F",
                     "target_prevfy_excelcol"       : "G",
                     "target_currfy_excelcol"       : "H",
                     "target_var_prevfy_excelcol"   : "I",
                     "target_var_currfy_excelcol"   : "J",
                     "target_ls_prevfy_excelcol"    : "K",
                     "target_ls_currfy_excelcol"    : "L",
                     "target_varname_excelcol"      : "M"
                     }

    def __init__(self, input_fp, output_fp, ocr_fp, fy, client_class, aic_name = "", mic_name = ""):

        self.input_fp       = input_fp
        self.output_fp      = output_fp
        self.ocr_fp         = ocr_fp
        self.fy             = fy
        self.client_class   = client_class
        self.aic_name       = aic_name
        
        # Get the template fp
        self.template_fp = os.path.join(
            luna.settings.LUNA_FOLDERPATH,
            "parameters", 
            "mas_forms_tb_mapping.xlsx")
        
        self.template_sheetname = "Form 3 - TB mapping"
        
        self.template_class = MASTemplateReader_Form3(
            self.template_fp, 
            self.template_sheetname)
        
        
        self.main()

    def main(self):

        self.read_files()
        self.write_output()

    def read_files(self):
        # Read the file with the values
        self.input_df = pd.read_excel(self.input_fp)
        self.ocr_df = pd.read_excel(self.ocr_fp)

    
    def build_varname_to_values(self, df):
        
        df = df.copy()
    
        varname_to_values = df[
            ["var_name", "Previous Balance", "Balance"]
            ].dropna(subset=["var_name"])
        
        # Check that var_name is unique
        if varname_to_values["var_name"].is_unique:
            varname_to_values = varname_to_values.set_index("var_name")
        else:
            raise Exception ("Variable name is not unique.")
                    
        # self.varname_to_values = varname_to_values.copy()
        
        return varname_to_values
    
    def _load_client_info(self):

        self.client_name = self.client_class.retrieve_client_info("CLIENTNAME")

    
    def _copy_columns(self, ws,
                      src_excelcol, dst_excelcol,
                      src_col_name, dst_col_name,
                      dst_value = None
                      ):
        for src, dst in zip(ws[f"{src_excelcol}:{src_excelcol}"], ws[f"{dst_excelcol}:{dst_excelcol}"]):
            if src.value == src_col_name:
                dst.value = dst_col_name
                dst.font = Font(bold = True)
                src.font = Font(bold = True)
                dst.alignment = Alignment(wrapText   = True,
                                          horizontal = 'center')
            elif src.value is not None and dst_value is None:
                dst.value = src.value
                dst.border = copy(src.border)
            elif src.value is not None and dst_value is not None:
                dst.value = dst_value
                dst.border = copy(src.border)
            else:
                dst.value = None
                dst.border = copy(src.border)

    def _copy_column_style(self, ws,
                           src_excelcol, dst_excelcol
                           ):
        for src, dst in zip(ws[f"{src_excelcol}:{src_excelcol}"], ws[f"{dst_excelcol}:{dst_excelcol}"]):
            if src.value is not None:
                dst.border = copy(src.border)


    def _create_header(self, ws):

        # Unmerge all cells
        for merge in list(ws.merged_cells):
            ws.unmerge_cells(range_string=str(merge))

        ws.delete_rows(idx = 1, amount = 3)
        ws.insert_rows(idx = 0, amount = 8)

        for merge in list(ws.merged_cells):
            ws.unmerge_cells(range_string=str(merge))

        row = 1

        # header title
        ws[f"A{row}"].value = "Form 3 - Statement relating to the accounts of a holder of a capital markets services licence"
        ws[f"A{row}"].font = Font(size = "16", bold = True)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="00C0C0C0")
        ws[f"A{row}"].border = Border(left   = Side(style = "thin"),
                                 right  = Side(style = "thin"),
                                 top    = Side(style = "thin"),
                                 bottom = Side(style = "thin")
                                 )
        ws.merge_cells(f"A{row}:F{row}")

        date_of_analysis = datetime.now().strftime("%d/%m/%Y")

        dct_of_fields = {"Client name"      : self.client_name,
                         "FY"               : self.fy,
                         "Date of analysis" : date_of_analysis,
                         "Prepared by"      : self.aic_name,
                         "Reviewed by"      : ""
                         }
        
        for key in dct_of_fields:
            
            row += 1

            self._create_header_row_field(key, dct_of_fields[key], row, ws)

    def _create_header_row_field(self, field_title, field_value, row, ws):

        border_style = Border(left   = Side(style = "thin"),
                              right  = Side(style = "thin"),
                              top    = Side(style = "thin"),
                              bottom = Side(style = "thin")
                              )
    
        ws[f"A{row}"].value = field_title
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="00C0C0C0")
        ws[f"A{row}"].font = Font(bold = True)
        ws[f"A{row}"].border = border_style
        ws.merge_cells(f"A{row}:C{row}")

        ws[f"D{row}"].value = field_value
        ws[f"D{row}"].fill = PatternFill("solid", fgColor="00FFFFFF")
        ws[f"D{row}"].alignment = Alignment(horizontal = 'left')
        ws[f"D{row}"].border = border_style
        ws.merge_cells(f"D{row}:F{row}")

    def _standardise_number_format(self, ws, lst_of_excelcols, row):
        for excelcol in lst_of_excelcols:
            ws[f"{excelcol}{row}"].number_format = '#,##0.00'

    def _replace_empty_value(self, ws, excelcol, row):
        cell = ws[f"{excelcol}{row}"]

        if cell.value in [999999999, '999999999']:
            cell.value = None

    def _create_var_formula(self, ws, excelcol, excelrow, val):
        
        cell = ws[f"{excelcol}{excelrow}"]

        var_target_excelcol = self._get_col_letter_from_ref(excelcol, 2)
        ocr_target_excelcol = self._get_col_letter_from_ref(excelcol, -2)
        if cell.value == val:
            ws[f"{var_target_excelcol}{excelrow}"].value = f"= {ocr_target_excelcol}{excelrow+5} - {excelcol}{excelrow+5}"
        # TODO : excelrow+5 is wrong

    def _get_col_letter_from_ref(self, ref_excelcol, mvmt):
        ref_colno = column_index_from_string(ref_excelcol)
        target_colno = ref_colno + int(mvmt)
        target_excelcol = get_column_letter(target_colno)

        return target_excelcol
    
    def _section_column_formatting(self, ws, section_name, starting_excelcol):
        section_color_dict = {'recalculated': '00CCFFFF',
                              'ocr'         : '00CCCCFF',
                              'var'         : '00CCFFCC'
                              }
        fill_setting = PatternFill("solid", fgColor = section_color_dict.get(section_name))
        ending_excelcol = self._get_col_letter_from_ref(starting_excelcol, 1)
        cols = ws[f"{starting_excelcol}8:{ending_excelcol}98"] #TODO: look into how to dynamically reference rows
        for cell in cols:
            cell[0].fill = fill_setting
            cell[1].fill = fill_setting

    def _create_col_header_1(self, ws, excelcol, row, header_name):
            cell = ws[f"{excelcol}{row}"]
            cell.value = header_name
            cell.font = Font(bold = True)
            cell.alignment = Alignment(horizontal = 'center')
            end_col = self._get_col_letter_from_ref(excelcol, 1)
            ws.merge_cells(f"{excelcol}{row}:{end_col}{row}")
            
    def write_output(self):
        templ_wb = openpyxl.load_workbook(self.template_fp)
        templ_ws = templ_wb[self.template_sheetname]

        sheets_to_remove = [sheet_name for sheet_name in templ_wb.sheetnames if sheet_name != self.template_sheetname]        
        
        for sheet_name in sheets_to_remove:
            templ_wb.remove(templ_wb[sheet_name])

        self._load_client_info()
        
        # Template as df
        colname_to_excelcol = self.template_class.colname_to_excelcol
        varname_to_index = self.template_class.varname_to_index

        # Rename index of colname_to_excelcol series
        new_index = ["Header 1", "Header 2", "Header 3", "Header 4", "Previous Balance", "Balance", "var_name"]
        if colname_to_excelcol.shape[0] == len(new_index):
            colname_to_excelcol = colname_to_excelcol.set_axis(new_index, axis = 0) 

        # Get the data
        varname_to_values = self.build_varname_to_values(self.input_df)
        self.varname_to_values = varname_to_values.copy()

        # Get data from ocr
        varname_to_values_ocr = self.build_varname_to_values(self.ocr_df)
        self.varname_to_values_ocr = varname_to_values_ocr.copy()

        # Process input_df
        input_df = self.input_df
        input_df = input_df[["Header 3", "var_name"]]
        input_df = input_df[input_df["var_name"].notnull()]
        input_df = input_df.set_index("var_name")
        
        # Save column index
        prevfy_excelcol = colname_to_excelcol.at["Previous Balance"]
        currfy_excelcol = colname_to_excelcol.at["Balance"]
        header_3_excelcol = colname_to_excelcol.at["Header 3"]
        varname_excelcol = colname_to_excelcol.at["var_name"]
        
        # Initialise excelcol
        excelcol_mapper = OutputFormatter.COLUMN_MAPPER
        target_varname_excelcol     = excelcol_mapper.get("target_varname_excelcol")
        target_ls_prevfy_excelcol   = excelcol_mapper.get("target_ls_prevfy_excelcol")
        target_ls_currfy_excelcol   = excelcol_mapper.get("target_ls_currfy_excelcol")
        target_ocr_prevfy_excelcol  = excelcol_mapper.get("target_ocr_prevfy_excelcol")
        target_ocr_currfy_excelcol  = excelcol_mapper.get("target_ocr_currfy_excelcol")
        target_var_prevfy_excelcol  = excelcol_mapper.get("target_var_prevfy_excelcol")
        target_var_currfy_excelcol  = excelcol_mapper.get("target_var_currfy_excelcol")
        target_prevfy_excelcol      = excelcol_mapper.get("target_prevfy_excelcol")
        target_currfy_excelcol      = excelcol_mapper.get("target_currfy_excelcol")

        # Copy to another column
        self._copy_columns(templ_ws,
                           varname_excelcol, target_varname_excelcol,
                           "var_name", "var_name",
                           dst_value = None)
        self._copy_columns(templ_ws,
                           prevfy_excelcol, target_prevfy_excelcol,
                           "Previous year\n<<<previous_fy>>>\n$", f"Previous year\n{int(self.fy)-1}\n$",
                           dst_value = None)
        self._copy_columns(templ_ws,
                           currfy_excelcol, target_currfy_excelcol,
                           "Current year\n<<<current_fy>>>\n$", f"Current year\n{self.fy}\n$",
                           dst_value = None)
        self._copy_columns(templ_ws,
                           prevfy_excelcol, target_ls_prevfy_excelcol,
                           "Previous year\n<<<previous_fy>>>\n$", f"Previous year\n{int(self.fy)-1}\n$",
                           dst_value = None)
        self._copy_columns(templ_ws,
                           currfy_excelcol, target_ls_currfy_excelcol,
                           "Current year\n<<<current_fy>>>\n$", f"Current year\n{self.fy}\n$",
                           dst_value = None)
        self._copy_columns(templ_ws,
                           prevfy_excelcol, target_ocr_prevfy_excelcol,
                           "Previous year\n<<<previous_fy>>>\n$", f"Previous year\n{int(self.fy)-1}\n$",
                           dst_value = "")
        
        self._copy_columns(templ_ws,
                           currfy_excelcol, target_ocr_currfy_excelcol,
                           "Current year\n<<<current_fy>>>\n$", f"Current year\n{int(self.fy)}\n$",
                           dst_value = "")
        
        # shifting original amt and subtotal excelcol to another col
        prevfy_excelcol = target_prevfy_excelcol
        currfy_excelcol = target_currfy_excelcol
    
        for varname in varname_to_values.index:
            prevfy = varname_to_values.at[varname, "Previous Balance"]
            currfy = varname_to_values.at[varname, "Balance"]
            if re.match("rev_other_revenue_\d+.*", varname) or re.match("exp_other_expense_\d+.*", varname):
                sig_acct_name = input_df.loc[varname, "Header 3"]
            
            # Get the location to update
            row = varname_to_index.at[varname]

            # Update
            templ_ws[f"{prevfy_excelcol}{row}"].value = prevfy
            templ_ws[f"{currfy_excelcol}{row}"].value = currfy
            if re.match("rev_other_revenue_\d+.*", varname) or re.match("exp_other_expense_\d+.*", varname):
                templ_ws[f"{header_3_excelcol}{row}"].value = sig_acct_name

            # Replace values declared with 999999999 with None instead
            self._replace_empty_value(templ_ws, target_ls_prevfy_excelcol, row)
            self._replace_empty_value(templ_ws, target_ls_currfy_excelcol, row)

            # Create formula for variance column to compare values of client vs rsm
            self._create_var_formula(templ_ws, prevfy_excelcol, row, prevfy)
            self._create_var_formula(templ_ws, currfy_excelcol, row, currfy)

            # Update number format
            lst_of_excelcols = [prevfy_excelcol, currfy_excelcol,
                                target_var_prevfy_excelcol, target_var_currfy_excelcol,
                                target_ocr_prevfy_excelcol, target_ocr_currfy_excelcol
                                ]
            self._standardise_number_format(templ_ws, lst_of_excelcols, row)

        for varname in varname_to_values_ocr.index:
            prevfy_ocr = varname_to_values_ocr.at[varname, "Previous Balance"]
            currfy_ocr = varname_to_values_ocr.at[varname, "Balance"]
            
            # Get the location to update
            row = varname_to_index.at[varname]
            
            # Update
            templ_ws[f"{target_ocr_prevfy_excelcol}{row}"].value = prevfy_ocr
            templ_ws[f"{target_ocr_currfy_excelcol}{row}"].value = currfy_ocr

        templ_ws[f"{target_var_prevfy_excelcol}4"].value = f"Previous year\n{int(self.fy)-1}\n$"
        templ_ws[f"{target_var_prevfy_excelcol}4"].font = Font(bold = True)
        templ_ws[f"{target_var_prevfy_excelcol}4"].alignment = Alignment(wrapText   = True,
                                                                         horizontal = 'center')
        templ_ws[f"{target_var_currfy_excelcol}4"].value = f"Current year\n{self.fy}\n$"
        templ_ws[f"{target_var_currfy_excelcol}4"].font = Font(bold = True)
        templ_ws[f"{target_var_currfy_excelcol}4"].alignment = Alignment(wrapText   = True,
                                                                         horizontal = 'center')


        self._copy_column_style(templ_ws, prevfy_excelcol, target_var_prevfy_excelcol)
        self._copy_column_style(templ_ws, currfy_excelcol, target_var_currfy_excelcol)

        self._create_header(templ_ws)

        self._section_column_formatting(templ_ws, "recalculated", prevfy_excelcol)
        self._section_column_formatting(templ_ws, "ocr", target_ocr_prevfy_excelcol)
        self._section_column_formatting(templ_ws, "var", target_var_prevfy_excelcol)
        

        # titles
        row = 8 #TODO: should not hardcode
        self._create_col_header_1(templ_ws, target_ls_prevfy_excelcol, row, "L/S")
        self._create_col_header_1(templ_ws, target_ocr_prevfy_excelcol, row, "Client")
        self._create_col_header_1(templ_ws, target_prevfy_excelcol, row, "RSM")
        self._create_col_header_1(templ_ws, target_var_prevfy_excelcol, row, "Variance")


        # conditional formatting
        redFill = PatternFill(start_color='EE1111',
                              end_color='EE1111',
                              fill_type='solid')
        templ_ws.conditional_formatting.add(f"{target_var_prevfy_excelcol}10:{target_var_currfy_excelcol}131",
                                      CellIsRule(operator='greaterThan', formula=['0.01'], stopIfTrue=True, fill=redFill))
        templ_ws.conditional_formatting.add(f"{target_var_prevfy_excelcol}10:{target_var_currfy_excelcol}131",
                                      CellIsRule(operator='lessThan', formula=['-0.01'], stopIfTrue=True, fill=redFill))
   
        # NOTE: forgot what this was for
        # templ_ws[f"{prevfy_excelcol}10"] = f"Previous year\n{int(self.fy)-1}\n$"
        # templ_ws[f"{prevfy_excelcol}10"].alignment = Alignment(wrapText   = True,
        #                                                       horizontal = 'center')
        # templ_ws[f"{currfy_excelcol}10"] = f"Current year\n{self.fy}\n$"
        # templ_ws[f"{currfy_excelcol}10"].alignment = Alignment(wrapText   = True,
        #                                                       horizontal = 'center')
        
        templ_ws.column_dimensions[target_varname_excelcol].hidden = True
        templ_ws.title = "Form 3 (Recalculated)"
        
        templ_wb.save(self.output_fp)
        templ_wb.close()
4
if __name__ == "__main__":

    if True: # Testing
        client_no   = 7167
        fy          = 2022

        #input_fp    = r"D:\workspace\luna\personal_workspace\tmp\mas_form1_7167_2022.xlsx"
        #template_fp = r"D:\workspace\luna\parameters\mas_forms_tb_mapping.xlsx"
        #output_fp   = r"D:\workspace\luna\personal_workspace\tmp\mas_form1_formatted_71679_2022.xlsx"

        client_class = tables.client.ClientInfoLoader_From_LunaHub(client_no)

        input_fp = rf"D:\workspace\luna\personal_workspace\tmp\mas_form3_{client_no}_{fy}.xlsx"
        ocr_fp = rf"D:\workspace\luna\personal_workspace\tmp\mas_form3_{client_no}_{fy}_ocr.xlsx"
        output_fp = rf"D:\workspace\luna\personal_workspace\tmp\mas_form3_formatted_{client_no}_{fy}.xlsx"
        
        aic_name = "John Smith"
        mic_name = "Jane Doe"
        
        self = OutputFormatter(input_fp     = input_fp,
                               output_fp    = output_fp,
                               ocr_fp       = ocr_fp,
                               fy           = fy,
                               client_class = client_class,
                               aic_name     = aic_name
                               )
        
    if True:
        import webbrowser
        webbrowser.open(output_fp)
        

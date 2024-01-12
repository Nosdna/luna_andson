'''
Sample script for Dacia and Jia Wey
'''

# Import standard libs
import os
import datetime
import sys
import pandas as pd
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
import numpy as np
import logging

# Initialise logger
logger = logging.getLogger()
if not(logger.hasHandlers()):
    logger.addHandler(logging.StreamHandler())

# Import luna package and fsvi package

import luna
import luna.common as common
import luna.fsvi as fsvi
from luna.fsvi.mas.template_reader import MASTemplateReader_Form3
import luna.lunahub.tables as tables


class MASForm3_Generator:
    SALARY_LS_CODES = [pd.Interval(7300, 7400, closed = 'left')]
    NET_FOREX_LS_CODES  = [pd.Interval(7410.400, 7410.400, closed = "both")]
    EXCLUDE_LS_CODES = [pd.Interval(5000,6900.4, closed = "both"), 
                      pd.Interval(7410.100,7410.400, closed='both')]
    EXP_EXCLUDE_LS_CODES= [pd.Interval(5000.000, 6900.400, closed = 'both'), 
                      pd.Interval(7410.100,7410.400, closed='both'), 
                      pd.Interval(7500.000, 7500.000, closed = "both")]
    
    FOREX_LS_CODES = [pd.Interval(7410.200, 7410.300, closed = "both")]
    EXCLUDE_TAX = [pd.Interval(7000.000,7500.000, closed = 'left')]
        

    
    def __init__(self, 
                 tb_class, mapper_class,
                 sig_acc_output_fp,
                 client_number,
                 fy, 
                 user_inputs = None):
        
        
        self.tb_class       = tb_class
        self.mapper_class   = mapper_class
        self.sig_acc_output_fp = sig_acc_output_fp
        self.client_number  = client_number
        self.fy             = fy
        self.user_inputs    = user_inputs 

        
        self.main()


       
    def main(self):
        
        # Will extract the tb and mapper of varname to lscodes
        self._map_varname_to_lscodes()

        # Prepare output container -> copy from mapper_class
        self._prepare_output_container()

        # Output debt accounts
        self.output_debt_accounts()

        # If user_inputs not specified, collect manual inputs interactively
        if self.user_inputs is None: 
            self.collect_manual_inputs()
        else: 
            pass
        
        # Map fields 
        self.map_debts()
        self.map_salary()
        self.map_director()           

        self.map_comm_rebates()
        self.map_reits_baseperf()
        self.map_reits_trans()
        
        self.map_mgmt_fees()
        self.map_adv_fees()
        self.map_corpfinfees_ipo()
        self.map_corpfinfees_others()
        self.map_trust_custdn()
        self.map_int_rev()
        self.map_div()        

        self.map_int_exp()
        self.map_tax()

        self.map_other_rev()
        self.map_total_rev()
        self.map_total_exp()
        self.map_other_exp()
        self.map_extraord_items()
        self.map_net_profit()

        self.create_sig_acct_output()
        self.design_acct_output()


    def _prepare_output_container(self):

      	# make a copy of the input template
        outputdf = self.mapper_class.df_processed.copy()

        # Add an output column
        outputdf["Balance"] = None

        # Save as attr
        self.outputdf = outputdf

    def add_bal_to_template_by_varname(self, varname, value):
        
        row = self.mapper_class.varname_to_index.at[varname]

        # Map value to template's Balance column
        self.outputdf.loc[row, "Balance"] = value
                
    def _map_varname_to_lscodes(self):
        
        mapper_class = self.mapper_class
        tb_class     = self.tb_class
        
        # get varname to ls code from mapper
        varname_to_lscodes = mapper_class.varname_to_lscodes

        # NOTE: edited by SJ to accommodate formulas on mapping template 20240111
        varname_to_lscodes_temp = varname_to_lscodes.copy()
        varname_to_lscodes_temp = varname_to_lscodes_temp.to_frame()
        pattern = ".*Interval.*"
        varname_to_lscodes_temp["filter"] = np.where(varname_to_lscodes_temp["L/S (intervals)"].astype(str).str.match(pattern), "yes", "no")
        varname_to_lscodes_ls = varname_to_lscodes_temp[varname_to_lscodes_temp["filter"] == "yes"]["L/S (intervals)"].squeeze()
        varname_to_lscodes_formula = varname_to_lscodes_temp[varname_to_lscodes_temp["filter"] == "no"]["L/S (intervals)"].squeeze()
        
        varname_to_lscodes = varname_to_lscodes_ls
       
        
        # get the tb for the current fy
        tb_df = tb_class.get_data_by_fy(self.fy).copy()
        tb_columns = tb_df.columns
        
        # screen thru
        for varname in varname_to_lscodes.index:
            #varname = "puc_pref_share_noncumulative"
            lscode_intervals = varname_to_lscodes.at[varname]
            
            # Get the true and false matches
            is_overlap, tb_df_true, tb_df_false = tb_class.filter_tb_by_fy_and_ls_codes(
                self.fy, lscode_intervals)
            
            # Update the main table
            tb_df[varname] = is_overlap
        varname_to_lscodes = pd.concat([varname_to_lscodes, varname_to_lscodes_formula], axis = 0)

        #
        self.tb_columns_main = tb_columns
        self.tb_with_varname = tb_df.copy()

    def filter_tb_by_varname(self, varname):
        
        tb = self.tb_with_varname.copy()
        
        filtered_tb = tb[tb[varname]][self.tb_columns_main]
        
        return filtered_tb

    def _check_empty_accts(self, varname):

        # Get df by varname
        filtered_tb = self.filter_tb_by_varname(varname)

        if filtered_tb.empty:
            value = 0.00
            logger.info(f"No accounts found. The amount will be set to 0.00 {filtered_tb}")

        else:
            value = abs(filtered_tb["Value"].sum())
        
        return value 


    def output_debt_accounts(self):
        '''
        Agreed treatment: 
        1) Filter account names with keywords "bad debt" or  "doubtful debt" 
        2) Output the list of accounts
        3) Auditor to select which accounts should map to (2a) Bad debts writted off or (2b) Provision for doubtful debts
        
        '''           
        tb = self.tb_class.get_data_by_fy(self.fy)

        baddebt_df = tb[tb["Name"].str.match(fr'(?i)bad\sdebts?')]
        provdebt_df = tb[tb["Name"].str.match(fr'(?i)pro.*\sdebts?|do.+\sdebts?')]
        
        #Save as attr 
        self.baddebt_df = baddebt_df
        self.provdebt_df = provdebt_df

    def collect_manual_inputs(self):
        '''
        To run only when the user_inputs parameter is not specified when
        the class is initialised.
        '''

        question_list = [
            "Enter the Account No for bad debts (NA if none): ",  
            "Enter the Account No for provision debts (NA if none): ", 
            "Enter Director's remuneration (0.00 if None): $ "]
        

        varname_list = ["exp_bad_debts", 
                        "exp_prov_dtf_debts", 
                        "exp_dir_renum"
                        ]

        self.user_inputs = pd.DataFrame({"Question": question_list, 
                                       "Answer": ""}, index = varname_list)
        
        self.user_inputs["Answer"] = self.user_inputs["Question"].apply(input)
        
    def map_debts(self):

        varname1 = "exp_bad_debts"
        varname2 = "exp_prov_dtf_debts"

        bad_debt_answer = self.user_inputs.at["exp_bad_debts", "Answer"]
        provision_debt_answer = self.user_inputs.at["exp_prov_dtf_debts", "Answer"]

        if bad_debt_answer is None:
            client_baddebt_list = "NA"
        else:
            client_baddebt_list = [i.strip() for i in bad_debt_answer.split(",")]

        if provision_debt_answer is None:
            client_provdebt_list = "NA"
        else:
            client_provdebt_list= [i.strip() for i in provision_debt_answer.split(",")]


        if (client_baddebt_list == "NA") & (client_provdebt_list == "NA"):
            logger.info(f"No accounts found. The amount will be set to 0.00.")
            bad_debt = 0.00
            provision_debt = 0.00

        else:
            # Amt 
            bad_debt = self.baddebt_df[self.baddebt_df["Account No"].isin(client_baddebt_list)]["Value"].sum()
            provision_debt = self.provdebt_df[self.provdebt_df["Account No"].isin(client_provdebt_list)]["Value"].sum()
            
        # Map amt to Template
        self.add_bal_to_template_by_varname(varname1, bad_debt)
        self.add_bal_to_template_by_varname(varname2, provision_debt)

        # Logger 
        logger.info(f"Mapped {varname1}: ${bad_debt}") 
        logger.info(f" Mapped {varname2} : ${provision_debt}")      

    def map_director(self):
        varname = "exp_dir_renum"

        # Amt
        answer = self.user_inputs.at["exp_dir_renum","Answer"]
        director = float(answer)

        # Map amt to Template 
        self.add_bal_to_template_by_varname(varname, director)

        # Logger 
        logger.info(f"Mapped {varname}: ${director}")   

        return director

    def map_salary(self):
        '''
        Treatment: 7300.000 - (h) Director's remuneration

        '''


        director = self.map_director()

        varname = "exp_slry"

        # Get 7300 series         
        boolean, true_match, false_match = \
            self.tb_class.filter_tb_by_fy_and_ls_codes(self.fy, MASForm3_Generator.SALARY_LS_CODES)
        
        # Amt 
        total_salary = true_match["Value"].sum()
        salary = total_salary - director   

        # Map amt to Template
        self.add_bal_to_template_by_varname(varname, salary)

       # Logger 
        logger.info(f"Mapped {varname}: ${salary}")   

    def map_comm_rebates(self):
        varname = "rev_comm_rebates"

        # If no accts, the amt will be set to 0
        commission_rebate = self._check_empty_accts(varname)

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, commission_rebate)  
            
        # Logger 
        logger.info(f"Mapped {varname}: ${commission_rebate}")   


    def map_reits_baseperf(self):     
        varname = "rev_reit_mgtfees_baseperf"

        # If no accts, the amt will be set to 0
        base_perf = self._check_empty_accts(varname)

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, base_perf)  
        
        # Logger 
        logger.info(f"Mapped {varname}: ${base_perf}")   

    def map_reits_trans(self):
        varname = "rev_reit_mgtfees_trans"
        
        # If no accts, the amt will be set to 0
        trans = self._check_empty_accts(varname)

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, trans)  

        # Logger 
        logger.info(f"Mapped {varname}: ${trans}")   

    def map_mgmt_fees(self):
        varname = "rev_pfl_mgtfees_mgmt_fees"

        # If no accts, the amt will be set to 0
        mgmt_fees = self._check_empty_accts(varname)

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, mgmt_fees)  

        # Logger 
        logger.info(f"Mapped {varname}: ${mgmt_fees}")      

    def map_adv_fees(self):
        varname = "rev_pfl_mgtfees_adv_fees"

        # If no accts, the amt will be set to 0
        advisory_fees = self._check_empty_accts(varname)

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, advisory_fees)
        
        # Logger 
        logger.info(f"Mapped {varname}: ${advisory_fees}")   
 
    def map_corpfinfees_ipo(self):
        varname = "rev_corp_finfees_ipo"

        # If no accts, the amt will be set to 0
        corpfinfees_ipo = self._check_empty_accts(varname)

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, corpfinfees_ipo)

        # Logger 
        logger.info(f"Mapped {varname}: ${corpfinfees_ipo}")   
    
    def map_corpfinfees_others(self):
        varname = "rev_corp_finfees_others"

        # If no accts, the amt will be set to 0
        corpfinfees_others = self._check_empty_accts(varname)

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, corpfinfees_others)      

        # Logger 
        logger.info(f"Mapped {varname}: ${corpfinfees_others}")  

    def map_trust_custdn(self):

        varname = "rev_trust_custdn"

        # If no accts, the amt will be set to 0
        trust_custdn = self._check_empty_accts(varname)

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, trust_custdn)   
            
        # Logger 
        logger.info(f"Mapped {varname}: ${trust_custdn}")   
   
    def map_int_rev(self):
        varname = "rev_int_others"

        # If no accts, the amt will be set to 0
        rev_int_others = self._check_empty_accts(varname)

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, rev_int_others)

        # Logger 
        logger.info(f"Mapped {varname}: ${rev_int_others}")   

    def map_div(self):

        varname = "rev_dividend"
        
        # If no accts, the amt will be set to 0
        dividend = self._check_empty_accts(varname)

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, dividend)

        # Logger 
        logger.info(f"Mapped {varname}: ${dividend}")   

    def map_int_exp(self):
        varname = "exp_int_expense"

        # If no accts, the amt will be set to 0
        exp_int_expense = self._check_empty_accts(varname)

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, exp_int_expense)

        # Logger 
        logger.info(f"Mapped {varname}: ${exp_int_expense}")   

    def map_tax(self):
        varname = "tax"

        # If no accts, the amt will be set to 0
        corp_tax = self._check_empty_accts(varname)

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, corp_tax)
                   
        # Logger 
        logger.info(f"Mapped {varname}: ${corp_tax}")   


    def map_other_rev(self):
        '''
        7410.100 to be assigned as 'Other revenue' (if <0 --> it is credit (-ve) balance)
        7410.2 Forex gain 
        7410.4 Net forex gain/loss (If < 0 -> it is credit (-ve) balance -> Forex gain)

        
        why need < 0 : 
        e.g. In CI's 2021, the account 'Loss on Fixed Asset Written of' is 7410.1 and is classified as Revenue-Other.
        But it has a positive value of 714, so it should be under other expense. 

        '''
        varname = "rev_other_revenue"

         # Get df by varname
        filtered_tb = self.filter_tb_by_varname(varname)

        if filtered_tb.empty:
            other_rev = 0.00
            logger.info(f"No Other revenue accounts. The amount will be set to 0.00 {filtered_tb}")

        else: 
            
            boolean, true_match, false_match =\
                  self.tb_class.filter_tb_by_fy_and_ls_codes(self.fy, MASForm3_Generator.NET_FOREX_LS_CODES)  
            
            # Sum L/S 7410.4
            forex = true_match["Value"].sum() 

            # Sum negative balances 
            other_rev = filtered_tb[~filtered_tb["L/S"].isin(['7410.4'])]
            other_rev = other_rev[other_rev["Value"] <0]
            other_rev = other_rev["Value"].sum()


            # if forex gain, add to other_rev
            if forex < 0:
                
                other_rev = abs(other_rev) + abs(forex)

            else: 
                other_rev = abs(other_rev)

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, other_rev)  

        # Logger 
        logger.info(f"Mapped {varname}: ${other_rev}")   
  
    def map_total_rev(self):
        ''' 
        
        To use TB values and Form 3 total from fields above to validate if numbers tally
        
        1) filter away 5000-6900.4 as we only want the L/S 7000 onwards, which are revenue and expense accounts
        2) filter away 7410.100-7410.400 as they are alr computed in map_other_rev
        3) If Total Revenue does not match, should throw an exception and log an error to indicate that there is computation issue.
        
        '''

    # Get other rev from Template
        otherrev_varname = "rev_other_revenue"
        otherrev_row = self.mapper_class.varname_to_index.at[otherrev_varname]
        other_rev = self.outputdf.at[otherrev_row,"Balance"]

    # First method: Compute from Template 
        varname = "rev_total_revenue"

        # Get index row
        row = self.mapper_class.varname_to_index.at[varname]

        # Amt
        total_revenue = abs(self.outputdf.loc[:row, "Balance"].sum())

        logger.info(f"total revenue computed from Template: {total_revenue}")

    # Second method: Compute from TB's Value
        
        # All revenue accounts without L/S 7410.400 (net forex), 
        # L/S 7410.100 (other rev), L/S 7410.200 (forex gain), L/S 7410.300(forex loss)

        boolean, true_match, false_match = \
            self.tb_class.filter_tb_by_fy_and_ls_codes(self.fy, MASForm3_Generator.EXCLUDE_LS_CODES)
        
        rev = false_match[false_match["Class"].str.match(r'Revenue.+')]
        rev = rev["Value"].sum()
        
        # Logger 
        logger.info(f"rev is {rev}")      

        total_rev = abs(rev) + abs(other_rev)
        
        try: 
            if total_revenue == total_rev:
                # Logger 
                logger.info(f"Total Revenue Match: {total_revenue}, {total_rev}")  
            else: 
                raise ValueError("Total Revenue does not Match")
        
        except ValueError as e:
                # Logger
                logger.error(f"Total Revenue does not Match: {total_revenue}, {total_rev}", exc_info=True)
                
                # Exit the script with a non-zero status to indicate an error
                sys.exit(1) 

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, total_rev)

        # save total_rev
        self.total_rev = total_rev  
          
    def map_total_exp(self):
        '''
        1) Filter for Expense class and sum the amts 
        2) Sum L/S 7410.1 that are positive 
        3) Sum exp w forex loss to get total exp
        
        '''
        
    # Compute from TB's Value 
        
        # Accounts without L/S 7410.400 (net forex), L/S 7410.100 (other rev),
        # L/S 7410.200 (forex gain), L/S 7410.300(forex loss), L/S 7500 (tax)

        boolean, true_match, false_match = \
            self.tb_class.filter_tb_by_fy_and_ls_codes(self.fy, MASForm3_Generator.EXP_EXCLUDE_LS_CODES)
        
        # Filter for Expense accounts and sum amt
        exp = false_match[false_match["Class"].str.match(r'Expense.+')]
        exp = exp["Value"].sum()

        # Sum L/S 7410.1 thar are positive balances
        varname = "rev_other_revenue"
        filtered_tb = self.filter_tb_by_varname(varname)
        other_rev_pos = filtered_tb[filtered_tb["L/S"].isin(["7410.1"])]
        other_rev_pos = other_rev_pos[other_rev_pos["Value"] > 0]["Value"].sum()

        # Update total expense amt 
        exp = exp + other_rev_pos
        
        # Filter if L/S code is '7410.200' or '7410.300' or '7410.400'
        # if net forex is <0 then we add this to total expense

        boolean, true_match, false_match = self.tb_class.filter_tb_by_fy_and_ls_codes(self.fy, MASForm3_Generator.FOREX_LS_CODES)    
        
        if not true_match.empty:

            net_forex = true_match["Value"].sum()

            if net_forex > 0: 
                logger.info(f"Forex loss: ${net_forex}")   
                total_exp = exp + net_forex

            else: 
                logger.info(f"Forex gain: ${net_forex}")  
                total_exp = exp

        else:

            boolean, true_match, false_match = self.tb_class.filter_tb_by_fy_and_ls_codes(self.fy, MASForm3_Generator.NET_FOREX_LS_CODES)    
        
            net_forex = true_match["Value"].sum()

            if net_forex > 0: 
                logger.info(f"Forex loss: ${net_forex}") 
                total_exp = exp + net_forex

            else: 
                logger.info(f"Forex gain: ${net_forex}") 
                total_exp = exp
        
        # Map amt to Template  
        varname = "exp_total_expense"
        self.add_bal_to_template_by_varname(varname, total_exp)
        
        #save total_exp
        self.total_exp = total_exp

        # Logger 
        logger.info(f"Mapped {varname}: ${total_exp}")   

    def map_other_exp(self):
        
        '''
        Treatment: Calculate other expenses by total expenses minus the sum of the balances from 2(a) to (i)
        
        '''

        # Get Total exp amt from Template 
        totalexp_varname = "exp_total_expense"
        totalexp_row = self.mapper_class.varname_to_index.at[totalexp_varname]
        total_exp = self.outputdf.at[totalexp_row,"Balance"]

        # Get index row 
        salary_varname = "exp_slry" 
        debt_varname = "exp_bad_debts"
        debt_row = self.mapper_class.varname_to_index.at[debt_varname] 
        salary_row = self.mapper_class.varname_to_index.at[salary_varname]
        
        # Amt (Total exp - sum of balances from (2a) Debts to (2i) Salaries)
        other_exp = total_exp - self.outputdf.loc[debt_row:salary_row, "Balance"].sum()

        
        # Map amt to Template 
        varname = "exp_other_expense"
        self.add_bal_to_template_by_varname(varname, other_exp)  

        # Logger 
        logger.info(f"Mapped {varname}: ${other_exp}")   

    def map_extraord_items(self):

        varname = "extraord_items"


        extraord_items = 0.00

        # Map amt to Template  
        self.add_bal_to_template_by_varname(varname, extraord_items)

        # Logger 
        logger.info(f"Mapped {varname}: ${extraord_items}")   
        
    def map_net_profit(self):
    # Net profit before tax
        # First method: Compute from TB's (all 7000++ accounts except tax)

        boolean, true_match, false_match = \
            self.tb_class.filter_tb_by_fy_and_ls_codes(self.fy, MASForm3_Generator.EXCLUDE_TAX)
        
        net_profit = true_match["Value"].sum() *-1

        # Second Method: Compute from Template
        # Get total rev from Template
        rev_varname = "rev_total_revenue"
        total_rev_row = self.mapper_class.varname_to_index.at[rev_varname]
        total_rev = self.outputdf.at[total_rev_row,"Balance"]

        # Get total exp from Template
        exp_varname = "exp_total_expense"
        total_exp_row = self.mapper_class.varname_to_index.at[exp_varname]
        total_exp = self.outputdf.at[total_exp_row,"Balance"]
        net_profit_2 = total_rev - total_exp

        # try: 
        if net_profit_2 == net_profit:
            # Logger 
            logger.info(f"Net profit before tax Match: {net_profit}, {net_profit_2}")  
        else: 
            pass
    
        # Map amt to Template 
        npbt_varname = "npbt"
        self.add_bal_to_template_by_varname(npbt_varname, net_profit)  

    # Net profit after tax but before extraordinary items
        # Get tax amt from Template
        tax_varname = "tax"
        tax_row = self.mapper_class.varname_to_index.at[tax_varname]
        tax = self.outputdf.loc[tax_row,"Balance"]

        # After tax but before extraordinary items
        npbt_row = self.mapper_class.varname_to_index.at[npbt_varname]    
        npat_bef_extraord = self.outputdf.loc[npbt_row, "Balance"] - tax
        npat_bef_varname = "npat_bef_extraord"
        self.add_bal_to_template_by_varname(npat_bef_varname, npat_bef_extraord)      

        # Get extraordinary items from Template 
        extraord_varname = "extraord_items"
        extraord_row = self.mapper_class.varname_to_index.at[extraord_varname]
        extraord = self.outputdf.loc[extraord_row,"Balance"]

        # Net profit after tax and extraordinary items for the year
        npat_varname = "npat"
        npat = self.outputdf.loc[npbt_row +2, "Balance"] + extraord
        self.add_bal_to_template_by_varname(npat_varname, npat)   

        # Logger 
        logger.info(f"Mapped {npat_bef_varname}: ${npat_bef_extraord}")   
        logger.info(f"Mapped {npat_varname}: ${npat}")   


    def _calculate_percent_of_total(self,
                                   account_type : str
                                   ):
        
        if account_type.lower() in ['rev', 'revenue']:

            total = self.total_rev

        elif account_type.lower() in ['exp', 'expenses', 'expense']:

            total = self.total_exp

        else:

            logger.error(f"Account type '{account_type}' specified is not supported."
                         "Please indicate a different account type.")

        return abs(total * 0.00005)
    
    def filter_tb_by_amount(self, amount, filter_type, tb):

        query = f"Value {filter_type} {amount}"

        filtered_tb = tb.query(query)

        return filtered_tb

        
    def load_sig_accts_from_datahub(self, account_type, fy):

        # placeholder for db_reader

        # placeholder dataframe
        if False:
            placeholder_data = {"Account No"    : ["1973558", "1973550", "1973558"],
                                "Name"          : ["Realised Ex (Gain)/Loss (C)", "Placeholder Account", "Realised Ex (Gain)/Loss (C)"],
                                "L/S"           : ["7410.4", "7410.4", "7410.4"],
                                "Class"         : ["Revenue - other", "Revenue - other", "Revenue - other"],
                                "L/S (interval)": [[7410.4, 7410.4], [7410.4, 7410.4], [7410.4, 7410.4]],
                                "Value"         : [10568.0, 500000, 10569.0],
                                "Completed FY?" : [True, True, True],
                                "Group"         : ["Realised Ex (Gain)/Loss", "Placeholder", "Realised Ex (Gain)/Loss"],
                                "Type"          : ["Revenue", "Revenue", "Revenue"],
                                "FY"            : [2021, 2021, 2020],
                                "Indicator"     : ["Declared in prev FY", ">= 5% of total", ">= 5% of total"]
                                }
            
            placeholder_df = pd.DataFrame(placeholder_data)

            df = placeholder_df
        
        # Read from lunahub for current year
        reader_class = tables.fs_masf3_sig_accts.MASForm3SigAccts_DownloaderFromLunaHub(self.client_number, 
                                                                                        fy, lunahub_obj=None)
        reader_class.main()
        df = reader_class.df_processed
        
        # Read from lunahub for previous year
        prev_fy = int(fy) - 1
        reader_class_prevfy = tables.fs_masf3_sig_accts.MASForm3SigAccts_DownloaderFromLunaHub(self.client_number, 
                                                                                        prev_fy, lunahub_obj=None)
        reader_class_prevfy.main()
        df_prevfy = reader_class_prevfy.df_processed

        # Concat for both years
        df_concat = pd.concat([df, df_prevfy], axis=0)
        
        filtered_df = df_concat[df_concat["Type"] == account_type]

        return filtered_df
    
    def validate_and_update_df(self, df, required_columns):
        # Step 1: Check if all required columns are present in the DataFrame
        missing_columns = set(required_columns) - set(df.columns)

        # Step 2: Append empty columns for missing ones
        for col in missing_columns:
            df[col] = pd.Series(dtype='object')

        # Step 3: Filter out extra columns
        df = df[required_columns]  # This will only keep the required columns
        
        return df


    def filter_for_sig_acct(self, fy, acc_type):

        total_5_percent = self._calculate_percent_of_total(acc_type)

        tb = self.tb_with_varname.copy()

        cols_lst = []

        if acc_type.lower() in ["rev", "revenue"]:
            search_query = "rev_.*"
            type_search = "Revenue"
        elif acc_type.lower() in ["exp", "expense", "expenses"]:
            search_query = "exp_.*"
            type_search = "Expenses"
        else:
            logger.error(f"Type '{acc_type}' specified is not supported."
                         "Please indicate a different account type.")

        for col in tb.columns:
            if re.search(search_query, col):
                cols_lst.append(col)

        query_str = ""
        
        for col in cols_lst:
            query_str += f"({col} == True)"
            if col != cols_lst[-1]:
                query_str += " or "
        filtered_tb = tb.query(query_str)

        filtered_tb = self.filter_tb_by_amount(total_5_percent, ">=", filtered_tb)

        # check if returned df is empty
        if filtered_tb.empty:
            logger.info("There are no accounts in the TB that have values above "
                        f"the threshold of {total_5_percent}.")
        else:
            filtered_tb["Indicator"] = f">= 5% of total"

        prev_fy_sig_accts = self.load_sig_accts_from_datahub(type_search, fy-1)

        cols_to_keep = ["Account No", "Name", "L/S", "Value", "Group", "FY",
                        "Indicator"
                        ]
        
        filtered_tb = self.validate_and_update_df(filtered_tb, cols_to_keep)
        prev_fy_sig_accts = self.validate_and_update_df(prev_fy_sig_accts, cols_to_keep)

        combined_df = pd.concat([prev_fy_sig_accts, filtered_tb], axis = 0, ignore_index = True)

        combined_df.sort_values(by = "FY", inplace = True)

        combined_df["Indicator"] = combined_df["Indicator"] + " for FY" + combined_df["FY"].astype(str)
        
        combined_df["Group"].fillna(value = "", inplace = True)
        combined_df["Group"] = combined_df["Group"].astype(str)
        

        groupby_cols = ["Account No", "Name"]
        
        final_df = combined_df.groupby(groupby_cols).agg({"L/S"         : "first",
                                                          "Value"       : "first",
                                                          "Indicator"   : lambda x : ", ".join(pd.unique(x)),
                                                          "Group"       : lambda x : ", ".join(pd.unique(x))}
                                                          ).reset_index()
        
        final_df["Type"] = type_search

        col_order = ["Account No", "Name", "L/S", "Value", "Type", "Indicator", "Group"]
        final_df = final_df[col_order]

        return final_df
    
    def create_sig_acct_output(self):

        rev_accts = self.filter_for_sig_acct(fy   = self.fy,
                                             acc_type = "rev"
                                             )
        
        exp_accts = self.filter_for_sig_acct(fy   = self.fy,
                                             acc_type = "exp"
                                             )
        
        acct_output = pd.concat([rev_accts, exp_accts],
                                axis = 0,
                                ignore_index = True
                                )
        
        self.acct_output = acct_output

        return acct_output
    
    def design_acct_output(self):

        fp = self.sig_acc_output_fp
        
        if fp is None:
            # no need to save, i.e. when run for previous year
            pass
        
        else:
            
            df = self.acct_output
    
            df.fillna(0, inplace = True)
    
            df.to_excel(fp, index = False, sheet_name = "Sheet1")
    
            wb = openpyxl.load_workbook(fp)
            ws = wb.active
    
            col_letter = openpyxl.utils.get_column_letter(df.shape[1]+1)
    
            # create new col and format header
            ws[f'{col_letter}1'] = f'Declare for current FY?'
            old_cell = ws['A1']
            new_cell = ws[f'{col_letter}1']
            new_cell.border = copy(old_cell.border)
            new_cell.font = copy(old_cell.font)
        
            dv = DataValidation(type     = "list",
                                formula1 = '"Yes, No"',
                                allow_blank =True
                                )
            
            # Only add a last column if df has data
            if df.shape[0] > 0:    
                dv.add(f'{col_letter}2:{col_letter}{df.shape[0] + 1}')
                ws.add_data_validation(dv)
            else:
                pass

            wb.save(fp)
            wb.close()
            
    def write_output(self, output_fp = None):
        
        if output_fp is None:
            logger.warning(f"Output not saved as output_fp = {output_fp}.")
        else:
            self.outputdf.to_excel(output_fp)
            logger.info(f"Output saved to {output_fp}.")
            

if __name__ == "__main__":
        
    if True:
        # Get the luna folderpath 
        luna_init_file = luna.__file__
        luna_folderpath = os.path.dirname(luna_init_file)
        logger.info(f"Your luna library is at {luna_folderpath}.")
        
        # Get the template folderpath
        template_folderpath = os.path.join(luna_folderpath, "templates")
        
        client_number = 40709
        engagement = 'ci'
    
        question_list = [
                "Enter the Account No for bad debts (NA if none): ",  
                "Enter the Account No for provision debts (NA if none): ", 
                "Enter Director's remuneration (0.00 if None): $ "]   
    
        varname_list = ["exp_bad_debts", 
                        "exp_prov_dtf_debts", 
                        "exp_dir_renum"
                        ]
        
        user_inputs = pd.DataFrame({"Question": question_list}, index = varname_list)
    
        user_inputs_currfy = user_inputs.copy()
        user_inputs_prevfy = user_inputs.copy()
    
        # CrossInvest
        input_dict_currfy_ci = {'exp_bad_debts'         : '1590054',
                                'exp_prov_dtf_debts'    : 'CW9',
                                'exp_dir_renum'         : '677754'
                                }
        input_dict_prevfy_ci = {'exp_bad_debts'         : '1590054',
                                'exp_prov_dtf_debts'    : 'CW9',
                                'exp_dir_renum'         : '819860'
                                }
        
        # MG
        input_dict_currfy_mg = {'exp_bad_debts'         : 'NA',
                                'exp_prov_dtf_debts'    : 'NA',
                                'exp_dir_renum'         : '748216'
                                }
        input_dict_prevfy_mg = {'exp_bad_debts'         : 'NA',
                                'exp_prov_dtf_debts'    : 'NA',
                                'exp_dir_renum'         : '792856'
                                }
        
        # ICM
        input_dict_currfy_icm = {'exp_bad_debts'         : 'NA',
                                 'exp_prov_dtf_debts'    : 'NA',
                                 'exp_dir_renum'         : '619423'
                                }
        input_dict_prevfy_icm = {'exp_bad_debts'         : 'NA',
                                 'exp_prov_dtf_debts'    : 'NA',
                                 'exp_dir_renum'         : '0'
                                }
        
        if engagement == 'mg':
            user_inputs_currfy["Answer"] = [input_dict_currfy_mg[var] for var in user_inputs_currfy.index]
            user_inputs_prevfy["Answer"] = [input_dict_prevfy_mg[var] for var in user_inputs_prevfy.index]
            tb_fp = r"P:\YEAR 2023\TECHNOLOGY\Technology users\FS Vertical\TB with updated LS codes\Myer Gold Investment Management - 2022 TB.xlsx"
            fy = 2022
        elif engagement == 'ci':
            user_inputs_currfy["Answer"] = [input_dict_currfy_ci[var] for var in user_inputs_currfy.index]
            user_inputs_prevfy["Answer"] = [input_dict_prevfy_ci[var] for var in user_inputs_prevfy.index]
            tb_fp = r"P:\YEAR 2023\TECHNOLOGY\Technology users\FS Vertical\TB with updated LS codes\Crossinvest TB reclassed.xlsx"
            fy = 2022
        elif engagement == 'icm':
            user_inputs_currfy["Answer"] = [input_dict_currfy_icm[var] for var in user_inputs_currfy.index]
            user_inputs_prevfy["Answer"] = [input_dict_prevfy_icm[var] for var in user_inputs_prevfy.index]
            tb_fp = r"P:\YEAR 2023\TECHNOLOGY\Technology users\FS Vertical\TB with updated LS codes\icm TB reformatted.xlsx"
            fy = 2023
        else:
            logger.warning("Engagement {engagement} is not recognised.")
    
        # # to use when reading from Alteryx
        # alteryx_fp = {'user_inputs_currfy': ui_currfy_fp,
        #               'user_inputs_prevfy': ui_prevfy_fp} # fp will be created when working on alteryx
        # user_inputs_currfy = pd.read_excel(alteryx_fp['user_inputs_currfy'])
        # user_inputs_prevfy = pd.read_excel(alteryx_fp['user_inputs_prevfy'])
    
    
        fp_dict = {'aged_receivables_fp'      : r"D:\Desktop\owgs\CODES\luna\personal_workspace\dacia\aged_receivables_template.xlsx",
                   'tb_fp'                    : tb_fp,
                   'sig_acct_output_fp'       : r"D:\Documents\Project\Internal Projects\20231206 Code review\acc_output.xlsx",
                   'sig_acct_prevfy_output_fp': r"D:\Documents\Project\Internal Projects\20231206 Code review\acc_output_prevfy.xlsx",
                   'output_fp'                : r"D:\Documents\Project\Internal Projects\20231206 Code review\form_3_output.xlsx"}
        

    # TB
    if False:
        #tb_fp = os.path.join(template_folderpath, "tb.xlsx")
        #tb_fp = r"D:\Desktop\owgs\CODES\luna\personal_workspace\dacia\Myer Gold Investment Management - 2022 TB.xlsx"
        
        tb_fp = fp_dict['tb_fp']
        logger.info(f"Your tb_filepath is at {tb_fp}.")
        
        # Load the tb
        fy_end_date = datetime.date(2022, 12, 31)
        tb_class = common.TBReader_ExcelFormat1(tb_fp, 
                                                sheet_name = 0,
                                                fy_end_date = fy_end_date)
        
        
        # Get data by fy
        # fy = 2022
        tb2022 = tb_class.get_data_by_fy(fy)
    if True:
        # Load tb class from LunaHub
        tb_class = common.TBLoader_From_LunaHub(client_number, fy)
        

    
    # Form 3 mapping 
    if True:
        
        mas_tb_mapping_fp = os.path.join(luna_folderpath, "parameters", "mas_forms_tb_mapping.xlsx")
        logger.info(f"Your mas_tb_mapping_fp is at {mas_tb_mapping_fp}.")
        
        # Load the class
        mapper_class = MASTemplateReader_Form3(mas_tb_mapping_fp, sheet_name = "Form 3 - TB mapping")
    
        # process df is here:
        df_processed = mapper_class.df_processed  # need to build methods
        


    # CLASS
    
    sig_acc_output_fp = fp_dict['sig_acct_output_fp']
    self = MASForm3_Generator(tb_class,
                              mapper_class,
                              sig_acc_output_fp,
                              client_number,
                              fy,
                              user_inputs = user_inputs_currfy
                              ) # is this an instance of a class?

    # COMMENT THIS PORTION OUT FOR ICM--#
    # Get previous balance
    prevfy = fy-1
    sig_acc_output_fp_prevfy = fp_dict['sig_acct_prevfy_output_fp']

    temp = MASForm3_Generator(tb_class,
                              mapper_class,
                              sig_acc_output_fp,
                              client_number,
                              fy=prevfy,
                              user_inputs = user_inputs_prevfy)
    self.outputdf['Previous Balance'] = temp.outputdf["Balance"]

    # Reorder columns by index 
    column_order = [0,1, 2, 3, 4, 5, 6, 7, 8, 9, 11, 10] 
    self.outputdf = self.outputdf.iloc[:, column_order]  
    # PORTION END--#

    # Output to excel 
    self.outputdf.to_excel(fp_dict['output_fp']) 



# MG
    # NA
    # NA
    # 748216

    # NA
    # NA
    # 792856

# CrossInvest
    # 1590054
    # CW9
    # 677754
    
    # 1590054
    # CW9 
    # 819860

# ICM
    # change fy to 2023
    # NA
    # NA
    # 619423
    
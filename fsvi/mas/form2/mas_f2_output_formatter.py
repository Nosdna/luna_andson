import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd
from copy import copy
from datetime import datetime
import re

import luna
from luna.fsvi.mas.template_reader import MASTemplateReader_Form1
from luna.lunahub import tables
import os

class OutputFormatter:

    COLUMN_MAPPER = {"target_ocr_amt_excelcol"      : "E",
                     "target_ocr_subtotal_excelcol" : "F",
                     "target_amt_excelcol"          : "G",
                     "target_subtotal_excelcol"     : "H",
                     "target_var_amt_excelcol"      : "I",
                     "target_var_subtotal_excelcol" : "J",
                     "target_ls_amt_excelcol"       : "K",
                     "target_ls_subtotal_excelcol"  : "L",
                     "target_varname_excelcol"      : "M"
                     }
    
    def __init__(self, input_fp, output_fp, ocr_fp, client_class, fy, aic_name = ""):

        self.input_fp       = input_fp
        self.output_fp      = output_fp
        self.ocr_fp         = ocr_fp
        self.client_class   = client_class
        self.fy             = fy
        self.aic_name       = aic_name
        
        # Get the template fp
        self.template_fp = os.path.join(
            luna.settings.LUNA_FOLDERPATH,
            "parameters", 
            "mas_forms_tb_mapping.xlsx")
        
        self.template_sheetname = "Form 2 - TB mapping"
        
        self.template_class = MASTemplateReader_Form1(
            self.template_fp, 
            self.template_sheetname)
        
        
        self.main()

    def main(self):

        self.read_files()
        self.write_output()

    def read_files(self):
        # Read the file with the values
        self.input_df = pd.read_excel(self.input_fp)
        self.ocr_df = pd.read_excel(self.ocr_fp)

    
    def build_varname_to_values(self, df):
        
        df = df.copy()
    
        varname_to_values = df[
            ["var_name", "Amount", "Subtotal"]
            ].dropna(subset=["var_name"])
        
        # Check that var_name is unique
        if varname_to_values["var_name"].is_unique:
            varname_to_values = varname_to_values.set_index("var_name")
        else:
            raise Exception ("Variable name is not unique.")
                    
        # self.varname_to_values = varname_to_values.copy()
        
        return varname_to_values
    
    def _load_client_info(self):

        self.client_name = self.client_class.retrieve_client_info("CLIENTNAME")

    
    def _copy_columns(self, ws,
                      src_excelcol, dst_excelcol,
                      src_col_name, dst_value = None
                      ):
        for src, dst in zip(ws[f"{src_excelcol}:{src_excelcol}"], ws[f"{dst_excelcol}:{dst_excelcol}"]):
            if src.value == src_col_name:
                dst.value = src_col_name
                dst.font = Font(bold = True)
                src.font = Font(bold = True)
            elif src.value is not None and dst_value is None:
                dst.value = src.value
                dst.border = copy(src.border)
            elif src.value is not None and dst_value is not None:
                dst.value = dst_value
                dst.border = copy(src.border)
            else:
                dst.value = None
                dst.border = copy(src.border)

    def _copy_column_style(self, ws,
                           src_excelcol, dst_excelcol
                           ):
        for src, dst in zip(ws[f"{src_excelcol}:{src_excelcol}"], ws[f"{dst_excelcol}:{dst_excelcol}"]):
            if src.value is not None:
                dst.border = copy(src.border)


    def _create_header(self, ws):

        # Unmerge all cells
        for merge in list(ws.merged_cells):
            ws.unmerge_cells(range_string=str(merge))

        ws.delete_rows(idx = 1, amount = 5)
        ws.insert_rows(idx = 0, amount = 7)

        # for idx in range(4, 6):
        #     ws.row_dimensions[idx].hidden = True #TODO: this should be temporary

        for merge in list(ws.merged_cells):
            ws.unmerge_cells(range_string=str(merge))

        row = 1

        # header title
        ws[f"A{row}"].value = "Form 2 - Statement of financial resources, total risk requirement and aggregate indebtedness"
        ws[f"A{row}"].font = Font(size = "16", bold = True)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="00C0C0C0")
        ws[f"A{row}"].border = Border(left   = Side(style = "thin"),
                                 right  = Side(style = "thin"),
                                 top    = Side(style = "thin"),
                                 bottom = Side(style = "thin")
                                 )
        ws.merge_cells(f"A{row}:F{row}")

        date_of_analysis = datetime.now().strftime("%d/%m/%Y")

        dct_of_fields = {"Client name"      : self.client_name,
                         "FY"               : self.fy,
                         "Date of analysis" : date_of_analysis,
                         "Prepared by"      : self.aic_name,
                         "Reviewed by"      : ""
                         }
        
        for key in dct_of_fields:
            
            row += 1

            self._create_header_row_field(key, dct_of_fields[key], row, ws)

    def _create_header_row_field(self, field_title, field_value, row, ws):

        border_style = Border(left   = Side(style = "thin"),
                              right  = Side(style = "thin"),
                              top    = Side(style = "thin"),
                              bottom = Side(style = "thin")
                              )
    
        ws[f"A{row}"].value = field_title
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="00C0C0C0")
        ws[f"A{row}"].font = Font(bold = True)
        ws[f"A{row}"].border = border_style
        ws.merge_cells(f"A{row}:C{row}")

        ws[f"D{row}"].value = field_value
        ws[f"D{row}"].fill = PatternFill("solid", fgColor="00FFFFFF")
        ws[f"D{row}"].alignment = Alignment(horizontal = 'left')
        ws[f"D{row}"].border = border_style
        ws.merge_cells(f"D{row}:F{row}")

    def _standardise_number_format(self, ws, lst_of_excelcols, row):
        for excelcol in lst_of_excelcols:
            ws[f"{excelcol}{row}"].number_format = '#,##0.00'

    def _replace_ls_null_value(self, ws, excelcol, row):
        cell = ws[f"{excelcol}{row}"]
        cell_value_str = str(cell.value)
        # print(f"cell: {cell} and {cell_value_str}")
        if cell_value_str is None:
            print(f"cell: {cell} and {cell_value_str}")
            cell.value = "<<<No L/S code assigned>>>"

    def _replace_ls_total_value(self, ws, excelcol, row):
        cell = ws[f"{excelcol}{row}"]
        cell_value_str = str(cell.value)

        if re.match("= MIN.*", cell_value_str):
            cell.value = "<<<Threshold result>>>"
        elif re.match("= .*100\%.*", cell_value_str):
            cell.value = "<<<Ratio>>>"
        elif re.match("=.*", cell_value_str):
            cell.value = "<<<Total>>>"
        else:
            pass

    def _create_var_formula(self, ws, excelcol, excelrow, val):
        
        cell = ws[f"{excelcol}{excelrow}"]

        var_target_excelcol = self._get_col_letter_from_ref(excelcol, 2)
        ocr_target_excelcol = self._get_col_letter_from_ref(excelcol, -2)
        if cell.value == val:
            ws[f"{var_target_excelcol}{excelrow}"].value = f"= {ocr_target_excelcol}{excelrow+2} - {excelcol}{excelrow+2}"
        # TODO : excelrow+2 is wrong

    def _get_col_letter_from_ref(self, ref_excelcol, mvmt):
        ref_colno = column_index_from_string(ref_excelcol)
        target_colno = ref_colno + int(mvmt)
        target_excelcol = get_column_letter(target_colno)

        return target_excelcol
    
    def _section_column_formatting(self, ws, section_name, starting_excelcol):
        section_color_dict = {'recalculated': '00CCFFFF',
                              'ocr'         : '00CCCCFF',
                              'var'         : '00CCFFCC'
                              }
        fill_setting = PatternFill("solid", fgColor = section_color_dict.get(section_name))
        ending_excelcol = self._get_col_letter_from_ref(starting_excelcol, 1)
        cols = ws[f"{starting_excelcol}6:{ending_excelcol}132"] #TODO: look into how to dynamically reference rows
        for cell in cols:
            cell[0].fill = fill_setting
            cell[1].fill = fill_setting

    def _create_col_header_1(self, ws, excelcol, row, header_name):
            cell = ws[f"{excelcol}{row}"]
            cell.value = header_name
            cell.font = Font(bold = True)
            cell.alignment = Alignment(horizontal = 'center')
            end_col = self._get_col_letter_from_ref(excelcol, 1)
            ws.merge_cells(f"{excelcol}{row}:{end_col}{row}")
            
    def write_output(self):
        templ_wb = openpyxl.load_workbook(self.template_fp)
        templ_ws = templ_wb[self.template_sheetname]

        self._load_client_info()
        
        sheets_to_remove = [sheet_name for sheet_name in templ_wb.sheetnames if sheet_name != self.template_sheetname]        
        
        for sheet_name in sheets_to_remove:
            templ_wb.remove(templ_wb[sheet_name])
        
        # Template as df
        colname_to_excelcol = self.template_class.colname_to_excelcol
        varname_to_index = self.template_class.varname_to_index
        
        # Get the data
        varname_to_values = self.build_varname_to_values(self.input_df)
        self.varname_to_values = varname_to_values.copy()

        # Get data from ocr
        varname_to_values_ocr = self.build_varname_to_values(self.ocr_df)
        self.varname_to_values_ocr = varname_to_values_ocr.copy()
        
        # Save column index
        amt_excelcol = colname_to_excelcol.at["Amount"]
        subtotal_excelcol = colname_to_excelcol.at["Subtotal"]
        varname_excelcol = colname_to_excelcol.at["var_name"]
        
        # Initialise excelcol
        excelcol_mapper = OutputFormatter.COLUMN_MAPPER
        target_varname_excelcol         = excelcol_mapper.get("target_varname_excelcol")
        target_ls_amt_excelcol          = excelcol_mapper.get("target_ls_amt_excelcol")
        target_ls_subtotal_excelcol     = excelcol_mapper.get("target_ls_subtotal_excelcol")
        target_ocr_amt_excelcol         = excelcol_mapper.get("target_ocr_amt_excelcol")
        target_ocr_subtotal_excelcol    = excelcol_mapper.get("target_ocr_subtotal_excelcol")
        target_var_amt_excelcol         = excelcol_mapper.get("target_var_amt_excelcol")
        target_var_subtotal_excelcol    = excelcol_mapper.get("target_var_subtotal_excelcol")
        target_amt_excelcol             = excelcol_mapper.get("target_amt_excelcol")
        target_subtotal_excelcol        = excelcol_mapper.get("target_subtotal_excelcol")

        # TODO: check on below
        for cell in templ_ws[f"{amt_excelcol}"]:
            if cell.value in ["'999999999", '999999999', 999999999]:
                cell.value = None
        for cell in templ_ws[f"{subtotal_excelcol}"]:
            if cell.value in ["'999999999", '999999999', 999999999]:
                cell.value = None

         # Copy to another column
        self._copy_columns(templ_ws,
                           varname_excelcol, target_varname_excelcol,
                           "var_name", dst_value = None)
        self._copy_columns(templ_ws,
                           amt_excelcol, target_amt_excelcol,
                           "Amount", dst_value = None)
        self._copy_columns(templ_ws,
                           subtotal_excelcol, target_subtotal_excelcol,
                           "Subtotal", dst_value = None)
        self._copy_columns(templ_ws,
                           amt_excelcol, target_ls_amt_excelcol,
                           "Amount", dst_value = None)
        self._copy_columns(templ_ws,
                           subtotal_excelcol, target_ls_subtotal_excelcol,
                           "Subtotal", dst_value = None)
        self._copy_columns(templ_ws,
                           amt_excelcol, target_ocr_amt_excelcol,
                           "Amount", dst_value = "")
        self._copy_columns(templ_ws,
                           subtotal_excelcol, target_ocr_subtotal_excelcol,
                           "Subtotal", dst_value = "")
        
        amt_excelcol = target_amt_excelcol
        subtotal_excelcol = target_subtotal_excelcol

        # filtered_varname_to_values = varname_to_values[varname_to_values.index.str.match(r"^total_.*")]

        varname_to_values_temp = varname_to_values.copy()
        varname_to_values_temp["Subtotal"] = varname_to_values["Subtotal"].astype(str)
        filtered_varname_to_values = varname_to_values_temp[~varname_to_values_temp["Subtotal"].str.contains("= SUM\(.*\)")]

        # Update amount and subtotal column with recalculated values
        for varname in filtered_varname_to_values.index:
            amt = varname_to_values.at[varname, "Amount"]
            subtotal = varname_to_values.at[varname, "Subtotal"]
            
            # Get the location to update
            row = varname_to_index.at[varname]
            
            # Update
            templ_ws[f"{amt_excelcol}{row}"].value = amt
            templ_ws[f"{subtotal_excelcol}{row}"].value = subtotal

            # Replace values declared with 999999999 with None instead
            self._replace_ls_null_value(templ_ws, target_ls_amt_excelcol, row)
            self._replace_ls_null_value(templ_ws, target_ls_subtotal_excelcol, row)

            # # TODO: temporary measure for H56
            # if templ_ws[f"{subtotal_excelcol}{row}"].value in [999999999, '999999999']:
            #     templ_ws[f"{subtotal_excelcol}{row}"].value = None
            # if templ_ws[f"{amt_excelcol}{row}"].value in [999999999, '999999999']:
            #     templ_ws[f"{amt_excelcol}{row}"].value = None

            # Create formula for variance column to compare values of client vs rsm
            self._create_var_formula(templ_ws, amt_excelcol, row, amt)
            self._create_var_formula(templ_ws, subtotal_excelcol, row, subtotal)

            # Update number format
            lst_of_excelcols = [amt_excelcol, subtotal_excelcol,
                                target_var_amt_excelcol, target_var_subtotal_excelcol,
                                target_ocr_amt_excelcol, target_ocr_subtotal_excelcol
                                ]
            self._standardise_number_format(templ_ws, lst_of_excelcols, row)

        for varname in varname_to_values_ocr.index:
            amt_ocr = varname_to_values_ocr.at[varname, "Amount"]
            subtotal_ocr = varname_to_values_ocr.at[varname, "Subtotal"]
            
            # Get the location to update
            row = varname_to_index.at[varname]
            
            # Update
            templ_ws[f"{target_ocr_amt_excelcol}{row}"].value = amt_ocr
            templ_ws[f"{target_ocr_subtotal_excelcol}{row}"].value = subtotal_ocr

        templ_ws[f"{target_var_amt_excelcol}7"].value = "Amount"
        templ_ws[f"{target_var_amt_excelcol}7"].font = Font(bold = True)
        templ_ws[f"{target_var_subtotal_excelcol}7"].value = "Subtotal"
        templ_ws[f"{target_var_subtotal_excelcol}7"].font = Font(bold = True)

        self._copy_column_style(templ_ws, amt_excelcol, target_var_amt_excelcol)
        self._copy_column_style(templ_ws, subtotal_excelcol, target_var_subtotal_excelcol)

        self._section_column_formatting(templ_ws, "recalculated", amt_excelcol)
        self._section_column_formatting(templ_ws, "ocr", target_ocr_amt_excelcol)
        self._section_column_formatting(templ_ws, "var", target_var_amt_excelcol)

        self._create_header(templ_ws)

        # adjust total and subtotal formulas
        # filtered_varname_to_values = varname_to_values[varname_to_values.index.str.match(r"^total_.*")]
        varname_to_values_temp = varname_to_values.copy()
        varname_to_values_temp["Subtotal"] = varname_to_values["Subtotal"].astype(str)
        lst_of_formula_varname = self.template_class.get_varname_to_formula().index.tolist()
        filtered_varname_to_values = varname_to_values_temp[varname_to_values_temp.index.isin(lst_of_formula_varname)]
        
        for varname in filtered_varname_to_values.index:

            MODIFIER = 2

            subtotal = varname_to_values.at[varname, "Subtotal"]

            row = varname_to_index.at[varname] + MODIFIER

            formula_subtotal_value = str(templ_ws[f"{target_ls_subtotal_excelcol}{row}"].value)

            # = SUM(A1:A5) OR = SUM(A1,A5)
            pattern = "^= SUM\(([A-Z]+)(\d+)\s*(.)\s*([A-Z]+)(\d+)\)$"
            formula_str = re.search(pattern, formula_subtotal_value)
            if formula_str is not None:
                ori_start_letter = formula_str.group(1)
                ori_start_row = formula_str.group(2)
                char = formula_str.group(3)
                ori_end_letter = formula_str.group(4)
                ori_end_row = formula_str.group(5)

                new_start_row = str(int(ori_start_row) + MODIFIER)
                new_end_row = str(int(ori_end_row) + MODIFIER)

                new_start_letter = self._get_col_letter_from_ref(ori_start_letter, 2)
                new_end_letter = self._get_col_letter_from_ref(ori_end_letter, 2)

                new_formula = f"= SUM({new_start_letter}{new_start_row}{char}{new_end_letter}{new_end_row})"

                templ_ws[f'{subtotal_excelcol}{row}'].value = new_formula

            # = SUM(A1:A5) - SUM(A10:A15) OR = SUM(A1:A5) + SUM(A10:A15) OR  SUM(A1,A5) + SUM(A10:A15) ETC...
            pattern = "^= SUM\(([A-Z]+)(\d+)\s*(.)\s*([A-Z]+)(\d+)\)\s*(.)\s*SUM\(([A-Z]+)(\d+)\s*(.)\s*([A-Z]+)(\d+)\)$"
            formula_str = re.search(pattern, formula_subtotal_value)
            if formula_str is not None:
                ori_start_letter = formula_str.group(1)
                ori_start_row = formula_str.group(2)
                char = formula_str.group(3)
                ori_end_letter = formula_str.group(4)
                ori_end_row = formula_str.group(5)
                char2 = formula_str.group(6)
                ori_start_letter2 = formula_str.group(7)
                ori_start_row2 = formula_str.group(8)
                char3 = formula_str.group(9)
                ori_end_letter2 = formula_str.group(10)
                ori_end_row2 = formula_str.group(11)

                new_start_row = str(int(ori_start_row) + MODIFIER)
                new_end_row = str(int(ori_end_row) + MODIFIER)
                new_start_row2 = str(int(ori_start_row2) + MODIFIER)
                new_end_row2 = str(int(ori_end_row2) + MODIFIER)

                new_start_letter = self._get_col_letter_from_ref(ori_start_letter, 2)
                new_end_letter = self._get_col_letter_from_ref(ori_end_letter, 2)
                new_start_letter2 = self._get_col_letter_from_ref(ori_start_letter2, 2)
                new_end_letter2 = self._get_col_letter_from_ref(ori_end_letter2, 2)

                new_formula = f"= SUM({new_start_letter}{new_start_row}{char}{new_end_letter}{new_end_row}){char2}SUM({new_start_letter2}{new_start_row2}{char3}{new_end_letter2}{new_end_row2})"

                templ_ws[f'{subtotal_excelcol}{row}'].value = new_formula

            # = A1 - SUM(A10:A15)
            pattern = "^= ([A-Z]+)(\d+)\s*(.)\s*SUM\(([A-Z]+)(\d+)\s*(.)\s*([A-Z]+)(\d+)\)$"
            formula_str = re.search(pattern, formula_subtotal_value)
            if formula_str is not None:
                ori_net_letter = formula_str.group(1)
                ori_net_row = formula_str.group(2)
                net_char =  formula_str.group(3)
                ori_start_letter = formula_str.group(4)
                ori_start_row = formula_str.group(5)
                char = formula_str.group(6)
                ori_end_letter = formula_str.group(7)
                ori_end_row = formula_str.group(8)

                new_net_row = str(int(ori_net_row) + MODIFIER)
                new_start_row = str(int(ori_start_row) + MODIFIER)
                new_end_row = str(int(ori_end_row) + MODIFIER)

                new_net_letter = self._get_col_letter_from_ref(ori_net_letter, 2)
                new_start_letter = self._get_col_letter_from_ref(ori_start_letter, 2)
                new_end_letter = self._get_col_letter_from_ref(ori_end_letter, 2)

                new_formula = f"= {new_net_letter}{new_net_row} {net_char} SUM({new_start_letter}{new_start_row}{char}{new_end_letter}{new_end_row})"
    
                templ_ws[f'{subtotal_excelcol}{row}'].value = new_formula

            # = SUM(A10:A15) - A20
            pattern = "^= SUM\(([A-Z]+)(\d+)\s*(.)\s*([A-Z]+)(\d+)\)\s*(.)\s*([A-Z]+)(\d+)$"
            formula_str = re.search(pattern, formula_subtotal_value)
            if formula_str is not None:
                ori_start_letter = formula_str.group(1)
                ori_start_row = formula_str.group(2)
                char = formula_str.group(3)
                ori_end_letter = formula_str.group(4)
                ori_end_row = formula_str.group(5)
                net_char =  formula_str.group(6)
                ori_net_letter = formula_str.group(7)
                ori_net_row = formula_str.group(8)

                new_net_row = str(int(ori_net_row) + MODIFIER)
                new_start_row = str(int(ori_start_row) + MODIFIER)
                new_end_row = str(int(ori_end_row) + MODIFIER)

                new_net_letter = self._get_col_letter_from_ref(ori_net_letter, 2)
                new_start_letter = self._get_col_letter_from_ref(ori_start_letter, 2)
                new_end_letter = self._get_col_letter_from_ref(ori_end_letter, 2)

                new_formula = f"= SUM({new_start_letter}{new_start_row}{char}{new_end_letter}{new_end_row}) {net_char} {new_net_letter}{new_net_row}"

                templ_ws[f'{subtotal_excelcol}{row}'].value = new_formula

            # = A1
            pattern = "^= ([A-Z]+)(\d+)$"
            formula_str = re.search(pattern, formula_subtotal_value)
            if formula_str is not None:
                ori_start_letter = formula_str.group(1)
                ori_start_row = formula_str.group(2)

                new_start_letter = self._get_col_letter_from_ref(ori_start_letter, 2)

                new_start_row = str(int(ori_start_row) + MODIFIER)

                new_formula = f"= {new_start_letter}{new_start_row}"

                templ_ws[f'{subtotal_excelcol}{row}'].value = new_formula

            # = A1/A10 * 100%
            pattern = "^= ([A-Z]+)(\d+)\s*\/\s*([A-Z]+)(\d+)\s*\*\s*100$"
            formula_str = re.search(pattern, formula_subtotal_value)
            if formula_str is not None:
                ori_num_letter = formula_str.group(1)
                ori_num_row = formula_str.group(2)
                ori_denom_letter = formula_str.group(3)
                ori_denom_row = formula_str.group(4)

                new_num_row = str(int(ori_num_row) + MODIFIER)
                new_denom_row = str(int(ori_denom_row) + MODIFIER)

                new_num_letter = self._get_col_letter_from_ref(ori_num_letter, 2)
                new_denom_letter = self._get_col_letter_from_ref(ori_denom_letter, 2)

                new_formula = f"= {new_num_letter}{new_num_row} / {new_denom_letter}{new_denom_row} * 100"

                templ_ws[f'{subtotal_excelcol}{row}'].value = new_formula

            # = SUM(A1,A2,A3,A4,A5)
            pattern = "^= SUM\(([A-Z]+)(\d+)\s*,\s*([A-Z]+)(\d+)\s*,\s*([A-Z]+)(\d+)\s*,\s*([A-Z]+)(\d+)\s*,\s*([A-Z]+)(\d+)\)$"
            formula_str = re.search(pattern, formula_subtotal_value)
            if formula_str is not None:
                ori_letter1 = formula_str.group(1)
                ori_row1 = formula_str.group(2)
                ori_letter2 = formula_str.group(3)
                ori_row2 = formula_str.group(4)
                ori_letter3 = formula_str.group(5)
                ori_row3 = formula_str.group(6)
                ori_letter4 = formula_str.group(7)
                ori_row4 = formula_str.group(8)
                ori_letter5 = formula_str.group(9)
                ori_row5 = formula_str.group(10)

                new_row1 = str(int(ori_row1) + MODIFIER)
                new_row2 = str(int(ori_row2) + MODIFIER)
                new_row3 = str(int(ori_row3) + MODIFIER)
                new_row4 = str(int(ori_row4) + MODIFIER)
                new_row5 = str(int(ori_row5) + MODIFIER)

                new_letter1 = self._get_col_letter_from_ref(ori_letter1, 2)
                new_letter2 = self._get_col_letter_from_ref(ori_letter2, 2)
                new_letter3 = self._get_col_letter_from_ref(ori_letter3, 2)
                new_letter4 = self._get_col_letter_from_ref(ori_letter4, 2)
                new_letter5 = self._get_col_letter_from_ref(ori_letter5, 2)

                new_formula = f"= SUM({new_letter1}{new_row1}, {new_letter2}{new_row2}, {new_letter3}{new_row3}, {new_letter4}{new_row4}, {new_letter5}{new_row5})"

                templ_ws[f'{subtotal_excelcol}{row}'].value = new_formula

            # = MIN(A1 * 100, 10000000)
            pattern = "^= MIN\(([A-Z]+)(\d+)\s\*\s*100.\s*10000000\)$"
            formula_str = re.search(pattern, formula_subtotal_value)
            if formula_str is not None:
                ori_start_letter = formula_str.group(1)
                ori_start_row = formula_str.group(2)

                new_start_letter = self._get_col_letter_from_ref(ori_start_letter, 2)
                new_start_row = str(int(ori_start_row) + MODIFIER)

                new_formula = f"= MIN({new_start_letter}{new_start_row} * 100, 10000000)"

                templ_ws[f'{subtotal_excelcol}{row}'].value = new_formula

            # Replace values declared with L/S with <<<>>> indicator instead
            self._replace_ls_total_value(templ_ws, target_ls_subtotal_excelcol, row)


        # titles
        row = 8 #TODO: should not hardcode
        self._create_col_header_1(templ_ws, target_ls_amt_excelcol, row, "L/S")
        self._create_col_header_1(templ_ws, target_ocr_amt_excelcol, row, "Client")
        self._create_col_header_1(templ_ws, target_amt_excelcol, row, "RSM")
        self._create_col_header_1(templ_ws, target_var_amt_excelcol, row, "Variance")

       
        # conditional formatting
        redFill = PatternFill(start_color='EE1111',
                              end_color='EE1111',
                              fill_type='solid')
        templ_ws.conditional_formatting.add(f"{target_var_amt_excelcol}10:{target_var_subtotal_excelcol}131",
                                      CellIsRule(operator='greaterThan', formula=['0.01'], stopIfTrue=True, fill=redFill))
        templ_ws.conditional_formatting.add(f"{target_var_amt_excelcol}10:{target_var_subtotal_excelcol}131",
                                      CellIsRule(operator='lessThan', formula=['-0.01'], stopIfTrue=True, fill=redFill))
        

        templ_ws.column_dimensions[target_varname_excelcol].hidden = True
        templ_ws.title = "Form 2 (Recalculated)"

        templ_wb.save(self.output_fp)
        templ_wb.close()

if __name__ == "__main__":

    if True: # Testing
        client_no   = 7167
        fy          = 2022

        #input_fp    = r"D:\workspace\luna\personal_workspace\tmp\mas_form1_7167_2022.xlsx"
        #template_fp = r"D:\workspace\luna\parameters\mas_forms_tb_mapping.xlsx"
        #output_fp   = r"D:\workspace\luna\personal_workspace\tmp\mas_form1_formatted_71679_2022.xlsx"

        input_fp = rf"D:\workspace\luna\personal_workspace\tmp\mas_form2_{client_no}_{fy}.xlsx"
        ocr_fp = rf"D:\workspace\luna\personal_workspace\tmp\mas_form2_{client_no}_{fy}_ocr.xlsx"
        output_fp = rf"D:\workspace\luna\personal_workspace\tmp\mas_form2_formatted_{client_no}_{fy}.xlsx"
        
        client_class = tables.client.ClientInfoLoader_From_LunaHub(client_no)
        
        aic_name = "John Smith"
        
        self = OutputFormatter(input_fp     = input_fp,
                               output_fp    = output_fp,
                               ocr_fp       = ocr_fp,
                               client_class = client_class,
                               fy           = fy,
                               aic_name     = aic_name
                               )
        
    if True:
        import webbrowser
        webbrowser.open(output_fp)

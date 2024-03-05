import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd
import numpy as np
from copy import copy
from datetime import datetime
import re

import luna
from luna.fsvi.funds.invtmt_report_template_reader import FundsInvtmtTemplateReader
import luna.common as common
from luna.lunahub import tables
import os

import pyeasylib.excellib as excellib

class InvtmtOutputFormatter:

    LSCODES_NAV         = [pd.Interval(6900.0, 6900.4, closed='both')]
    LSCODES_BOND_INT    = [pd.Interval(7400.2, 7400.2, closed='both')]
    LSCODES_BOND_INTREC = [pd.Interval(5200.0, 5300.0, closed='left')]

    CONFIDENCE_THRESHOLD = 0.7

    def __init__(self, sublead_class, portfolio_class, recon_class, tb_class,
                 output_fp, mapper_fp, user_inputs,
                 client_class, fy, aic_name = ""
                 ):

        self.sublead_class  = sublead_class
        self.portfolio_class= portfolio_class
        self.recon_class    = recon_class
        self.tb_class       = tb_class
        self.output_fp      = output_fp
        self.mapper_fp      = mapper_fp
        self.user_inputs    = user_inputs
        self.client_class   = client_class
        self.fy             = int(fy)
        self.aic_name       = aic_name

        self.template_fp = os.path.join(luna.settings.LUNA_FOLDERPATH,
                                        "parameters",
                                        "investment_template.xlsx"
                                        )
        
        self.template_class = FundsInvtmtTemplateReader(self.template_fp)

        self.main()

    def main(self):

        self.get_data()
        self.write_sublead_output()
        self.write_recon_output()
        self.write_portfolio_output()


    def get_data(self):
        self.sublead_input_df = self.sublead_class.main()
        self.portfolio_input_df = self.portfolio_class.main()
        self.recon_input_df_detail = self.recon_class.main()
        self.portfolio_mapper_df = pd.read_excel(self.mapper_fp)
            
    def build_varname_to_values(self, df):
        
        df = df.copy()
    
        varname_to_values = df[
            ["VARNAME", "VALUE", "VALUEPREVFY"]
            ].dropna(subset=["VARNAME"])
        
        # Check that var_name is unique
        if varname_to_values["VARNAME"].is_unique:
            varname_to_values = varname_to_values.set_index("VARNAME")
        else:
            raise Exception ("Variable name is not unique.")
                    
        # self.varname_to_values = varname_to_values.copy()
        
        return varname_to_values

    def _load_client_info(self):

        self.client_name = self.client_class.retrieve_client_info("CLIENTNAME")
    
    def _standardise_number_format(self, ws, lst_of_excelcols, row):
        for excelcol in lst_of_excelcols:
            ws[f"{excelcol}{row}"].number_format = '#,##0.00'

    def _standardise_date_format(self, ws, lst_of_excelcols, row):
        for excelcol in lst_of_excelcols:
            ws[f"{excelcol}{row}"].number_format = 'DD/MM/YYYY'

    def _create_header(self, ws, title, del_row_no, add_row_no):

        # Unmerge all cells
        for merge in list(ws.merged_cells):
            ws.unmerge_cells(range_string=str(merge))

        if del_row_no != 0:
            ws.delete_rows(idx = 1, amount = del_row_no)
        ws.insert_rows(idx = 0, amount = 6 + add_row_no)

        for merge in list(ws.merged_cells):
            ws.unmerge_cells(range_string=str(merge))

        row = 1

        # header title
        ws[f"A{row}"].value = title
        ws[f"A{row}"].font = Font(size = "16", bold = True)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="00C0C0C0")
        ws[f"A{row}"].border = Border(left   = Side(style = "thin"),
                                 right  = Side(style = "thin"),
                                 top    = Side(style = "thin"),
                                 bottom = Side(style = "thin")
                                 )
        ws.merge_cells(f"A{row}:F{row}")

        date_of_analysis = datetime.now().strftime("%d/%m/%Y")

        dct_of_fields = {"Client name"      : self.client_name,
                         "FY"               : self.fy,
                         "Date of analysis" : date_of_analysis,
                         "Prepared by"      : self.aic_name,
                         "Reviewed by"      : ""
                         }
        
        for key in dct_of_fields:
            
            row += 1

            self._create_header_row_field(key, dct_of_fields[key], row, ws)

    def _create_header_row_field(self, field_title, field_value, row, ws):

        border_style = Border(left   = Side(style = "thin"),
                              right  = Side(style = "thin"),
                              top    = Side(style = "thin"),
                              bottom = Side(style = "thin")
                              )
    
        ws[f"A{row}"].value = field_title
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="00C0C0C0")
        ws[f"A{row}"].font = Font(bold = True)
        ws[f"A{row}"].border = border_style
        ws.merge_cells(f"A{row}:C{row}")

        ws[f"D{row}"].value = field_value
        ws[f"D{row}"].fill = PatternFill("solid", fgColor="00FFFFFF")
        ws[f"D{row}"].alignment = Alignment(horizontal = 'left')
        ws[f"D{row}"].border = border_style
        ws.merge_cells(f"D{row}:F{row}")

    def _format_header_cell(self, col_lst, row, ws):

        border_style = Border(left   = Side(style = "thin"),
                              right  = Side(style = "thin"),
                              top    = Side(style = "thin"),
                              bottom = Side(style = "thin")
                              )

        for col in col_lst:
            ws[f"{col}{row}"].fill = PatternFill("solid", fgColor="00C0C0C0")
            ws[f"{col}{row}"].font = Font(bold = True)
            ws[f"{col}{row}"].border = border_style
        
    def _standardise_cell_format(self, ws, excelcol, row):
        font_style = Font(size = "8", name = "Arial")
        alignment_style =  Alignment(horizontal = 'center')
        ws[f"{excelcol}{row}"].font = font_style
        ws[f"{excelcol}{row}"].alignment = alignment_style
    
    def _populate_portfolio_formula(self, ws, colname_to_excelcol, col, row, formula):
        ws[f"{colname_to_excelcol[col]}{row}"].value = formula
        ws[f"{colname_to_excelcol[col]}{row}"].alignment = Alignment(horizontal = 'left')
        self._standardise_cell_format(ws, colname_to_excelcol[col], row)

    def _adjust_col_width(self, ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # Get column letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length)  # Add some padding
            ws.column_dimensions[column_letter].width = adjusted_width

    def _copy_column_style(self, ws,
                           src_excelcol, src_excelrow,
                           dst_excelcol, dst_excelrow
                           ):
        for src, dst in zip(ws[f"{src_excelcol}{src_excelrow}"], ws[f"{dst_excelcol}{dst_excelrow}"]):
            dst.fill = copy(src.fill)
    
    def map_portfolio_columns(self):
        
        self.portfolio_mapper_df = self.portfolio_mapper_df[~self.portfolio_mapper_df["Standardised"].isna()]
        self.portfolio_mapper_dict = dict(zip(self.portfolio_mapper_df['Standardised'], self.portfolio_mapper_df['Formatted']))
        self.processed_portfolio_input_df = self.portfolio_input_df.rename(columns = self.portfolio_mapper_dict).reset_index(drop = True)
    
    def update_custodian_confidence_level_indicator(self):

        col = "Notes for reconciliation"
        df = self.processed_portfolio_input_df.copy()
        threshold = InvtmtOutputFormatter.CONFIDENCE_THRESHOLD

        # cond = [(df[col] >= threshold), (df[col] < threshold)]
        # res = [f"Confidence of '{df[col]}': Fuzzy matched", f"Confidence of '{df[col]}': Need investigation"]

        # self.processed_portfolio_input_df[col] = np.select(cond, res, default='')

        self.processed_portfolio_input_df[col] = df[col].apply(lambda x: f"Confidence of '{x}': Fuzzy matched" if x >= threshold else f"Confidence of '{df[col]}': Need investigation")


    def process_nav(self):

        is_overlap, true_match, false_match = self.tb_class.filter_tb_by_fy_and_ls_codes(self.fy, InvtmtOutputFormatter.LSCODES_NAV)

        total_equity = true_match['Value'].sum()

        return total_equity

    def filter_tb_for_bond_int(self):

        is_overlap, true_match, false_match = self.tb_class.filter_tb_by_fy_and_ls_codes(self.fy, InvtmtOutputFormatter.LSCODES_BOND_INT)

        filtered_tb = true_match.copy()

        filtered_tb['Include / Exclude'] = filtered_tb['Name'].apply(lambda x: 'Included' if re.match('(?i).*interest.*', x) else 'Excluded')

        return filtered_tb.set_index('Account No').copy()
    
    def filter_tb_for_bond_intrec(self):

        is_overlap, true_match, false_match = self.tb_class.filter_tb_by_fy_and_ls_codes(self.fy, InvtmtOutputFormatter.LSCODES_BOND_INTREC)

        filtered_tb = true_match.copy()

        filtered_tb['Include / Exclude'] = filtered_tb['Name'].apply(lambda x: 'Included' if re.match('(?i).*interest.*', x) else 'Excluded')

        return filtered_tb.set_index('Account No').copy()
    
    def create_hyperlink_in_sublead_to_tb(self, wb, source_sheetname, cell, reference_sheetname, title, filtered_tb):

        field_to_source_locations = {"<<<link>>>": cell}

        header_rows = ["Title", "Client name", "Data as of", "FY", 
                       "Prepared by", "Reviewed by"]

        field_to_data = {"<<<link>>>" : filtered_tb}

        hyperlink_class = common.hyperlinks.DataHyperlink(source_sheetname, field_to_source_locations, reference_sheetname, header_rows, field_to_data, wb)

        hyperlink_class.write_reference_data()
        
        hyperlink_class.write_source_data()

        ws = wb[reference_sheetname]

        self._create_header(wb[reference_sheetname], title, 6, 0)

        for c_idx, header_value in enumerate(filtered_tb.reset_index().columns, 1):
            col = openpyxl.utils.cell.get_column_letter(c_idx)
            self._format_header_cell([col], 9, ws)
        
    def write_sublead_output(self):

        sheet_name = "<5100-xx>Investment sub-lead"

        templ_wb = openpyxl.load_workbook(self.template_fp)
        templ_ws = templ_wb[sheet_name]

        self._load_client_info()

        # template as df
        colname_to_excelcol = self.template_class.sublead_colname_to_excelcol
        varname_to_index = self.template_class.sublead_varname_to_index

        # additional processing for sublead
        varname_to_index_df = varname_to_index.to_frame().reset_index()
        filtered_varname_to_index_df = varname_to_index_df[varname_to_index_df['var_name'].str.match(r".*///.*")]
        for i in range(len(filtered_varname_to_index_df)):
            i_idx = filtered_varname_to_index_df.index[i]
            varname_to_index_df.at[i_idx, 'var_name'] = filtered_varname_to_index_df.at[i_idx, 'var_name'].split(' /// ')
        varname_to_index = varname_to_index_df.explode('var_name').set_index('var_name').squeeze()
                
        # get the data
        varname_to_values = self.build_varname_to_values(self.sublead_input_df)
        self.varname_to_values = varname_to_values.copy()

        # save column index
        varname_excelcol = colname_to_excelcol.at["var_name"]
        cfy_excelcol = colname_to_excelcol.at["Current FY"]
        pfy_excelcol = colname_to_excelcol.at["Previous FY"]

        for varname in varname_to_values.index:
            cfy = varname_to_values.at[varname, "VALUE"]
            pfy = varname_to_values.at[varname, "VALUEPREVFY"]

            row = varname_to_index.at[varname]

            templ_ws[f"{cfy_excelcol}{row}"].value = cfy
            templ_ws[f"{pfy_excelcol}{row}"].value = pfy

            self._standardise_number_format(templ_ws,
                                            [cfy_excelcol, pfy_excelcol],
                                            row
                                            )
            self._standardise_cell_format(templ_ws, cfy_excelcol, row)
            self._standardise_cell_format(templ_ws, pfy_excelcol, row)

        fy_excelrow = min(varname_to_index) - 3
        fcy_lst = self.sublead_input_df["FUNCTIONALCURRENCY"].to_list()
        if len(set(fcy_lst)) > 1:
            raise Exception("More than one unique value found in FUNCTIONALCURRENCY column:"
                            f"{fcy_lst}.")
        else:
            fcy = fcy_lst[0]

        templ_ws[f"{cfy_excelcol}{fy_excelrow}"].value = self.fy
        templ_ws[f"{pfy_excelcol}{fy_excelrow}"].value = self.fy-1
        templ_ws[f"{cfy_excelcol}{fy_excelrow+1}"].value = fcy
        templ_ws[f"{pfy_excelcol}{fy_excelrow+1}"].value = fcy
            
        
        self._create_header(templ_ws, sheet_name, 6, 0)

        templ_ws.column_dimensions[varname_excelcol].hidden = True

        filtered_tb_bond_int = self.filter_tb_for_bond_int().drop(["L/S (interval)"], axis = 1)
        filtered_tb_bond_intrec = self.filter_tb_for_bond_intrec().drop(["L/S (interval)"], axis = 1)

        self.create_hyperlink_in_sublead_to_tb(templ_wb, sheet_name, "E39",
                                               "Interest - Bonds TB",
                                               "Trial Balance for Interest - Bonds",
                                               filtered_tb_bond_int)
        self.create_hyperlink_in_sublead_to_tb(templ_wb, sheet_name, "E41",
                                               "Interest Receivables - Bonds TB", 
                                               "Trial Balance for Interest Receivables - Bonds",
                                               filtered_tb_bond_intrec)

        templ_ws["E39"].fill = copy(templ_ws["D39"].fill)
        templ_ws["E41"].fill = copy(templ_ws["D41"].fill)

        templ_wb.save(self.output_fp)
        templ_wb.close()

    def write_recon_output(self):

        summary_sheet_name = "Investment recon summary"
        detail_sheet_name = "Investment recon detail"

        templ_wb = openpyxl.load_workbook(self.output_fp)

        # summary
        templ_ws = templ_wb[summary_sheet_name]

        # summary content writing
        self.recon_input_df_summary = self.recon_input_df_detail.groupby(by="EXCEPTIONINDICATOR").agg(
            # MATCHINGINDICATOR       = ('MATCHINGINDICATOR', lambda x: ','.join(x.unique())),
            EXCEPTIONINDICATOR      = ('EXCEPTIONINDICATOR', lambda x: ','.join(x.unique())),
            MARKETVALUEFUNDADMIN    = ("MARKETVALUEFUNDADMIN", 'sum'),
            MARKETVALUEBROKER       = ('MARKETVALUEBROKER', 'sum'),
            VALUEDIFFERENCE         = ("VALUEDIFFERENCE", 'sum')
            )
        for c_idx, header_value in enumerate(self.recon_input_df_summary.columns, 1):
            templ_ws.cell(row=1, column=c_idx, value=header_value)
            col = openpyxl.utils.cell.get_column_letter(c_idx)
            self._format_header_cell([col], 1, templ_ws)

        for r_idx, row in enumerate(self.recon_input_df_summary.values, 2):
            for c_idx, value in enumerate(row, 1):
                templ_ws.cell(row=r_idx, column=c_idx, value=value)
            self._standardise_number_format(templ_ws, ['B', 'C', 'D', 'E', 'F', 'G'], r_idx) #TODO: hardcoded excelcols
        
        self._create_header(templ_ws, summary_sheet_name, 0, 1)

        # detail
        templ_ws = templ_wb[detail_sheet_name]

        self.recon_input_df_detail = self.recon_input_df_detail.drop(['CLIENTNUMBER', 'FY', 'UPLOADER',
                                                                     'UPLOADDATETIME', 'COMMENT1', 'COMMENT2',
                                                                     'COMMENT3'
                                                                     ], axis = 1)

        # detail content writing
        for c_idx, header_value in enumerate(self.recon_input_df_detail.columns, 1):
            templ_ws.cell(row=1, column=c_idx, value=header_value)
            col = openpyxl.utils.cell.get_column_letter(c_idx)
            self._format_header_cell([col], 1, templ_ws)

        for r_idx, row in enumerate(self.recon_input_df_detail.values, 2):
            for c_idx, value in enumerate(row, 1):
                templ_ws.cell(row=r_idx, column=c_idx, value=value)
            self._standardise_number_format(templ_ws, ['B', 'H', 'N', 'O'], r_idx) #TODO: hardcoded excelcols
        
        self._create_header(templ_ws, detail_sheet_name, 0, 1)


        templ_wb.save(self.output_fp)
        templ_wb.close()

    def write_portfolio_output(self):

        sheet_name = "<5100-xx>Investment Portfolio"

        templ_wb = openpyxl.load_workbook(self.output_fp)
        templ_ws = templ_wb[sheet_name]

        self._load_client_info()

        colname_to_excelcol = self.template_class.portfolio_colname_to_excelcol

        self.map_portfolio_columns()
        self.update_custodian_confidence_level_indicator()

        self._create_header(templ_ws, sheet_name, 2, 0)

        row = 14
        templ_ws.merge_cells(f"A{row}:H{row}") # per client
        templ_ws.merge_cells(f"I{row}:L{row}") # per custodian
        templ_ws.merge_cells(f"M{row}:R{row}") # per client
        templ_ws.merge_cells(f"S{row}:Y{row}") # per rsm
        templ_ws.merge_cells(f"AA{row}:AD{row}") # per rsm
        templ_ws.merge_cells(f"AF{row}:AI{row}") # if ltp is not within bis ask spread
        
        input_length = len(self.processed_portfolio_input_df)
        if input_length > 25:
            templ_ws.insert_rows(idx = 40, amount = input_length - 25 + 2)


        transposed_df = self.processed_portfolio_input_df.T

        for col in transposed_df.index:
            row = 13 + 5
            for i in range(len(transposed_df.columns)):
                try:
                    val = transposed_df.at[col, i]
                    templ_ws[f"{colname_to_excelcol[col]}{row}"].value = val
                    self._standardise_cell_format(templ_ws, colname_to_excelcol[col], row)
                    row += 1
                except:
                    pass
                self._standardise_number_format(templ_ws,
                                                ['J', 'K', 'M', 'O', 'P', 'Q',
                                                 'R', 'S', 'U', 'V', 'W', 'AA',
                                                 'AB', 'AC', 'AF', 'AG', 'AH',
                                                 'AI'], # TODO: hardcoded
                                                row
                                                )
                self._standardise_date_format(templ_ws, ['H'], row) # TODO: format not showing

        row = 13 + 5

        # # this will be KIV as a later enhancement
        # portfolio_formulas = self.template_class.portfolio_df_processed.iloc[0,:]
        # portfolio_formulas = portfolio_formulas[~portfolio_formulas.isna()]
        # portfolio_formulas_colname = portfolio_formulas
        # portfolio_formulas = portfolio_formulas.reset_index().set_index(12) #TODO: col reference hardcoded
        # mapping = portfolio_formulas.to_dict()['index']
        # col = 'Market Value at Last Trade Price (Base)'
        # formula = portfolio_formulas_colname[col]
        # # Regular expression pattern to match variable names
        # pattern = re.compile(r'\(\w\)')
        # # Iterate over each match and replace it with corresponding value
        # for var in pattern.findall(formula):
        #     formula = re.sub(re.escape(var), mapping.get(var, var), formula)
        #     print(formula)


        for i in range(len(transposed_df.columns)):

            # create formula for diff in holdings
            col = 'Diff in holdings?'
            diff_in_holdings = f"= {colname_to_excelcol['Holdings per confirmation @']}{row} - {colname_to_excelcol['Holdings']}{row}"
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, diff_in_holdings)

            # create formula for market value at last trade price (base)
            col = 'Market Value at Last Trade Price (Base)'
            mv_at_ltp_fcy = f"= {colname_to_excelcol['Holdings']}{row} * {colname_to_excelcol['Last Trade Price per unit (Local Currency)']}{row} * {colname_to_excelcol['Exchange Rate @']}{row}"
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, mv_at_ltp_fcy)

            # create formula for % of nav
            col = r'% of NAV'
            percent_of_nav = f"= {colname_to_excelcol['Market Value at Last Trade Price (Base)']}{row} / B9"
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, percent_of_nav)

            # create formula for market value per rsm (base)
            col = 'Market Value per RSM (Base)'
            mv_per_rsm_fcy = f"= {colname_to_excelcol['Holdings']}{row} * {colname_to_excelcol['Exchange Rate @']}{row} *  {colname_to_excelcol['Price Obtained from']}{row} "
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, mv_per_rsm_fcy)

            # create formula for diff in value (base)
            col = 'Diff in Value (Base)'
            diff_in_val_fcy = f"= {colname_to_excelcol['Market Value per RSM (Base)']}{row} - {colname_to_excelcol['Market Value at Last Trade Price (Base)']}{row}"
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, diff_in_val_fcy)

            # create formula for as a % of NAV
            col = r'As a % of NAV'
            percent_of_nav = f"= {colname_to_excelcol['Diff in Value (Base)']}{row} / B9"
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, percent_of_nav)

            # create formula for exception y/n
            col = r'Exception (Y/N)'
            exception_1 = f'= IF({colname_to_excelcol["Diff in Value (Base)"]}{row} = 0, "N", "Y")'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, exception_1)

            # create formula for lvl hierarchy
            col = 'Level hierarchy'
            lvl_hierarchy = f'= IF(ISBLANK({colname_to_excelcol["Price Obtained from"]}{row}), "", 1)'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, lvl_hierarchy)

            # create formula for ltp between range?
            col = 'Last Trade price between Bid/Ask range?\n(between / not between)'
            min_con_formula = f"{colname_to_excelcol['Last Trade Price per unit (Local Currency)']}{row}>=MIN({colname_to_excelcol['Bid Price Obtained from']}{row}, {colname_to_excelcol['Ask Price Obtained from']}{row})"
            max_con_formula = f"{colname_to_excelcol['Last Trade Price per unit (Local Currency)']}{row}<=MAX({colname_to_excelcol['Bid Price Obtained from']}{row}, {colname_to_excelcol['Ask Price Obtained from']}{row})"
            ltp_btw_bid_ask = f'= IF(AND({min_con_formula},{max_con_formula}), "Between", "Not between")'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, ltp_btw_bid_ask)

            # create formula for exception y/n 2
            col = 'Exception\n(Y/N)'
            excelcol = colname_to_excelcol["Last Trade price between Bid/Ask range?\n(between / not between)"]
            exception_2 = f'= IF({excelcol}{row} = "Between", "N", "Y")'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, exception_2)

            # create formula for price at bid
            col = r'Price per client'
            ref_excelcol = colname_to_excelcol['Exception\n(Y/N)']
            price_per_client = f'= IF({ref_excelcol}{row} = "Y", {colname_to_excelcol["Last Trade Price per unit (Local Currency)"]}{row}, "")'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, price_per_client)

            # create formula for price at bid
            col = r'Price at Bid'
            ref_excelcol = colname_to_excelcol['Exception\n(Y/N)']
            price_at_bid = f'= IF({ref_excelcol}{row} = "Y", {colname_to_excelcol["Bid Price Obtained from"]}{row}, "")'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, price_at_bid)
            
            # create formula for price at ask
            col = r'Price at Ask'
            ref_excelcol = colname_to_excelcol['Exception\n(Y/N)']
            price_at_ask = f'= IF({ref_excelcol}{row} = "Y", {colname_to_excelcol["Ask Price Obtained from"]}{row}, "")'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, price_at_ask)

            # create formula for price at ask
            col = r'Max difference'
            ref_excelcol = colname_to_excelcol['Exception\n(Y/N)']
            max_diff = f'= IF({ref_excelcol}{row} = "Y", ({colname_to_excelcol["Ask Price Obtained from"]}{row} - {colname_to_excelcol["Bid Price Obtained from"]}{row}) * {colname_to_excelcol["Exchange Rate @"]}{row} * {colname_to_excelcol["Holdings"]}{row}, "")'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, max_diff)
            self._standardise_cell_format(templ_ws, colname_to_excelcol[col], row)

            row += 1

        # TODO: all cols and rows hardcoded
        templ_ws[f"B9"].value = self.process_nav()
        templ_ws[f"B10"].value = float(self.user_inputs.at["om", "Answer"])
        templ_ws[f"B11"].value = float(self.user_inputs.at["pm", "Answer"])
        templ_ws[f"B12"].value = re.sub(r'(?<=\=B)5(?=\*0\.05%)', '9', templ_ws[f"B12"].value)

        # update total at bottom of sheet
        if input_length > 25:
            excelcol = colname_to_excelcol['Market Value at Last Trade Price (Base)']
            templ_ws[f"{excelcol}{38+6+input_length-25}"].value = f"=SUM({excelcol}{14+4}:{excelcol}{36+6+input_length-25})"
            templ_ws[f"{excelcol}{40+6+input_length-25}"].value = float(self.user_inputs.at["manual_adjustment", "Answer"])
            templ_ws[f"{excelcol}{41+6+input_length-25}"].value = f"=SUM({excelcol}{38+6+input_length-25}:{excelcol}{40+6+input_length-25})"
            templ_ws[f"{excelcol}{44+6+input_length-25}"].value = f"={excelcol}{41+6+input_length-25}-{excelcol}{43+6+input_length-25}"

            excelcol = colname_to_excelcol['Diff in Value (Base)']
            templ_ws[f"{excelcol}{38+6+input_length-25}"].value = f"=SUM({excelcol}{14+4}:{excelcol}{36+6+input_length-25})"


        self._adjust_col_width(templ_ws)

        templ_wb.save(self.output_fp)
        templ_wb.close()




if __name__ == "__main__":

    client_no = 50060
    fy        = 2023

    # recon_input_fp = r"D:\Documents\Project\Internal Projects\20240122 FS Funds\Recon output.xlsx"
    output_fp = r"D:\workspace\luna\personal_workspace\db\funds_test.xlsx"
    portfolio_mapper_fp = r"D:\workspace\luna\parameters\invtmt_portfolio_mapper.xlsx"

    client_class = tables.client.ClientInfoLoader_From_LunaHub(client_no)
    sublead_class = tables.fs_funds_output_sublead.FundsSublead_DownloaderFromLunaHub(client_no, fy)
    portfolio_class = tables.fs_funds_output_portfolio.FundsPortfolio_DownloaderFromLunaHub(client_no, fy)
    recon_class = tables.fs_funds_recon_details.FundsReconDetail_DownloaderFromLunaHub(client_no, fy)
    tb_class = common.TBLoader_From_LunaHub(client_no, fy)

    aic_name = "DS Team"

    for attempt in range(12):
        user_response_class = tables.fs_funds_userresponse.FundsUserResponse_DownloaderFromLunaHub(
            client_no,
            fy)
        user_inputs = user_response_class.main()
        if user_inputs is not None:
            break
        elif user_inputs is None and attempt == 11:
            raise Exception (f"Data not found for specified client {client_no} or FY {fy}.")
        else:
            continue
    
    self = InvtmtOutputFormatter(sublead_class  = sublead_class,
                                 portfolio_class= portfolio_class,
                                 recon_class    = recon_class,
                                 tb_class       = tb_class,
                                 output_fp      = output_fp,
                                 mapper_fp      = portfolio_mapper_fp,
                                 user_inputs    = user_inputs,
                                 client_class   = client_class,
                                 fy             = fy,
                                 aic_name       = aic_name
                                 )


    if True:

        import webbrowser
        webbrowser.open(output_fp)
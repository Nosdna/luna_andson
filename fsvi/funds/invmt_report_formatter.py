import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd
import numpy as np
from copy import copy
from datetime import datetime
import re

import luna
from luna.fsvi.funds.invmt_report_template_reader import FundsInvmtTemplateReader
import luna.common as common
from luna.lunahub import tables
import os

import pyeasylib.excellib as excellib

class InvmtOutputFormatter:

    LSCODES_NAV                 = [pd.Interval(6900.0, 6900.4, closed='both')]
    LSCODES_BOND_INT            = [pd.Interval(7400.2, 7400.2, closed='both')]
    LSCODES_BOND_INTREC         = [pd.Interval(5200.0, 5300.0, closed='left')]
    LSCODES_UNREALISED_GAINLOSS = [pd.Interval(7000.2, 7000.2, closed='both')]
    LSCODES_REALISED_GAINLOSS   = [pd.Interval(7000.1, 7000.1, closed='both')]
    #Adding in parameter for Txn related
    TRANSACTIONTYPERSM_PURC = "buy"
    TRANSACTIONTYPERSM_SALES = "sell"

    DATABASE_MISC_COLS  = ['WORKFLOWDATE', 'CLIENTNUMBER','FY',
                           'UPLOADER', 'UPLOADDATETIME',
                           'COMMENT1', 'COMMENT2', 'COMMENT3'
                           ]
    
    DATABASE_SUBLEAD_MISC_COLS = ['REPORTNAME','INSTITUTION','FILEPATH','FILENAME','SHEETNAME','WORKFLOWDATE', 'CLIENTNUMBER','FY',
                           'UPLOADER', 'UPLOADDATETIME',
                           'COMMENT1', 'COMMENT2', 'COMMENT3'
                           ]

    CONFIDENCE_THRESHOLD = 0.45

    def __init__(self, sublead_class, portfolio_class, recon_class,
                 broker_class, custodian_class, tb_class, 
                 processedtransaction_class, processedportfolio_class,
                 output_fp, mapper_fp, user_inputs,
                 client_class, fy, aic_name = ""
                 ):

        self.sublead_class  = sublead_class
        self.portfolio_class= portfolio_class
        self.recon_class    = recon_class
        self.broker_class   = broker_class
        self.custodian_class= custodian_class
        self.tb_class       = tb_class
        self.output_fp      = output_fp
        self.mapper_fp      = mapper_fp
        self.user_inputs    = user_inputs
        self.client_class   = client_class
        self.processedtransaction_class = processedtransaction_class
        self.processedportfolio_class = processedportfolio_class
        self.fy             = int(fy)
        self.aic_name       = aic_name

        self.template_fp = os.path.join(luna.settings.LUNA_FOLDERPATH,
                                        "parameters",
                                        "investment_template.xlsx"
                                        )
        
        self.template_class = FundsInvmtTemplateReader(self.template_fp)

        self.main()

    def main(self):

        self.get_data()
        self.write_sublead_output()
        self.write_recon_output()
        self.write_portfolio_output()
        self.write_processed_broker()
        self.write_processed_custodian_confirmation()
        #self.write_readme(readme_table)


    def get_data(self):
        self.sublead_input_df = self.sublead_class.main()
        self.portfolio_input_df = self.portfolio_class.main()
        self.recon_input_df_detail = self.recon_class.main()
        self.portfolio_mapper_df = pd.read_excel(self.mapper_fp)
        self.broker_df = self.broker_class.main()
        self.custodian_confirmation_df = self.custodian_class.main()
        self.transaction_df = self.processedtransaction_class.main()
        self.portfolio_df = self.processedportfolio_class.main()
            
    def build_varname_to_values(self, df):
        
        df = df.copy()
    
        varname_to_values = df[
            ["VARNAME", "VALUE", "VALUEPREVFY"]
            ].dropna(subset=["VARNAME"])
        
        # Check that var_name is unique
        if varname_to_values["VARNAME"].is_unique:
            varname_to_values = varname_to_values.set_index("VARNAME")
        else:
            raise Exception ("Variable name is not unique.")
                    
        # self.varname_to_values = varname_to_values.copy()
        
        return varname_to_values

    def _load_client_info(self):

        self.client_name = self.client_class.retrieve_client_info("CLIENTNAME")
    
    def _standardise_number_format(self, ws, lst_of_excelcols, row):
        for excelcol in lst_of_excelcols:
            ws[f"{excelcol}{row}"].number_format = '#,##0.00'

    def _standardise_date_format(self, ws, lst_of_excelcols, row):
        for excelcol in lst_of_excelcols:
            ws[f"{excelcol}{row}"].number_format = 'DD/MM/YYYY'

    def _create_header(self, ws, title, del_row_no, add_row_no):

        # Unmerge all cells
        for merge in list(ws.merged_cells):
            ws.unmerge_cells(range_string=str(merge))

        if del_row_no != 0:
            ws.delete_rows(idx = 1, amount = del_row_no)
        ws.insert_rows(idx = 0, amount = 6 + add_row_no)

        for merge in list(ws.merged_cells):
            ws.unmerge_cells(range_string=str(merge))

        row = 1

        # header title
        ws[f"A{row}"].value = title
        ws[f"A{row}"].font = Font(size = "16", bold = True)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="00C0C0C0")
        ws[f"A{row}"].border = Border(left   = Side(style = "thin"),
                                 right  = Side(style = "thin"),
                                 top    = Side(style = "thin"),
                                 bottom = Side(style = "thin")
                                 )
        ws.merge_cells(f"A{row}:F{row}")

        date_of_analysis = datetime.now().strftime("%d/%m/%Y")

        dct_of_fields = {"Client name"      : self.client_name,
                         "FY"               : self.fy,
                         "Date of analysis" : date_of_analysis,
                         "Prepared by"      : self.aic_name,
                         "Reviewed by"      : ""
                         }
        
        for key in dct_of_fields:
            
            row += 1

            self._create_header_row_field(key, dct_of_fields[key], row, ws)

    def _create_header_row_field(self, field_title, field_value, row, ws):

        border_style = Border(left   = Side(style = "thin"),
                              right  = Side(style = "thin"),
                              top    = Side(style = "thin"),
                              bottom = Side(style = "thin")
                              )
    
        ws[f"A{row}"].value = field_title
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="00C0C0C0")
        ws[f"A{row}"].font = Font(bold = True)
        ws[f"A{row}"].border = border_style
        ws.merge_cells(f"A{row}:C{row}")

        ws[f"D{row}"].value = field_value
        ws[f"D{row}"].fill = PatternFill("solid", fgColor="00FFFFFF")
        ws[f"D{row}"].alignment = Alignment(horizontal = 'left')
        ws[f"D{row}"].border = border_style
        ws.merge_cells(f"D{row}:F{row}")

    def _format_header_cell(self, col_lst, row, ws):

        border_style = Border(left   = Side(style = "thin"),
                              right  = Side(style = "thin"),
                              top    = Side(style = "thin"),
                              bottom = Side(style = "thin")
                              )

        for col in col_lst:
            ws[f"{col}{row}"].fill = PatternFill("solid", fgColor="1F497D") # 20240424 changed to dark blue
            ws[f"{col}{row}"].font = Font(bold = True, color = "00FFFFFF")
            ws[f"{col}{row}"].border = border_style
        
    def _standardise_cell_format(self, ws, excelcol, row):
        font_style = Font(size = "10", name = "Arial")
        alignment_style =  Alignment(horizontal = 'center')
        ws[f"{excelcol}{row}"].font = font_style
        ws[f"{excelcol}{row}"].alignment = alignment_style
    
    def _populate_portfolio_formula(self, ws, colname_to_excelcol, col, row, formula):
        ws[f"{colname_to_excelcol[col]}{row}"].value = formula
        ws[f"{colname_to_excelcol[col]}{row}"].alignment = Alignment(horizontal = 'left')
        self._standardise_cell_format(ws, colname_to_excelcol[col], row)

    def _adjust_col_width(self, ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # Get column letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length)  # Add some padding
            ws.column_dimensions[column_letter].width = adjusted_width

    def _copy_column_style(self, ws,
                           src_excelcol, src_excelrow,
                           dst_excelcol, dst_excelrow
                           ):
        for src, dst in zip(ws[f"{src_excelcol}{src_excelrow}"], ws[f"{dst_excelcol}{dst_excelrow}"]):
            dst.fill = copy(src.fill)
    
    def _get_column_lst_letters_by_dtype(self, df, data_type):

        # Convert data_type list to dtype objects
        data_type_dtypes = [np.dtype(dtype) for dtype in data_type]

        # Get the list of columns whose data types match those in data_type
        lst_of_wanted_cols = df.columns[df.dtypes.isin(data_type_dtypes)].tolist()

        lst_of_wanted_cols_indices = [df.columns.get_loc(col) for col in lst_of_wanted_cols]

        lst_of_wanted_cols_letters = [openpyxl.utils.cell.get_column_letter(idx+1) for idx in lst_of_wanted_cols_indices]

        return lst_of_wanted_cols_letters
    
    def _get_column_lst_letters_by_substr(self, df, substr):

        mask = df.columns.str.contains(substr)

        # Filter columns based on the mask
        lst_of_wanted_cols = df.columns[mask].tolist()

        # Get the indices of the columns
        lst_of_wanted_cols_indices = df.columns.get_indexer(lst_of_wanted_cols)

        # Get the column letters corresponding to the indices
        lst_of_wanted_cols_letters = [openpyxl.utils.cell.get_column_letter(idx+1) for idx in lst_of_wanted_cols_indices]

        return lst_of_wanted_cols_letters
    
    def _get_col_letter_from_ref(self, ref_excelcol, mvmt):
        ref_colno = column_index_from_string(ref_excelcol)
        target_colno = ref_colno + int(mvmt)
        target_excelcol = get_column_letter(target_colno)

        return target_excelcol
    
    def _create_tb_reference_formula_for_sublead(self, name_of_reference_sheet):

        return f"=SUMIF('{name_of_reference_sheet}'!H:H, \"Included\", '{name_of_reference_sheet}'!F:F)"
    
    def map_portfolio_columns(self):
        
        self.portfolio_mapper_df = self.portfolio_mapper_df[~self.portfolio_mapper_df["Standardised"].isna()]
        self.portfolio_mapper_dict = dict(zip(self.portfolio_mapper_df['Standardised'], self.portfolio_mapper_df['Formatted']))
        self.processed_portfolio_input_df = self.portfolio_input_df.rename(columns = self.portfolio_mapper_dict).reset_index(drop = True)
    
    def update_custodian_confidence_level_indicator(self):

        col = "Notes for reconciliation"
        df = self.processed_portfolio_input_df.copy()
        threshold = InvmtOutputFormatter.CONFIDENCE_THRESHOLD

        # cond = [(df[col] >= threshold), (df[col] < threshold)]
        # res = [f"Confidence of '{df[col]}': Fuzzy matched", f"Confidence of '{df[col]}': Need investigation"]

        # self.processed_portfolio_input_df[col] = np.select(cond, res, default='')

        self.processed_portfolio_input_df[col] = df[col].apply(lambda x: f"Confidence of '{x:.2f}': Matched" if x >= threshold else f"Confidence of '{x:.2f}': Possible name mismatch")

    def sort_portfolio_input_df(self):
        df = self.processed_portfolio_input_df.copy()
        col = 'diff_in_holdings'
        df[col] = df['Holdings per confirmation @'].astype(float) - df['Holdings'].astype(float)

        cond = [(df[col] < 0.01) & (df[col] > -0.01),
                ((df[col] > 0.01) | (df[col] > 0.01)) & (df["Ref in (Custodian's Confirmation)"].isnull()) & (df["Security Name"].notnull())]
        res = [1, 2]

        df['rank'] = np.select(cond, res, default=3)

        self.processed_portfolio_input_df = df.sort_values(by = ['rank', 'Notes for reconciliation'],
                                                           ascending=[True, False],
                                                           axis = 0,
                                                           ignore_index = True
                                                           )

    # Function to insert row in between -> But does not work with hyperlink
    # def _insert_disclaimer(self, ws,column_to_insert, row_to_insert, message):
    #     max_row = ws.max_row
    #     max_column = get_column_letter(ws.max_column)
    #     ws.move_range(f"A{row_to_insert}:{max_column}{max_row}", rows=1, cols=0)
    #     ws[f"{column_to_insert}{row_to_insert}"].value = message


    def make_sheet_active(self, wb: openpyxl.workbook.workbook.Workbook, sheet_name: str):
        wb.active = wb[sheet_name]
        for ws in wb:
            ws.views.sheetView[0].tabSelected = ws.title == sheet_name

    def process_nav(self):

        is_overlap, true_match, false_match = self.tb_class.filter_tb_by_fy_and_ls_codes(self.fy, InvmtOutputFormatter.LSCODES_NAV)

        total_equity = true_match['Value'].sum()

        return total_equity
    
    def filter_tb_for_sublead_field(self, field):

        field_dict = {"bond_int"    : InvmtOutputFormatter.LSCODES_BOND_INT,
                      "bond_intrec" : InvmtOutputFormatter.LSCODES_BOND_INTREC,
                      "unrealised_gainloss": InvmtOutputFormatter.LSCODES_UNREALISED_GAINLOSS,
                      "realised_gainloss": InvmtOutputFormatter.LSCODES_REALISED_GAINLOSS
                      }
        
        ls_codes = field_dict[field]
        
        is_overlap, true_match, false_match = self.tb_class.filter_tb_by_fy_and_ls_codes(self.fy, ls_codes)

        filtered_tb = true_match.copy()

        if field in ["bond_int", "bond_intrec"]:
            filtered_tb['Include / Exclude'] = filtered_tb['Name'].apply(lambda x: 'Included' if re.match('(?i).*interest.*', x) else 'Excluded')
        elif field in ["unrealised_gainloss", "realised_gainloss"]:
            filtered_tb['Include / Exclude'] = filtered_tb['Name'].apply(lambda x: 'Excluded' if re.match('(?i).*fx|foreign exchange|forex.*', x) else 'Included')
        else:
            pass
        
        filtered_tb = filtered_tb.sort_values(by = 'Include / Exclude', ascending = False, ignore_index = True)

        # drop cols
        filtered_tb = filtered_tb.drop(["L/S (interval)", "Completed FY?"], axis = 1)

        return filtered_tb.set_index('Account No').copy()
    

    #Creating filter Transaction for sublead - A
    def filter_txn_for_sublead_field(self, field):

        field_dict = {"purc"    : InvmtOutputFormatter.TRANSACTIONTYPERSM_PURC,
                      "unreal_txn_cost"    : InvmtOutputFormatter.TRANSACTIONTYPERSM_PURC,
                      "sales" : InvmtOutputFormatter.TRANSACTIONTYPERSM_SALES,
                      "report_cost" : InvmtOutputFormatter.TRANSACTIONTYPERSM_SALES,
                      "report_sales" : InvmtOutputFormatter.TRANSACTIONTYPERSM_SALES,
                      "real_txn_cost" : InvmtOutputFormatter.TRANSACTIONTYPERSM_SALES,
                      }

        transaction_types = field_dict[field]

        is_type, true_match, false_match = self.processedtransaction_class.filter_txn_by_txntype(self.fy,transaction_types)

        #filtered_txn = true_match.copy()
        filtered_txn = pd.concat([true_match,false_match])

        #filtered_df = df[df.apply(lambda row: row['type'] == 'buy', axis=1)]
        if field in ["purc", "unreal_txn_cost"]:
            #filtered_txn = filtered_txn[filtered_txn['TRANSACTIONTYPERSM'] == InvmtOutputFormatter.TRANSACTIONTYPERSM_PURC]
            filtered_txn['Include / Exclude'] = filtered_txn['TRANSACTIONTYPERSM'].apply(lambda x: 'Included' if x==InvmtOutputFormatter.TRANSACTIONTYPERSM_PURC else 'Excluded')
        elif field in ["sales", "report_cost","report_sales","real_txn_cost"]:
            #filtered_txn = filtered_txn[filtered_txn['TRANSACTIONTYPERSM'] == InvmtOutputFormatter.TRANSACTIONTYPERSM_SALES]
            filtered_txn['Include / Exclude'] = filtered_txn['TRANSACTIONTYPERSM'].apply(lambda x: 'Included' if x==InvmtOutputFormatter.TRANSACTIONTYPERSM_SALES else 'Excluded')
        else:
            pass

        return filtered_txn.set_index('SECURITYNAME').copy()
    
    #Created function to remove index column from df_processed and future use
    def filter_position_for_sublead_field(self, field):
        
        filtered_port = processedportfolio_class.df_processed.copy()

        if field in ["cost_at_end", "fv_valuation_report","unreal_port_cost"]:
            filtered_port['Include / Exclude'] = "Included"
        else:
            pass        

        return filtered_port.set_index('SECURITYNAME').copy()


    def create_hyperlink_in_sublead_to_tb(self, wb, source_sheetname, return_source_cell, cell, reference_sheetname, title, filtered_tb, filter_logic):

        field_to_source_locations = {"<<<link>>>": cell}

        header_rows = ["Title", "Client name", "Data as of", "FY", 
                       "Prepared by", "Reviewed by"]

        #Dropping unused columns for txn and portfolio
        print(self.DATABASE_SUBLEAD_MISC_COLS)
        if all(col in filtered_tb.columns for col in self.DATABASE_SUBLEAD_MISC_COLS):
            cols_to_drop = self.DATABASE_SUBLEAD_MISC_COLS.copy()
            filtered_tb = filtered_tb.drop(cols_to_drop, errors = 'ignore', axis = 1)
        else:
            pass

        field_to_data = {"<<<link>>>" : filtered_tb}

        hyperlink_class = common.hyperlinks.DataHyperlink(source_sheetname, "A9", field_to_source_locations, reference_sheetname, header_rows, field_to_data, wb)
        hyperlink_class.write_reference_data()
        hyperlink_class.write_source_data()

        ws = wb[reference_sheetname]

        self._create_header(wb[reference_sheetname], title, 6, 1)

        for c_idx, header_value in enumerate(filtered_tb.reset_index().columns, 1):
            col = openpyxl.utils.cell.get_column_letter(c_idx)
            self._format_header_cell([col], 10, ws)

        for r_idx, row in enumerate(filtered_tb.reset_index().values, 11):
            self._standardise_number_format(ws, ['F'], r_idx) #TODO: hardcoded excelcols and rows  
            self._standardise_date_format(ws, ['E'], r_idx) #TODO: hardcoded excelcols and rows  

        field_to_source_locations = {"<<<Return to Sub-lead>>>": "A9"}

        ws["B9"].value = "Filter logic: " + filter_logic
        ws["A8"].value = "Data has been extracted and standardised to all formats, and may not be same as followed in source document. For detailed explaination, view README"
        ws.merge_cells("A8:F8") #Readme message
        ws.merge_cells("B9:F9") #Filter logic

        hyperlink_class = common.hyperlinks.DataHyperlink(reference_sheetname, return_source_cell, field_to_source_locations, source_sheetname, [], {}, wb)
        hyperlink_class.write_reference_data()
        hyperlink_class.write_source_data()

        #self._insert_disclaimer(ws,"A", "8", "Data has been extracted and standardised to all formats, and may not be same as followed in source document. For detailed explaination, view README")

    def process_readme(self, ws_readme):
        ws_readme.merge_cells("A2:D3")

    def write_sublead_output(self):

        sheet_name = "<5100-xx>Investment sub-lead"

        templ_wb = openpyxl.load_workbook(self.template_fp)
        templ_ws = templ_wb[sheet_name]

        self._load_client_info()

        # template as df
        colname_to_excelcol = self.template_class.sublead_colname_to_excelcol
        varname_to_index = self.template_class.sublead_varname_to_index

        # additional processing for sublead
        varname_to_index_df = varname_to_index.to_frame().reset_index()
        filtered_varname_to_index_df = varname_to_index_df[varname_to_index_df['var_name'].str.match(r".*///.*")]
        for i in range(len(filtered_varname_to_index_df)):
            i_idx = filtered_varname_to_index_df.index[i]
            varname_to_index_df.at[i_idx, 'var_name'] = filtered_varname_to_index_df.at[i_idx, 'var_name'].split(' /// ')
        varname_to_index = varname_to_index_df.explode('var_name').set_index('var_name').squeeze()
                
        # get the data
        varname_to_values = self.build_varname_to_values(self.sublead_input_df)
        varname_to_values.at['cost_roll_sales_txn_report_sales', 'VALUEPREVFY'] = varname_to_values.at['cost_roll_sales_txn_report_sales', 'VALUE']
        varname_to_values.at['cost_roll_sales_txn_report_sales', 'VALUE'] = varname_to_values.at['cost_roll_sales_txn_report_cost', 'VALUE']
        varname_to_values = varname_to_values.drop("cost_roll_sales_txn_report_cost")
        self.varname_to_values = varname_to_values.copy()

        # save column index
        varname_excelcol = colname_to_excelcol.at["var_name"]
        cfy_excelcol = colname_to_excelcol.at["Current FY"]
        pfy_excelcol = colname_to_excelcol.at["Previous FY"]


        for varname in varname_to_values.index:
            cfy = varname_to_values.at[varname, "VALUE"]
            pfy = varname_to_values.at[varname, "VALUEPREVFY"]

            row = varname_to_index.at[varname]

            templ_ws[f"{cfy_excelcol}{row}"].value = cfy
            templ_ws[f"{pfy_excelcol}{row}"].value = pfy

            self._standardise_number_format(templ_ws,
                                            [cfy_excelcol, pfy_excelcol],
                                            row
                                            )
            self._standardise_cell_format(templ_ws, cfy_excelcol, row)
            self._standardise_cell_format(templ_ws, pfy_excelcol, row)

        fy_excelrow = min(varname_to_index) - 3
        fcy_lst = self.sublead_input_df["FUNCTIONALCURR"].to_list()
        if len(set(fcy_lst)) > 1:
            raise Exception("More than one unique value found in FUNCTIONALCURR column:"
                            f"{fcy_lst}.")
        else:
            fcy = fcy_lst[0]

        templ_ws[f"{cfy_excelcol}{fy_excelrow}"].value = self.fy
        templ_ws[f"{pfy_excelcol}{fy_excelrow}"].value = self.fy-1
        templ_ws[f"{cfy_excelcol}{fy_excelrow+1}"].value = fcy
        templ_ws[f"{pfy_excelcol}{fy_excelrow+1}"].value = fcy
            
        
        self._create_header(templ_ws, sheet_name, 6, 0)

        templ_ws.column_dimensions[varname_excelcol].hidden = True

        filtered_tb_bond_int = self.filter_tb_for_sublead_field("bond_int")
        filtered_tb_bond_intrec = self.filter_tb_for_sublead_field("bond_intrec")
        filtered_tb_unrealised_gainloss = self.filter_tb_for_sublead_field("unrealised_gainloss")
        filtered_tb_realised_gainloss = self.filter_tb_for_sublead_field("realised_gainloss")

        #pointing to filter txn for sublead
        filtered_txn_purc = self.filter_txn_for_sublead_field("purc")
        filtered_txn_unreal_txn_cost = self.filter_txn_for_sublead_field("unreal_txn_cost")
        filtered_txn_sales = self.filter_txn_for_sublead_field("sales")
        filtered_txn_report_cost = self.filter_txn_for_sublead_field("report_cost")
        filtered_txn_report_sales = self.filter_txn_for_sublead_field("report_sales")
        filtered_txn_real_txn_cost = self.filter_txn_for_sublead_field("real_txn_cost")

        #pointing to filter portfolio for sublead
        filtered_portfolio_cost_at_end = self.filter_position_for_sublead_field("cost_at_end")
        filtered_portfolio_fv_valuation_report = self.filter_position_for_sublead_field("fv_valuation_report")
        filtered_portfolio_unreal_port_cost = self.filter_position_for_sublead_field("unreal_port_cost")
        
        self.create_hyperlink_in_sublead_to_tb(templ_wb, sheet_name, "E43", "E43",
                                               "Interest - Bonds TB",
                                               "Trial Balance for Interest - Bonds",
                                               filtered_tb_bond_int,
                                               "Filter for LSCODE = 7400.200 and for ACCOUNTNAME containing 'Interest'.")
        self.create_hyperlink_in_sublead_to_tb(templ_wb, sheet_name, "E45", "E45",
                                               "Interest Receivables - Bonds TB", 
                                               "Trial Balance for Interest Receivables - Bonds",
                                               filtered_tb_bond_intrec,
                                               "Filter for LSCODE >= 5200.000 and <5201.000 and for ACCOUNTNAME containing 'Interest'.")
        self.create_hyperlink_in_sublead_to_tb(templ_wb, sheet_name, "E37", "E37",
                                               "Unrealised Gain or Loss TB", 
                                               "Trial Balance for Unrealised Gain/Loss",
                                               filtered_tb_unrealised_gainloss,
                                               "Filter for LSCODE = 7000.200 and for ACCOUNTNAME not containing 'FX' or 'Foreign exchange'.")
        self.create_hyperlink_in_sublead_to_tb(templ_wb, sheet_name, "E49", "E49",
                                               "Realised Gain or Loss TB", 
                                               "Trial Balance for Realised Gain/Loss",
                                               filtered_tb_realised_gainloss,
                                               "Filter for LSCODE = 7000.100 and for ACCOUNTNAME not containing 'FX' or 'Foreign exchange'.")
        
        #Creating new hyperlinks per Kelly feedback - A

        #Transaction - Purchases
        self.create_hyperlink_in_sublead_to_tb(templ_wb, sheet_name, "E16", "E16",
                                               "Purchases Transaction", 
                                               "Transaction Report for Purchases",
                                               filtered_txn_purc,
                                               "Filter TRANSACTIONTYPERSM = buy and sum of (TRANSACTIONAMOUNTLCY-REALISEDPRICELCY) * TRADEDATERATE")
        #Transaction - Cost of Sales
        self.create_hyperlink_in_sublead_to_tb(templ_wb, sheet_name, "E17", "E17",
                                               "Sales Transaction", 
                                               "Transaction Report for Sales",
                                               filtered_txn_sales,
                                               "Filter TRANSACTIONTYPERSM = sell and sum of (TRANSACTIONAMOUNTLCY-REALISEDPRICELCY) * TRADEDATERATE")
        
        #Create hyperlink for E20, same as sales - harde
        hyperlink = "#'Sales Transaction'!A9"        
        ws = templ_wb[sheet_name]
        ws["E24"].hyperlink = hyperlink
        ws["E24"].value = "<<<link>>>"
        ws["E24"].style = 'Hyperlink'

        #Portfolio - Cost at end of year/period per client
        self.create_hyperlink_in_sublead_to_tb(templ_wb, sheet_name, "E19", "E19",
                                               "Cost at End of Year Portfolio", 
                                               "Investment Portfolio for Cost",
                                               filtered_portfolio_cost_at_end,
                                               "Sum of (MARKETVALUEFCY / MARKETVALUELCY) * COSTLCY")
        #Transaction - Sale Proceeds
        self.create_hyperlink_in_sublead_to_tb(templ_wb, sheet_name, "G24", "G24",
                                               "Sales Proceeds Transaction", 
                                               "Transaction Report for Sales Proceeds",
                                               filtered_txn_report_sales,
                                               "Filter TRANSACTIONTYPERSM = buy and sum of TRANSACTIONAMOUNTFCY")
        #Portfolio - FV of investment
        self.create_hyperlink_in_sublead_to_tb(templ_wb, sheet_name, "E28", "E28",
                                               "FV of Investment Portfolio", 
                                               "Investment Portfolio for FV of Investment",
                                               filtered_portfolio_fv_valuation_report,
                                               "Sum of MARKETVALUEFCY")
        #Portfolio - Cost of investment
        self.create_hyperlink_in_sublead_to_tb(templ_wb, sheet_name, "E32", "E32",
                                               "Cost of Investment Portfolio", 
                                               "Investment Portfolio for Cost of Investment",
                                               filtered_portfolio_unreal_port_cost,
                                               "Sum of (MARKETVALUEFCY / MARKETVALUELCY) * COSTLCY")
        
        #Transaction - Unreal Transaction Cost
        self.create_hyperlink_in_sublead_to_tb(templ_wb, sheet_name, "E36", "E36",
                                               "Unrealised Transaction Cost", 
                                               "Transaction Report for Unrealised Transaction Cost",
                                               filtered_txn_unreal_txn_cost,
                                               "Filter TRANSACTIONTYPERSM = buy and sum of BROKERCOMMISSIONFCY + Sum of OTHERFEESFCY ")

        #Transaction - Real Transaction Cost
        self.create_hyperlink_in_sublead_to_tb(templ_wb, sheet_name, "E48", "E48",
                                               "Realised Transaction Cost", 
                                               "Transaction Report for Realised Transaction Cost",
                                               filtered_txn_real_txn_cost,
                                               "Filter TRANSACTIONTYPERSM = sell and sum of COMMISSIONFCY + Sum of OTHERCHARGESFCY ")
        

        # TODO: check if this is what they want
        if False:
            templ_ws["D39"].value = self._create_tb_reference_formula_for_sublead("Interest - Bonds TB")
            templ_ws["D41"].value = self._create_tb_reference_formula_for_sublead("Interest Receivables - Bonds TB")
            templ_ws["D33"].value = self._create_tb_reference_formula_for_sublead("Unrealised Gain or Loss TB")
            templ_ws["D45"].value = self._create_tb_reference_formula_for_sublead("Realised Gain or Loss TB")

        templ_ws["E43"].fill = copy(templ_ws["D43"].fill)
        templ_ws["E45"].fill = copy(templ_ws["D45"].fill)
        templ_ws["E37"].fill = copy(templ_ws["D37"].fill)
        templ_ws["E49"].fill = copy(templ_ws["D49"].fill)
        templ_ws["E16"].fill = copy(templ_ws["D16"].fill)
        templ_ws["E17"].fill = copy(templ_ws["D17"].fill)
        templ_ws["E19"].fill = copy(templ_ws["D19"].fill)
        templ_ws["G24"].fill = copy(templ_ws["G24"].fill)
        templ_ws["E24"].fill = copy(templ_ws["D19"].fill)
        templ_ws["E28"].fill = copy(templ_ws["E29"].fill)
        templ_ws["E32"].fill = copy(templ_ws["E29"].fill)
        templ_ws["E36"].fill = copy(templ_ws["E29"].fill)
        templ_ws["E48"].fill = copy(templ_ws["E29"].fill)
        
        # merging auditor guidance and conclusion
        templ_ws.merge_cells("B8:F10") # Auditor Guidance
        templ_ws.merge_cells("B71:F73") # Conclusion

        # modifying readme merge
        self.process_readme(templ_wb["Readme"])

        templ_wb.save(self.output_fp)
        templ_wb.close()

    def write_recon_output(self):

        summary_sheet_name = "Investment txn recon summary"
        detail_sheet_name = "Investment txn recon detail"

        templ_wb = openpyxl.load_workbook(self.output_fp)

        # summary tab
        templ_ws = templ_wb[summary_sheet_name]

        # summary content writing
        self.recon_input_df_summary = self.recon_input_df_detail.groupby(by="EXCEPTIONINDICATOR").agg(
            # MATCHINGINDICATOR       = ('MATCHINGINDICATOR', lambda x: ','.join(x.unique())),
            EXCEPTIONINDICATOR      = ('EXCEPTIONINDICATOR', lambda x: ','.join(x.unique())),
            MARKETVALUEFUNDADMIN    = ("MARKETVALUEFUNDADMIN", 'sum'),
            MARKETVALUEBROKER       = ('MARKETVALUEBROKER', 'sum'),
            VALUEDIFFERENCE         = ("VALUEDIFFERENCE", 'sum')
            )
        
        content_df = self.recon_input_df_summary.copy()

        for c_idx, header_value in enumerate(content_df.columns, 1):
            templ_ws.cell(row=1, column=c_idx, value=header_value)
            col = openpyxl.utils.cell.get_column_letter(c_idx)
            self._format_header_cell([col], 1, templ_ws)

        lst_of_number_cols = self._get_column_lst_letters_by_dtype(content_df, ["float64"])

        for r_idx, row in enumerate(self.recon_input_df_summary.values, 2):
            for c_idx, value in enumerate(row, 1):
                templ_ws.cell(row=r_idx, column=c_idx, value=value)
            self._standardise_number_format(templ_ws, lst_of_number_cols, r_idx) 
            
        self._create_header(templ_ws, summary_sheet_name, 0, 1)

        # detail tab
        templ_ws = templ_wb[detail_sheet_name]

        self.recon_input_df_detail["transaction_type_rsm_rank"] = self.recon_input_df_detail['TRANSACTIONTYPERSMBROKER'].apply(lambda x: 2 if x == "others" else 1)
        self.recon_input_df_detail["value_difference_abs"] = abs(self.recon_input_df_detail["VALUEDIFFERENCE"])
        self.recon_input_df_detail = self.recon_input_df_detail.sort_values(by = ['value_difference_abs', 'transaction_type_rsm_rank', 'MATCHINGINDICATORNAME', 'CONFIDENCELEVELNAME'],
                                                           ascending=[True, True, True, False],
                                                           axis = 0,
                                                           ignore_index = True
                                                           )
        
        content_df = self.recon_input_df_detail.copy()        
        cols_to_drop = InvmtOutputFormatter.DATABASE_MISC_COLS.copy()
        content_df = content_df.drop(cols_to_drop, axis = 1)

        # add formula columns
        content_df.loc[:, ["PRICEDIFFERENCE", "QUANTITYDIFFERENCE"]] = float(0.00)

        # 20240424 - commented out because we decided to group the same source together
        # content_df_col_order = ['SECURITYNAMEFUNDADMIN', 'SECURITYNAMEBROKER', 'ISINFUNDADMIN',
        #                         'ISINBROKER', 'TRADEDATEFUNDADMIN', 'TRADEDATEBROKER', 'PRICEFUNDADMIN',
        #                         'PRICEBROKER', 'PRICEDIFFERENCE', 'QUANTITYFUNDADMIN', 'QUANTITYBROKER',
        #                         'QUANTITYDIFFERENCE', 'MARKETVALUEFUNDADMIN', 'MARKETVALUEBROKER', 'VALUEDIFFERENCE',
        #                         'TRANSACTIONTYPEFUNDADMIN', 'TRANSACTIONTYPEBROKER',
        #                         'TRANSACTIONTYPERSMFUNDADMIN', 'TRANSACTIONTYPERSMBROKER',
        #                         'LOCALCURRFUNDADMIN', 'LOCALCURRBROKER', 'EXCEPTIONINDICATOR',
        #                         'CONFIDENCELEVELNAME', 'MATCHINGINDICATORNAME', 'INSTITUTIONFUNDADMIN',
        #                         'INSTITUTIONBROKER']
        content_df_col_order = ['SECURITYNAMEFUNDADMIN', 'ISINFUNDADMIN', 'TRADEDATEFUNDADMIN',
                               'PRICEFUNDADMIN', 'QUANTITYFUNDADMIN', 'MARKETVALUEFUNDADMIN', 
                               'TRANSACTIONTYPEFUNDADMIN', 'TRANSACTIONTYPERSMFUNDADMIN',
                                'LOCALCURRFUNDADMIN','INSTITUTIONFUNDADMIN',
                               'SECURITYNAMEBROKER', 'ISINBROKER', 'TRADEDATEBROKER',
                                'PRICEBROKER', 'QUANTITYBROKER', 'MARKETVALUEBROKER',
                                'TRANSACTIONTYPEBROKER', 'TRANSACTIONTYPERSMBROKER',
                                'LOCALCURRBROKER', 'INSTITUTIONBROKER',
                                'PRICEDIFFERENCE', 'QUANTITYDIFFERENCE', 'VALUEDIFFERENCE', 'EXCEPTIONINDICATOR',
                                'CONFIDENCELEVELNAME', 'MATCHINGINDICATORNAME']
        content_df = content_df[content_df_col_order]
        lst_of_number_cols = self._get_column_lst_letters_by_dtype(content_df, ["float64"])
        lst_of_date_cols = self._get_column_lst_letters_by_dtype(content_df, ["datetime64[ns]"] )
        lst_of_fundadmin_cols = self._get_column_lst_letters_by_substr(content_df, "FUNDADMIN")
        lst_of_broker_cols = self._get_column_lst_letters_by_substr(content_df, "BROKER")
        lst_of_difference_cols = self._get_column_lst_letters_by_substr(content_df, "DIFFERENCE")

        # detail content writing
        for c_idx, header_value in enumerate(content_df.columns, 1):
            cell = templ_ws.cell(row=1, column=c_idx, value=header_value)
            col = openpyxl.utils.cell.get_column_letter(c_idx)
            self._format_header_cell([col], 1, templ_ws)
            if col in lst_of_fundadmin_cols:
                cell.fill = PatternFill("solid", fgColor = "C5D9F1")
                cell.font = Font(bold = True,
                                 color = "00000000")
            elif col in lst_of_difference_cols:
                cell.fill = PatternFill("solid", fgColor = "00C0C0C0")
                cell.font = Font(bold = True,
                                 color = "00000000")
            elif col in lst_of_broker_cols:
                pass
            else:
                cell.fill = PatternFill("solid", fgColor = "00C0C0C0")
                cell.font = Font(bold = True,
                                 color = "00000000")
            

        for r_idx, row in enumerate(content_df.values, 2):
            for c_idx, value in enumerate(row, 1):
                # templ_ws.cell(row=r_idx, column=c_idx, value=value)
                cell = templ_ws.cell(row=r_idx, column=c_idx, value=value)
                excelcol = openpyxl.utils.cell.get_column_letter(c_idx)
                # # to colour whole column in a specific colour
                # if excelcol in lst_of_fundadmin_cols:
                #     cell.fill = PatternFill("solid", fgColor = "D9D9D9")
                # elif excelcol in lst_of_broker_cols:
                #     cell.fill = PatternFill("solid", fgColor = "F1F1F1")
                # elif excelcol in lst_of_difference_cols:
                #     fa_excelcol = self._get_col_letter_from_ref(excelcol, -2)
                #     br_excelcol = self._get_col_letter_from_ref(excelcol, -1)
                #     cell.value = f"= {fa_excelcol}{r_idx+7} - {br_excelcol}{r_idx+7}"
                #     cell.fill = PatternFill("solid", fgColor = "C5D9F1")
                # else:
                #     pass
                if excelcol in lst_of_difference_cols:
                    colname = templ_ws[f"{excelcol}1"].value
                    value_test = re.search("(.*?)DIFFERENCE", colname).group(1)
                    match value_test:

                        case "PRICE" | "QUANTITY":
                            fa_colname = f"{value_test}FUNDADMIN"
                            br_colname = f"{value_test}BROKER"
                            fa_dist_from_diff_col = content_df_col_order.index(fa_colname) - content_df_col_order.index(colname)
                            fa_excelcol = self._get_col_letter_from_ref(excelcol, fa_dist_from_diff_col)
                            br_dist_from_diff_col = content_df_col_order.index(br_colname) - content_df_col_order.index(colname)
                            br_excelcol = self._get_col_letter_from_ref(excelcol, br_dist_from_diff_col)
                            cell.value = f"= {fa_excelcol}{r_idx+7} - {br_excelcol}{r_idx+7}"

                        case "VALUE":
                            fa_colname = f"MARKET{value_test}FUNDADMIN"
                            br_colname = f"MARKET{value_test}BROKER"
                            fa_dist_from_diff_col = content_df_col_order.index(fa_colname) - content_df_col_order.index(colname)
                            fa_excelcol = self._get_col_letter_from_ref(excelcol, fa_dist_from_diff_col)
                            br_dist_from_diff_col = content_df_col_order.index(br_colname) - content_df_col_order.index(colname)
                            br_excelcol = self._get_col_letter_from_ref(excelcol, br_dist_from_diff_col)
                            cell.value = f"= {fa_excelcol}{r_idx+7} - {br_excelcol}{r_idx+7}"

                        case _:

                            pass
                    
            self._standardise_number_format(templ_ws, lst_of_number_cols, r_idx) 
            self._standardise_date_format(templ_ws, lst_of_date_cols, r_idx)
        self._create_header(templ_ws, detail_sheet_name, 0, 1)

        templ_wb.save(self.output_fp)
        templ_wb.close()

    def write_portfolio_output(self):

        sheet_name = "<5100-xx>Investment Portfolio"

        templ_wb = openpyxl.load_workbook(self.output_fp)
        templ_ws = templ_wb[sheet_name]

        self._load_client_info()

        colname_to_excelcol = self.template_class.portfolio_colname_to_excelcol

        self.map_portfolio_columns()
        self.update_custodian_confidence_level_indicator()

        self._create_header(templ_ws, sheet_name, 2, 0)

        row = 18
        templ_ws.merge_cells("A14:F16") # Auditor Guidance
        templ_ws.merge_cells(f"A{row}:H{row}") # per client
        templ_ws.merge_cells(f"I{row}:M{row}") # per custodian
        templ_ws.merge_cells(f"N{row}:S{row}") # per client
        templ_ws.merge_cells(f"T{row}:Z{row}") # per rsm
        templ_ws.merge_cells(f"AB{row}:AE{row}") # per rsm
        templ_ws.merge_cells(f"AG{row}:AJ{row}") # if ltp is not within bis ask spread

        content_df = self.processed_portfolio_input_df.copy()
        
        input_length = len(content_df)

        templ_ws.merge_cells(f"A{row+input_length+18}:F{row+input_length+20}") # Conclusion

        if input_length > 25:
            templ_ws.insert_rows(idx = 40, amount = input_length - 25 + 2)

        self.sort_portfolio_input_df()

        transposed_df = content_df.T

        lst_of_number_cols = self._get_column_lst_letters_by_dtype(content_df, ["float64"])
        lst_of_date_cols = self._get_column_lst_letters_by_dtype(content_df, ["datetime64[ns]"])

        for col in transposed_df.index:
            row = 17 + 5
            for i in range(len(transposed_df.columns)):
                try:
                    val = transposed_df.at[col, i]
                    templ_ws[f"{colname_to_excelcol[col]}{row}"].value = val
                    self._standardise_cell_format(templ_ws, colname_to_excelcol[col], row)
                    row += 1
                except:
                    pass
                self._standardise_number_format(templ_ws,
                                                ['J', 'K', 'N', 'P', 'Q', 'R',
                                                 'S', 'T', 'V', 'W', 'X', 'AB',
                                                 'AC', 'AD', 'AG', 'AH', 'AI',
                                                 'AJ'], # TODO: hardcoded
                                                row
                                                )
                self._standardise_date_format(templ_ws, ['H'], row)
                # self._standardise_number_format(templ_ws, lst_of_number_cols, row)
                # self._standardise_date_format(templ_ws, lst_of_date_cols, row) # TODO: format not showing
        
        # for r_idx, row in enumerate(content_df.values, 1):
        #     self._standardise_number_format(templ_ws, lst_of_number_cols, r_idx)
        #     self._standardise_date_format(templ_ws, lst_of_date_cols, r_idx)

        row = 17 + 5

        # # this will be KIV as a later enhancement
        # portfolio_formulas = self.template_class.portfolio_df_processed.iloc[0,:]
        # portfolio_formulas = portfolio_formulas[~portfolio_formulas.isna()]
        # portfolio_formulas_colname = portfolio_formulas
        # portfolio_formulas = portfolio_formulas.reset_index().set_index(12) #TODO: col reference hardcoded
        # mapping = portfolio_formulas.to_dict()['index']
        # col = 'Market Value at Last Trade Price (Base)'
        # formula = portfolio_formulas_colname[col]
        # # Regular expression pattern to match variable names
        # pattern = re.compile(r'\(\w\)')
        # # Iterate over each match and replace it with corresponding value
        # for var in pattern.findall(formula):
        #     formula = re.sub(re.escape(var), mapping.get(var, var), formula)
        #     print(formula)


        for i in range(len(transposed_df.columns)):

            # create formula for diff in holdings
            col = 'Diff in holdings?'
            diff_in_holdings = f"= {colname_to_excelcol['Holdings per confirmation @']}{row} - {colname_to_excelcol['Holdings']}{row}"
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, diff_in_holdings)

            # create formula for market value at last trade price (base)
            col = 'Market Value at Last Trade Price (Base)'
            mv_at_ltp_fcy = f"= {colname_to_excelcol['Holdings']}{row} * {colname_to_excelcol['Last Trade Price per unit (Local Currency)']}{row} * {colname_to_excelcol['Exchange Rate @']}{row}"
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, mv_at_ltp_fcy)

            # create formula for % of nav
            col = r'% of NAV'
            percent_of_nav = f"= {colname_to_excelcol['Market Value at Last Trade Price (Base)']}{row} / B9"
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, percent_of_nav)

            # create formula for market value per rsm (base)
            col = 'Market Value per RSM (Base)'
            mv_per_rsm_fcy = f"= {colname_to_excelcol['Holdings']}{row} * {colname_to_excelcol['Exchange Rate @']}{row} *  {colname_to_excelcol['Price Obtained from']}{row} "
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, mv_per_rsm_fcy)

            # create formula for diff in value (base)
            col = 'Diff in Value (Base)'
            diff_in_val_fcy = f"= {colname_to_excelcol['Market Value per RSM (Base)']}{row} - {colname_to_excelcol['Market Value at Last Trade Price (Base)']}{row}"
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, diff_in_val_fcy)

            # create formula for as a % of NAV
            col = r'As a % of NAV'
            percent_of_nav = f"= {colname_to_excelcol['Diff in Value (Base)']}{row} / B9"
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, percent_of_nav)

            # create formula for exception y/n
            col = r'Exception (Y/N)'
            exception_1 = f'= IF({colname_to_excelcol["Diff in Value (Base)"]}{row} = 0, "N", "Y")'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, exception_1)

            # create formula for lvl hierarchy
            col = 'Level hierarchy'
            lvl_hierarchy = f'= IF(ISBLANK({colname_to_excelcol["Price Obtained from"]}{row}), "", 1)'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, lvl_hierarchy)

            # create formula for ltp between range?
            col = 'Last Trade price between Bid/Ask range?\n(between / not between)'
            min_con_formula = f"{colname_to_excelcol['Last Trade Price per unit (Local Currency)']}{row}>=MIN({colname_to_excelcol['Bid Price Obtained from']}{row}, {colname_to_excelcol['Ask Price Obtained from']}{row})"
            max_con_formula = f"{colname_to_excelcol['Last Trade Price per unit (Local Currency)']}{row}<=MAX({colname_to_excelcol['Bid Price Obtained from']}{row}, {colname_to_excelcol['Ask Price Obtained from']}{row})"
            ltp_btw_bid_ask = f'= IF(AND({min_con_formula},{max_con_formula}), "Between", "Not between")'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, ltp_btw_bid_ask)

            # create formula for exception y/n 2
            col = 'Exception\n(Y/N)'
            excelcol = colname_to_excelcol["Last Trade price between Bid/Ask range?\n(between / not between)"]
            exception_2 = f'= IF({excelcol}{row} = "Between", "N", "Y")'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, exception_2)

            # create formula for price at bid
            col = r'Price per client'
            ref_excelcol = colname_to_excelcol['Exception\n(Y/N)']
            price_per_client = f'= IF({ref_excelcol}{row} = "Y", {colname_to_excelcol["Last Trade Price per unit (Local Currency)"]}{row}, "")'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, price_per_client)

            # create formula for price at bid
            col = r'Price at Bid'
            ref_excelcol = colname_to_excelcol['Exception\n(Y/N)']
            price_at_bid = f'= IF({ref_excelcol}{row} = "Y", {colname_to_excelcol["Bid Price Obtained from"]}{row}, "")'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, price_at_bid)
            
            # create formula for price at ask
            col = r'Price at Ask'
            ref_excelcol = colname_to_excelcol['Exception\n(Y/N)']
            price_at_ask = f'= IF({ref_excelcol}{row} = "Y", {colname_to_excelcol["Ask Price Obtained from"]}{row}, "")'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, price_at_ask)

            # create formula for price at ask
            col = r'Max difference'
            ref_excelcol = colname_to_excelcol['Exception\n(Y/N)']
            max_diff = f'= IF({ref_excelcol}{row} = "Y", ({colname_to_excelcol["Ask Price Obtained from"]}{row} - {colname_to_excelcol["Bid Price Obtained from"]}{row}) * {colname_to_excelcol["Exchange Rate @"]}{row} * {colname_to_excelcol["Holdings"]}{row}, "")'
            self._populate_portfolio_formula(templ_ws, colname_to_excelcol, col, row, max_diff)
            self._standardise_cell_format(templ_ws, colname_to_excelcol[col], row)

            row += 1

        # TODO: all cols and rows hardcoded
        templ_ws[f"B9"].value = self.process_nav()
        templ_ws[f"B10"].value = float(self.user_inputs.at["om", "Answer"])
        templ_ws[f"B11"].value = float(self.user_inputs.at["pm", "Answer"])
        templ_ws[f"B12"].value = re.sub(r'(?<=\=B)5(?=\*0\.05%)', '9', templ_ws[f"B12"].value)

        # update total at bottom of sheet
        if input_length > 25:
            excelcol = colname_to_excelcol['Market Value at Last Trade Price (Base)']
            templ_ws[f"{excelcol}{38+6+input_length-25}"].value = f"=SUM({excelcol}{14+4}:{excelcol}{36+6+input_length-25})"
            templ_ws[f"{excelcol}{40+6+input_length-25}"].value = float(self.user_inputs.at["manual_adjustment", "Answer"])
            templ_ws[f"{excelcol}{41+6+input_length-25}"].value = f"=SUM({excelcol}{38+6+input_length-25}:{excelcol}{40+6+input_length-25})"
            templ_ws[f"{excelcol}{44+6+input_length-25}"].value = f"={excelcol}{41+6+input_length-25}-{excelcol}{43+6+input_length-25}"

            excelcol = colname_to_excelcol['Diff in Value (Base)']
            templ_ws[f"{excelcol}{38+6+input_length-25}"].value = f"=SUM({excelcol}{14+4}:{excelcol}{36+6+input_length-25})"

            excelcol = colname_to_excelcol['Price per client']
            templ_ws[f"{excelcol}{38+6+input_length-25}"].value = f"=SUM({excelcol}{14+4}:{excelcol}{36+6+input_length-25})"

            excelcol = colname_to_excelcol['Price at Bid']
            templ_ws[f"{excelcol}{38+6+input_length-25}"].value = f"=SUM({excelcol}{14+4}:{excelcol}{36+6+input_length-25})"

            excelcol = colname_to_excelcol['Price at Ask']
            templ_ws[f"{excelcol}{38+6+input_length-25}"].value = f"=SUM({excelcol}{14+4}:{excelcol}{36+6+input_length-25})"

            excelcol = colname_to_excelcol['Max difference']
            templ_ws[f"{excelcol}{38+6+input_length-25}"].value = f"=SUM({excelcol}{14+4}:{excelcol}{36+6+input_length-25})"



        # self._adjust_col_width(templ_ws)

        templ_wb.save(self.output_fp)
        templ_wb.close()

    def write_processed_broker(self):

        sheet_name = "Broker invmt txn listing"

        templ_wb = openpyxl.load_workbook(self.output_fp)
        templ_ws = templ_wb[sheet_name]

        cols_to_drop = InvmtOutputFormatter.DATABASE_MISC_COLS.copy()
        content_df = self.broker_df.copy()

        content_df = content_df.drop(cols_to_drop, axis = 1)

        for c_idx, header_value in enumerate(content_df.columns, 1):
            templ_ws.cell(row=1, column=c_idx, value=header_value)
            col = openpyxl.utils.cell.get_column_letter(c_idx)
            self._format_header_cell([col], 1, templ_ws)

        lst_of_number_cols = self._get_column_lst_letters_by_dtype(content_df, ["float64"])
        lst_of_date_cols = self._get_column_lst_letters_by_dtype(content_df, ["datetime64[ns]"])

        for r_idx, row in enumerate(content_df.values, 2):
            for c_idx, value in enumerate(row, 1):
                templ_ws.cell(row=r_idx, column=c_idx, value=value)
            self._standardise_number_format(templ_ws, lst_of_number_cols, r_idx)
            self._standardise_date_format(templ_ws, lst_of_date_cols, r_idx)
        
        self._create_header(templ_ws, sheet_name, 0, 1)

        # self._adjust_col_width(templ_ws)

        templ_wb.save(self.output_fp)
        templ_wb.close()

    def write_processed_custodian_confirmation(self):

        sheet_name = "Custodian confirmation"

        templ_wb = openpyxl.load_workbook(self.output_fp)
        templ_ws = templ_wb[sheet_name]

        cols_to_drop = InvmtOutputFormatter.DATABASE_MISC_COLS.copy()
        cols_to_drop.append("COMPLETEDFY")

        content_df = self.custodian_confirmation_df.copy()

        content_df = content_df.drop(cols_to_drop, axis = 1)

        lst_of_number_cols = self._get_column_lst_letters_by_dtype(content_df, ["float64"])

        for c_idx, header_value in enumerate(content_df.columns, 1):
            templ_ws.cell(row=1, column=c_idx, value=header_value)
            col = openpyxl.utils.cell.get_column_letter(c_idx)
            self._format_header_cell([col], 1, templ_ws)

        for r_idx, row in enumerate(content_df.values, 2):
            for c_idx, value in enumerate(row, 1):
                templ_ws.cell(row=r_idx, column=c_idx, value=value)
            self._standardise_number_format(templ_ws, lst_of_number_cols, r_idx)
                
        self._create_header(templ_ws, sheet_name, 0, 1)

        # self._adjust_col_width(templ_ws)

        templ_wb.save(self.output_fp)
        templ_wb.close()

        col_adjust = excellib.width_methods.ColumnsWidthAdjuster(self.output_fp)
        col_adjust.main(sheetnames = True)

        templ_wb = openpyxl.load_workbook(self.output_fp)
        templ_ws = templ_wb["<5100-xx>Investment sub-lead"]
        templ_ws.column_dimensions["A"].hidden = True
        templ_ws = templ_wb["<5100-xx>Investment Portfolio"]
        templ_ws.column_dimensions["M"].hidden = True

        # Making sublead sheet default on open
        self.make_sheet_active(templ_wb,"<5100-xx>Investment sub-lead")
        
        templ_wb.save(self.output_fp)
        templ_wb.close()


"""     def write_readme(self, readme_table):

        sheet_name = "README"
        #creating temp wbws
        templ_wb = openpyxl.load_workbook(self.output_fp)
        templ_ws = templ_wb[sheet_name]

        #creating content less db specific content 
        #TODO: custodian_confirmation_df TO CHANGE)
        cols_to_drop = InvmtOutputFormatter.DATABASE_MISC_COLS.copy()
        content_df = readme_table.copy() 

        lst_of_number_cols = self._get_column_lst_letters_by_dtype(content_df, ["float64"])

        for c_idx, header_value in enumerate(content_df.columns, 1):
            templ_ws.cell(row=1, column=c_idx, value=header_value)
            col = openpyxl.utils.cell.get_column_letter(c_idx)
            self._format_header_cell([col], 1, templ_ws)

        for r_idx, row in enumerate(content_df.values, 2):
            for c_idx, value in enumerate(row, 1):
                templ_ws.cell(row=r_idx, column=c_idx, value=value)
            self._standardise_number_format(templ_ws, lst_of_number_cols, r_idx)
                
        self._create_header(templ_ws, sheet_name, 0, 1)

        # self._adjust_col_width(templ_ws)

        templ_wb.save(self.output_fp)
        templ_wb.close()

        col_adjust = excellib.width_methods.ColumnsWidthAdjuster(self.output_fp)
        col_adjust.main(sheetnames = True)
   
 """
if __name__ == "__main__":

    client_no = 50060
    fy        = 2023

    # recon_input_fp = r"D:\Documents\Project\Internal Projects\20240122 FS Funds\Recon output.xlsx"
    output_fp = r"D:\andsoncaimc\Desktop\Task6-App2Exceloutput\luna\personal_workspace\db\funds_test.xlsx"
    portfolio_mapper_fp = r"D:\andsoncaimc\Desktop\Task6-App2Exceloutput\luna\parameters\invmt_portfolio_mapper.xlsx"

    client_class    = tables.client.ClientInfoLoader_From_LunaHub(client_no)
    sublead_class   = tables.fs_funds_invmt_output_sublead.FundsSublead_DownloaderFromLunaHub(client_no, fy)
    portfolio_class = tables.fs_funds_invmt_output_portfolio.FundsPortfolio_DownloaderFromLunaHub(client_no, fy)
    recon_class     = tables.fs_funds_invmt_txn_recon_details.FundsInvmtTxnReconDetail_DownloaderFromLunaHub(client_no, fy)
    broker_class    = tables.fs_funds_broker_statement.FundsBrokerStatement_DownloaderFromLunaHub(client_no, fy)
    custodian_class = tables.fs_funds_custodian_confirmation.FundsCustodianConfirmation_DownloaderFromLunaHub(client_no, fy)
    processedtransaction_class = tables.fs_funds_fundadmin_txn.FundsFundAdminTxn_DownloaderFromLunaHub(client_no, fy)
    processedportfolio_class = tables.fs_funds_fundadmin_portfolio.FundsFundAdminPortfolio_DownloaderFromLunaHub(client_no, fy)
    tb_class        = common.TBLoader_From_LunaHub(client_no, fy)

    aic_name = "DS Team"

    for attempt in range(12):
        user_response_class = tables.fs_funds_userresponse.FundsUserResponse_DownloaderFromLunaHub(
            client_no,
            fy)
        user_inputs = user_response_class.main()
        if user_inputs is not None:
            break
        elif user_inputs is None and attempt == 11:
            raise Exception (f"Data not found for specified client {client_no} or FY {fy}.")
        else:
            continue


    self = InvmtOutputFormatter(sublead_class       = sublead_class,
                                 portfolio_class    = portfolio_class,
                                 recon_class        = recon_class,
                                 broker_class       = broker_class,
                                 custodian_class    = custodian_class,
                                 processedtransaction_class = processedtransaction_class,
                                 processedportfolio_class = processedportfolio_class,
                                 tb_class           = tb_class,
                                 output_fp          = output_fp,
                                 mapper_fp          = portfolio_mapper_fp,
                                 user_inputs        = user_inputs,
                                 client_class       = client_class,
                                 fy                 = fy,
                                 aic_name           = aic_name
                                 )


    if False:

        import webbrowser
        webbrowser.open(output_fp)